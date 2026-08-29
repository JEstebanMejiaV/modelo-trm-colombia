from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from trm_model.data.registry import load_source_registry
from trm_model.features.monthly_transforms import design_term_name
from trm_model.monthly.explanation import NO_CAUSAL_WARNING
from trm_model.monthly.specifications import (
    INTEGRATED_FACTOR_SPECS_4,
    REFERENCE_FACTOR_SPECS,
)
from trm_model.paths import project_paths


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
DATA = ROOT / "data"
VINTAGES = DATA / "vintages"
EXCEL = ROOT / "deliverables" / "modelo_trm_colombia.xlsx"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def assert_close(actual: object, expected: float, label: str, atol: float = 1e-8) -> None:
    if actual is None or not np.isclose(
        float(actual), expected, rtol=0.0, atol=atol
    ):
        raise AssertionError(f"{label}: Excel={actual!r}, CSV={expected!r}.")


def _assert_array_close(
    actual: object, expected: object, label: str, atol: float = 1e-10
) -> None:
    actual_values = pd.to_numeric(pd.Series(actual), errors="coerce").to_numpy(dtype=float)
    expected_values = pd.to_numeric(pd.Series(expected), errors="coerce").to_numpy(dtype=float)
    if (
        actual_values.shape != expected_values.shape
        or not np.isfinite(actual_values).all()
        or not np.isfinite(expected_values).all()
        or not np.allclose(actual_values, expected_values, rtol=0.0, atol=atol)
    ):
        maximum = np.nanmax(np.abs(actual_values - expected_values)) if actual_values.size else np.nan
        raise AssertionError(f"{label}: diferencia máxima={maximum!r}.")


def _assert_value_close(
    actual: object, expected: object, label: str, atol: float = 1e-8
) -> None:
    if pd.isna(expected):
        if actual is not None and not pd.isna(actual):
            raise AssertionError(f"{label}: se esperaba una celda vacía, hay {actual!r}.")
        return
    assert_close(actual, float(expected), label, atol=atol)


def _is_true(value: object) -> bool:
    return str(value).strip().lower() == "true"


def _factor_terms(factor_specs: dict[str, dict[str, object]]) -> dict[str, list[str]]:
    return {
        factor: [design_term_name(component, lag) for component, lag in spec["terminos"]]
        for factor, spec in factor_specs.items()
    }


def _parse_dates(frame: pd.DataFrame, label: str) -> pd.Series:
    if "fecha" not in frame.columns:
        raise AssertionError(f"{label} no contiene la columna fecha.")
    dates = pd.to_datetime(frame["fecha"], errors="coerce")
    if dates.isna().any():
        raise AssertionError(f"{label} contiene fechas inválidas.")
    if len(frame) != 240:
        raise AssertionError(f"{label} debe cubrir exactamente 240 meses; hay {len(frame)}.")
    if not dates.is_unique or not dates.is_monotonic_increasing:
        raise AssertionError(f"{label} debe tener fechas únicas y ordenadas.")
    return dates.reset_index(drop=True)


def _numeric_frame(
    frame: pd.DataFrame, columns: list[str], label: str
) -> pd.DataFrame:
    numeric = frame.loc[:, columns].apply(pd.to_numeric, errors="coerce")
    if numeric.isna().any().any() or not np.isfinite(numeric.to_numpy(dtype=float)).all():
        raise AssertionError(f"{label} contiene valores no numéricos, faltantes o no finitos.")
    return numeric


def _validate_factor_contribution_csv(
    relative_path: str,
    source_relative_path: str,
    factor_specs: dict[str, dict[str, object]],
) -> pd.DataFrame:
    path = RESULTS / relative_path
    source_path = RESULTS / source_relative_path
    if not path.exists():
        raise AssertionError(f"Falta la salida explicativa {relative_path}.")
    if not source_path.exists():
        raise AssertionError(f"Falta la fuente de conciliación {source_relative_path}.")

    frame = pd.read_csv(path)
    source = pd.read_csv(source_path)
    terms_by_factor = _factor_terms(factor_specs)
    factor_columns = list(terms_by_factor)
    expected_columns = [
        "fecha",
        *factor_columns,
        "otros_componentes",
        "suma_factores",
        "ajuste_total",
        "cierre_contable",
    ]
    if list(frame.columns) != expected_columns:
        raise AssertionError(
            f"{relative_path} no concilia con sus columnas esperadas: "
            f"{list(frame.columns)}."
        )

    dates = _parse_dates(frame, relative_path)
    source_dates = _parse_dates(source, source_relative_path)
    if not np.array_equal(dates.to_numpy(), source_dates.to_numpy()):
        raise AssertionError(f"Las fechas de {relative_path} no concilian con {source_relative_path}.")

    source_terms = [term for terms in terms_by_factor.values() for term in terms]
    missing_terms = sorted(set(source_terms).difference(source.columns))
    if missing_terms:
        raise AssertionError(
            f"{source_relative_path} no contiene términos necesarios: {missing_terms}."
        )
    numeric_columns = [*factor_columns, "otros_componentes", "suma_factores", "ajuste_total", "cierre_contable"]
    values = _numeric_frame(frame, numeric_columns, relative_path)
    source_numeric = _numeric_frame(
        source,
        [*source_terms, "ajuste_total"],
        source_relative_path,
    )

    expected_sum = values[factor_columns].sum(axis=1)
    _assert_array_close(values["suma_factores"], expected_sum, f"{relative_path}: suma_factores")
    other_terms = [
        column
        for column in source.columns
        if column not in {"fecha", "ajuste_total", *source_terms}
    ]
    expected_other = (
        source.loc[:, other_terms]
        .apply(pd.to_numeric, errors="coerce")
        .sum(axis=1)
        if other_terms
        else pd.Series(0.0, index=source.index)
    )
    if pd.to_numeric(expected_other, errors="coerce").isna().any():
        raise AssertionError(f"{source_relative_path} contiene otros componentes no numéricos.")
    _assert_array_close(values["otros_componentes"], expected_other, f"{relative_path}: otros_componentes")
    _assert_array_close(
        values["ajuste_total"],
        source_numeric["ajuste_total"],
        f"{relative_path}: ajuste_total",
    )
    _assert_array_close(
        values["suma_factores"] + values["otros_componentes"],
        values["ajuste_total"],
        f"{relative_path}: suma de factores más otros componentes",
    )
    closure = values["cierre_contable"].abs()
    if float(closure.max()) > 1e-10:
        raise AssertionError(
            f"{relative_path} no cierra contablemente; máximo={float(closure.max())}."
        )
    return frame


_INTERPRETATION_COLUMNS = [
    "modelo_id",
    "modelo",
    "factor",
    "grupo",
    "dominio",
    "descripcion",
    "pregunta_guia",
    "canal_descriptivo",
    "terminos",
    "terminos_legibles",
    "rezagos_modelo",
    "n_terminos",
    "es_compuesto",
    "coeficiente",
    "coeficientes_terminos",
    "signos_terminos",
    "p_valor_hac",
    "ic_95_inferior",
    "ic_95_superior",
    "ic95_cruza_cero",
    "p_valor_menor_05",
    "estado_inferencia",
    "contribucion_media_mensual_pct",
    "contribucion_mediana_mensual_pct",
    "contribucion_abs_media_mensual_pct",
    "meses_contribucion_positiva_pct",
    "ultima_contribucion_mensual_pct",
    "participacion_shapley_r2_pct",
    "aporte_shapley_r2_pp",
    "ic95_shapley_inferior_pct",
    "ic95_shapley_superior_pct",
    "probabilidad_shapley_top3_pct",
    "advertencia_interpretacion",
    "estabilidad_signo_factor_pct_submuestras",
    "estabilidad_signos_terminos_pct_submuestras",
    "estabilidad_signos_terminos_pct_2020_en_adelante",
    "peso_entre_factores_pct_2020_en_adelante",
    "lectura_dinamica",
]

_STABILITY_DETAIL_COLUMNS = [
    "submuestra",
    "inicio",
    "fin",
    "observaciones",
    "r2",
    "r2_ajustado",
    "factor",
    "grupo",
    "coeficiente",
    "p_valor_hac",
    "coeficientes_terminos",
    "p_valores_terminos",
    "terminos_evaluados",
    "terminos_con_signo_coincidente",
    "signos_terminos",
    "todos_los_terminos_mismo_signo",
    "shapley_r2",
    "peso_entre_factores_pct",
    "rango_peso",
    "signo_coincide_muestra_completa",
    "diferencia_peso_vs_completa_pp",
]


def _validate_factor_explanation_outputs(
    weights: pd.DataFrame,
    bootstrap: pd.DataFrame,
    stability_detail: pd.DataFrame,
    stability_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    expected_reference_factors = {
        "Términos de intercambio",
        "Remesas",
        "Diferencial de tasas",
        "Déficit fiscal",
        "Dólar amplio",
        "VIX",
    }
    expected_integrated_factors = {
        *expected_reference_factors,
        "Riesgo soberano EMBIG Colombia",
        "Reservas internacionales",
        "Balanza comercial cambiaria",
        "Flujos netos de capital",
        "Diferencial de compensación inflacionaria 5 años",
        "Actividad y precios domésticos",
        "Monedas regionales",
        "Condiciones financieras, commodities y actividad internacional",
    }
    if set(REFERENCE_FACTOR_SPECS) != expected_reference_factors:
        raise AssertionError("La especificación de controles externos cambió sus factores esperados.")
    if set(INTEGRATED_FACTOR_SPECS_4) != expected_integrated_factors:
        raise AssertionError("La especificación integrada cambió sus 14 factores esperados.")

    _validate_factor_contribution_csv(
        "explicacion/contribuciones_factores_controles_externos.csv",
        "explicacion/contribuciones_controles_externos.csv",
        REFERENCE_FACTOR_SPECS,
    )
    integrated_contributions = _validate_factor_contribution_csv(
        "explicacion/contribuciones_factores_marco_macro_integral.csv",
        "explicacion/contribuciones_marco_macro_integral.csv",
        INTEGRATED_FACTOR_SPECS_4,
    )

    interpretation_path = RESULTS / "explicacion/interpretacion_factores_marco_macro_integral.csv"
    if not interpretation_path.exists():
        raise AssertionError("Falta la ficha de interpretación por factor.")
    interpretation = pd.read_csv(interpretation_path)
    if list(interpretation.columns) != _INTERPRETATION_COLUMNS:
        raise AssertionError("La ficha de interpretación no concilia con su contrato de columnas.")
    if len(interpretation) != 14 or interpretation["factor"].duplicated().any():
        raise AssertionError("La ficha de interpretación debe contener 14 factores únicos.")
    if set(interpretation["factor"]) != expected_integrated_factors:
        raise AssertionError("La ficha de interpretación no cubre los 14 factores integrados.")

    terms_by_factor = _factor_terms(INTEGRATED_FACTOR_SPECS_4)
    interpretation_by_factor = interpretation.set_index("factor")
    weights_by_factor = weights.set_index("factor")
    bootstrap_by_factor = bootstrap.set_index("factor")
    if set(weights_by_factor.index) != expected_integrated_factors:
        raise AssertionError("El CSV Shapley no concilia con los 14 factores de la ficha.")
    if set(bootstrap_by_factor.index) != expected_integrated_factors:
        raise AssertionError("El bootstrap Shapley no concilia con los 14 factores de la ficha.")

    for factor, terms in terms_by_factor.items():
        row = interpretation_by_factor.loc[factor]
        weight = weights_by_factor.loc[factor]
        bootstrap_row = bootstrap_by_factor.loc[factor]
        is_composite = len(terms) > 1
        if _is_true(row["es_compuesto"]) != is_composite:
            raise AssertionError(f"La ficha no marca correctamente el carácter compuesto de {factor}.")
        if int(row["n_terminos"]) != len(terms):
            raise AssertionError(f"La ficha no concilia el número de términos de {factor}.")
        if row["terminos"] != ", ".join(terms):
            raise AssertionError(f"La ficha no concilia los términos de {factor}.")
        if is_composite:
            for column in ["coeficiente", "p_valor_hac", "ic_95_inferior", "ic_95_superior"]:
                if not pd.isna(row[column]):
                    raise AssertionError(f"El factor compuesto {factor} no debe tener {column} agregado.")
            if "sin coeficiente único" not in str(row["estado_inferencia"]).lower():
                raise AssertionError(f"Falta la advertencia de coeficiente único para {factor}.")
            if "no tiene un coeficiente ni un signo único" not in str(row["lectura_dinamica"]).lower():
                raise AssertionError(f"La lectura dinámica de {factor} no trata el bloque como compuesto.")
        else:
            for column in ["coeficiente", "p_valor_hac", "ic_95_inferior", "ic_95_superior"]:
                if pd.isna(row[column]) or not np.isfinite(float(row[column])):
                    raise AssertionError(f"El factor simple {factor} tiene {column} no finito.")
        if row["advertencia_interpretacion"] != NO_CAUSAL_WARNING:
            raise AssertionError(f"La ficha de {factor} no conserva la advertencia no causal.")
        dynamic_reading = str(row["lectura_dinamica"])
        if not dynamic_reading.strip() or NO_CAUSAL_WARNING not in dynamic_reading:
            raise AssertionError(f"La lectura dinámica de {factor} no contiene la advertencia no causal.")
        _assert_value_close(
            row["participacion_shapley_r2_pct"],
            weight["peso_entre_factores_pct"],
            f"Participación Shapley de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            row["aporte_shapley_r2_pp"],
            weight["aporte_r2_puntos_porcentuales"],
            f"Aporte Shapley de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            row["ic95_shapley_inferior_pct"],
            bootstrap_row["ic_95_inferior_pct"],
            f"IC Shapley inferior de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            row["ic95_shapley_superior_pct"],
            bootstrap_row["ic_95_superior_pct"],
            f"IC Shapley superior de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            row["probabilidad_shapley_top3_pct"],
            bootstrap_row["probabilidad_top3_pct"],
            f"Probabilidad top 3 de {factor}",
            atol=1e-10,
        )

    if list(stability_detail.columns) != _STABILITY_DETAIL_COLUMNS:
        raise AssertionError("La estabilidad detallada no contiene las columnas por término esperadas.")
    expected_subsamples = {
        "Muestra completa",
        "Primera mitad",
        "Segunda mitad",
        "Prepandemia",
        "2020 en adelante",
    }
    if set(stability_detail["submuestra"]) != expected_subsamples:
        raise AssertionError("La estabilidad detallada no cubre los cinco cortes esperados.")
    if len(stability_detail) != 70 or not stability_detail.groupby(["submuestra", "factor"]).size().eq(1).all():
        raise AssertionError("La estabilidad detallada debe tener una fila por factor y submuestra.")
    if not {"terminos_mismo_signo", "terminos_evaluados", "factores_mismo_signo_de_14"}.issubset(stability_summary.columns):
        raise AssertionError("El resumen de estabilidad no conserva los conteos por término.")
    if not pd.to_numeric(stability_summary["terminos_evaluados"], errors="coerce").eq(31).all():
        raise AssertionError("Cada corte de estabilidad debe evaluar 31 términos.")

    for _, record in stability_detail.iterrows():
        factor = str(record["factor"])
        terms = terms_by_factor[factor]
        evaluated = int(record["terminos_evaluados"])
        matching = int(record["terminos_con_signo_coincidente"])
        if evaluated != len(terms) or not 0 <= matching <= evaluated:
            raise AssertionError(f"El conteo de términos de estabilidad no concilia para {factor}.")
        if _is_true(record["todos_los_terminos_mismo_signo"]) != (matching == evaluated):
            raise AssertionError(f"El indicador de signo por términos no concilia para {factor}.")
        for column in ["coeficientes_terminos", "p_valores_terminos", "signos_terminos"]:
            text = str(record[column])
            if not text or text.lower() == "nan" or not all(f"{term}=" in text for term in terms):
                raise AssertionError(f"{column} no enumera todos los términos de {factor}.")
        if len(terms) > 1:
            if not pd.isna(record["coeficiente"]) or not pd.isna(record["p_valor_hac"]):
                raise AssertionError(f"El factor compuesto {factor} recibió un coeficiente único en estabilidad.")
        elif pd.isna(record["coeficiente"]) or pd.isna(record["p_valor_hac"]):
            raise AssertionError(f"El factor simple {factor} perdió su inferencia en estabilidad.")

    summary_by_subsample = stability_summary.set_index("submuestra")
    for subsample in expected_subsamples:
        detail_rows = stability_detail.loc[stability_detail["submuestra"].eq(subsample)]
        summary_row = summary_by_subsample.loc[subsample]
        matching = int(detail_rows["terminos_con_signo_coincidente"].sum())
        evaluated = int(detail_rows["terminos_evaluados"].sum())
        factors_matching = int(detail_rows["todos_los_terminos_mismo_signo"].map(_is_true).sum())
        if int(summary_row["terminos_mismo_signo"]) != matching:
            raise AssertionError(f"El resumen no concilia términos coincidentes en {subsample}.")
        if int(summary_row["terminos_evaluados"]) != evaluated:
            raise AssertionError(f"El resumen no concilia términos evaluados en {subsample}.")
        if int(summary_row["factores_mismo_signo_de_14"]) != factors_matching:
            raise AssertionError(f"El resumen no concilia factores con signo coincidente en {subsample}.")

    full_detail = stability_detail.loc[stability_detail["submuestra"].eq("Muestra completa")]
    if not full_detail["signo_coincide_muestra_completa"].map(_is_true).all() or not np.allclose(
        pd.to_numeric(full_detail["diferencia_peso_vs_completa_pp"], errors="coerce"),
        0.0,
        rtol=0.0,
        atol=1e-10,
    ):
        raise AssertionError("La muestra completa no es la referencia coherente de estabilidad.")

    for factor in terms_by_factor:
        factor_rows = stability_detail.loc[stability_detail["factor"].eq(factor)]
        non_full = factor_rows.loc[factor_rows["submuestra"].ne("Muestra completa")]
        recent = factor_rows.loc[factor_rows["submuestra"].eq("2020 en adelante")]
        expected_factor_share = 100.0 * non_full["signo_coincide_muestra_completa"].map(_is_true).mean()
        expected_term_share = 100.0 * (
            pd.to_numeric(non_full["terminos_con_signo_coincidente"], errors="coerce").sum()
            / pd.to_numeric(non_full["terminos_evaluados"], errors="coerce").sum()
        )
        recent_row = recent.iloc[0]
        expected_recent_term_share = 100.0 * float(recent_row["terminos_con_signo_coincidente"]) / float(recent_row["terminos_evaluados"])
        interpretation_row = interpretation_by_factor.loc[factor]
        _assert_value_close(
            interpretation_row["estabilidad_signo_factor_pct_submuestras"],
            expected_factor_share,
            f"Estabilidad de signo de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            interpretation_row["estabilidad_signos_terminos_pct_submuestras"],
            expected_term_share,
            f"Estabilidad por término de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            interpretation_row["estabilidad_signos_terminos_pct_2020_en_adelante"],
            expected_recent_term_share,
            f"Estabilidad reciente de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            interpretation_row["peso_entre_factores_pct_2020_en_adelante"],
            recent_row["peso_entre_factores_pct"],
            f"Peso reciente de {factor}",
            atol=1e-10,
        )

    return interpretation, integrated_contributions


def _table_bounds(ws, name: str, headers: list[str], data_rows: int) -> tuple[int, int, int, int]:
    try:
        table = ws.tables[name]
    except KeyError as error:
        raise AssertionError(f"No se encontró la tabla Excel {name}.") from error
    if not hasattr(table, "ref"):
        raise AssertionError(f"La tabla Excel {name} no tiene un rango válido.")
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    actual_headers = [ws.cell(min_row, column).value for column in range(min_col, max_col + 1)]
    if actual_headers != headers:
        raise AssertionError(f"Los encabezados de {name} no concilian: {actual_headers!r}.")
    if max_row != min_row + data_rows or max_col - min_col + 1 != len(headers):
        raise AssertionError(f"El rango de {name} no concilia con sus filas/columnas esperadas.")
    return min_col, min_row, max_col, max_row


def _validate_factor_workbook(
    workbook,
    interpretation: pd.DataFrame,
    factor_contributions: pd.DataFrame,
) -> None:
    summary = workbook["Resumen"]
    marco = workbook["Marco_macro_integral"]
    if "Ninguna tabla identifica efectos causales" not in str(summary["A3"].value):
        raise AssertionError("Resumen no contiene la advertencia explícita de no causalidad.")
    if "no son efectos causales" not in str(summary["A51"].value).lower():
        raise AssertionError("La ficha del Resumen no contiene su advertencia no causal.")
    if "no son efectos causales" not in str(marco["A2"].value).lower():
        raise AssertionError("Marco_macro_integral no contiene su advertencia no causal.")
    if "no causalidad" not in str(marco["A284"].value).lower():
        raise AssertionError("La tabla de contribuciones no contiene su advertencia no causal.")

    ficha_headers = [
        "Factor",
        "Grupo",
        "Tipo de lectura",
        "Coeficiente / signo",
        "IC 95% HAC",
        "Participación R² incremental",
        "Contribución media Δln TRM",
        "Lectura dinámica",
    ]
    min_col, min_row, _, _ = _table_bounds(
        summary,
        "FichaFactoresResumenTable",
        ficha_headers,
        len(interpretation),
    )
    interpretation_by_factor = interpretation.set_index("factor")
    excel_factors: set[str] = set()
    for offset in range(1, len(interpretation) + 1):
        row_number = min_row + offset
        factor = str(summary.cell(row_number, min_col).value)
        if factor in excel_factors:
            raise AssertionError(f"La ficha Excel repite el factor {factor}.")
        excel_factors.add(factor)
        if factor not in interpretation_by_factor.index:
            raise AssertionError(f"La ficha Excel contiene un factor inesperado: {factor}.")
        expected = interpretation_by_factor.loc[factor]
        composite = _is_true(expected["es_compuesto"])
        if str(summary.cell(row_number, min_col + 1).value) != str(expected["grupo"]):
            raise AssertionError(f"El grupo Excel no concilia para {factor}.")
        type_text = str(summary.cell(row_number, min_col + 2).value)
        coefficient_text = summary.cell(row_number, min_col + 3).value
        interval_text = summary.cell(row_number, min_col + 4).value
        if composite:
            if type_text != "Bloque compuesto" or coefficient_text != "Sin coeficiente único" or interval_text != "No aplica":
                raise AssertionError(f"La ficha Excel asigna un coeficiente único al bloque {factor}.")
        else:
            if type_text != "Coeficiente único" or coefficient_text in {None, "", "Sin coeficiente único"}:
                raise AssertionError(f"La ficha Excel no muestra el coeficiente simple de {factor}.")
            assert_close(
                coefficient_text,
                float(expected["coeficiente"]),
                f"Coeficiente Excel de {factor}",
                atol=1e-5,
            )
        _assert_value_close(
            summary.cell(row_number, min_col + 5).value,
            float(expected["participacion_shapley_r2_pct"]) / 100,
            f"Participación Excel de {factor}",
            atol=1e-10,
        )
        _assert_value_close(
            summary.cell(row_number, min_col + 6).value,
            float(expected["contribucion_media_mensual_pct"]) / 100,
            f"Contribución media Excel de {factor}",
            atol=1e-10,
        )
        reading = str(summary.cell(row_number, min_col + 7).value)
        if NO_CAUSAL_WARNING not in reading:
            raise AssertionError(f"La lectura Excel de {factor} no conserva la advertencia no causal.")
    if excel_factors != set(interpretation["factor"]):
        raise AssertionError("Los factores de FichaFactoresResumenTable no concilian con el CSV.")

    contribution_columns = [column for column in factor_contributions.columns if column != "fecha"]
    contribution_headers = ["Mes", *contribution_columns]
    min_col, min_row, _, _ = _table_bounds(
        marco,
        "ContribucionesFactoresMarcoMacroIntegralTable",
        contribution_headers,
        len(factor_contributions),
    )
    for offset, (_, record) in enumerate(factor_contributions.iterrows(), start=1):
        row_number = min_row + offset
        expected_date = pd.to_datetime(record["fecha"]).strftime("%Y-%m")
        actual_date = pd.to_datetime(marco.cell(row_number, min_col).value, errors="coerce")
        if pd.isna(actual_date) or actual_date.strftime("%Y-%m") != expected_date:
            raise AssertionError(f"La fecha Excel de contribuciones no concilia en la fila {row_number}.")
        for column_offset, column in enumerate(contribution_columns, start=1):
            expected_value = record[column]
            _assert_value_close(
                marco.cell(row_number, min_col + column_offset).value,
                expected_value,
                f"Contribución Excel {column} fila {row_number}",
                atol=1e-8,
            )


def main() -> None:
    metadata = json.loads((RESULTS / "metadata.json").read_text(encoding="utf-8"))
    for filename, expected_hash in metadata["proxies_snapshot_sha256"].items():
        actual_hash = sha256_file(DATA / "raw" / filename)
        if actual_hash != expected_hash:
            raise AssertionError(
                f"La instantánea raw {filename} no coincide con su SHA-256 registrado."
            )
    for filename, expected_hash in metadata["internal_snapshot_sha256"].items():
        actual_hash = sha256_file(DATA / "raw" / filename)
        if actual_hash != expected_hash:
            raise AssertionError(
                f"La instantánea interna {filename} no coincide con su SHA-256 registrado."
            )

    weights = pd.read_csv(RESULTS / "explicacion/pesos_explicativos_marco_macro_integral.csv")
    coverage_path = DATA / "base_global_cobertura.csv"
    if not coverage_path.exists():
        raise AssertionError("Falta el registro de cobertura de variables globales.")
    coverage = pd.read_csv(coverage_path).set_index("variable")
    internal_coverage_path = DATA / "variables_internas_cobertura.csv"
    if not internal_coverage_path.exists():
        raise AssertionError("Falta la matriz de cobertura de variables internas.")
    internal_coverage = pd.read_csv(internal_coverage_path).set_index("variable")
    for variable in ["ise_total_dane", "ipc_colombia_indice"]:
        if variable not in internal_coverage.index:
            raise AssertionError(f"Falta la cobertura documentada de {variable}.")
        record = internal_coverage.loc[variable]
        if record["estado"] != "activa" or not bool(record["cubre_muestra_completa"]):
            raise AssertionError(f"La variable interna activa {variable} no cubre la muestra completa.")
        if int(record["meses_faltantes_muestra"]) != 0:
            raise AssertionError(f"La variable interna activa {variable} tiene faltantes.")
    active_variables = {
        "yield_real_10y_tips_pct",
        "yield_real_5y_us_pct",
        "yield_2y_us_pct",
        "yield_10y_us_pct",
        "spread_10y_2y_us_pct",
        "breakeven_5y_us_pct",
        "breakeven_10y_us_pct",
        "brent_usd_barril",
        "commodities_index_imf",
        "epu_global",
        "estres_financiero_stl",
        "nfci_chicago",
        "anfci_chicago",
        "empleo_manufactura_us_miles",
        "produccion_industrial_us",
        "desempleo_us_pct",
        "fletes_transporte_us",
    }
    if not active_variables.issubset(coverage.index):
        raise AssertionError("El registro de cobertura no contiene todas las series activas.")
    active_coverage = coverage.loc[sorted(active_variables)]
    if not active_coverage["estado"].eq("activa").all() or not active_coverage["cubre_muestra_completa"].all():
        raise AssertionError("Una serie declarada activa no cubre la muestra completa.")
    documented_candidates = {
        "ted_spread_pct",
        "high_yield_oas_pct",
        "desempleo_us_bls_pct",
        "precios_importacion_china",
        "produccion_industrial_china",
        "indicador_lider_china",
        "ipc_china",
    }
    if not documented_candidates.issubset(coverage.index):
        raise AssertionError("Faltan candidatos globales en el registro de cobertura.")
    if {"Global", "Global ampliado"}.intersection(set(weights.get("grupo", []))):
        raise AssertionError("Persisten nombres genéricos de grupos globales.")
    if "Variables globales nuevas" in set(weights["factor"]):
        raise AssertionError("Persistió el nombre no descriptivo del factor global.")
    if len(weights) != 14:
        raise AssertionError("La descomposición debe contener exactamente 14 factores.")
    if (weights["shapley_r2"] < -1e-12).any():
        raise AssertionError("Un aporte Shapley dentro de muestra resultó negativo.")
    if not np.isclose(weights["peso_entre_factores_pct"].sum(), 100.0, atol=1e-8):
        raise AssertionError("Los pesos Shapley no suman 100%.")
    if not np.isclose(
        weights["shapley_r2"].sum(), weights["r2_incremental"].iloc[0], atol=1e-10
    ):
        raise AssertionError("Los aportes Shapley no cierran contra el R² incremental.")

    bootstrap = pd.read_csv(RESULTS / "explicacion/intervalos_bootstrap_pesos_shapley.csv")
    if set(bootstrap["factor"]) != set(weights["factor"]) or len(bootstrap) != 14:
        raise AssertionError("Los intervalos bootstrap no cubren los 14 factores Shapley.")
    if not bootstrap["replicas_validas"].eq(200).all():
        raise AssertionError("Los intervalos Shapley deben usar 200 réplicas válidas.")
    if not bootstrap["bloque_meses"].eq(12).all():
        raise AssertionError("El bootstrap Shapley debe conservar bloques de 12 meses.")
    if not bootstrap["probabilidad_top3_pct"].between(0, 100).all():
        raise AssertionError("Una probabilidad top 3 quedó fuera de [0, 100].")
    if not np.isclose(bootstrap["probabilidad_top3_pct"].sum(), 300.0, atol=1e-8):
        raise AssertionError("Las probabilidades top 3 no suman tres posiciones.")
    bootstrap_by_factor = bootstrap.set_index("factor")
    weights_by_factor = weights.set_index("factor")
    if not np.allclose(
        bootstrap_by_factor.loc[weights_by_factor.index, "peso_puntual_pct"],
        weights_by_factor["peso_entre_factores_pct"],
        rtol=0.0,
        atol=1e-10,
    ):
        raise AssertionError("Los pesos puntuales del bootstrap no concilian con Shapley exacto.")

    stability_detail = pd.read_csv(
        RESULTS / "explicacion/estabilidad_submuestras_marco_macro_integral.csv"
    )
    stability_summary = pd.read_csv(RESULTS / "explicacion/estabilidad_submuestras_resumen.csv")
    expected_subsamples = {
        "Muestra completa",
        "Primera mitad",
        "Segunda mitad",
        "Prepandemia",
        "2020 en adelante",
    }
    if set(stability_summary["submuestra"]) != expected_subsamples:
        raise AssertionError("Faltan cortes de estabilidad en el resumen.")
    if len(stability_detail) != 70 or not stability_detail.groupby("submuestra").size().eq(14).all():
        raise AssertionError("La estabilidad detallada debe tener 14 factores en cinco cortes.")
    if not stability_summary["correlacion_spearman_rangos_vs_completa"].between(-1, 1).all():
        raise AssertionError("Una correlación de rangos de estabilidad quedó fuera de [-1, 1].")

    baseline_manifest = json.loads(
        (VINTAGES / "2026-08-23" / "manifest.json").read_text(encoding="utf-8")
    )
    if baseline_manifest["origin_date"] != "2026-08-23" or not baseline_manifest["immutable"]:
        raise AssertionError("El baseline de vintages no está marcado como inmutable.")
    registry = load_source_registry(paths=project_paths(ROOT))
    expected_sources = {
        (str(source["source_id"]), str(source["raw_path"]))
        for source in registry.active_sources
    }
    baseline_sources = {
        (str(record.get("id") or record.get("source_id")), str(record["raw_path"]))
        for record in baseline_manifest["files"]
    }
    if baseline_sources != expected_sources:
        raise AssertionError(
            "El baseline de vintages no concilia IDs/rutas con las 26 fuentes activas: "
            f"missing={sorted(expected_sources - baseline_sources)}, "
            f"extra={sorted(baseline_sources - expected_sources)}"
        )
    if len(baseline_sources) != len(registry.active_sources):
        raise AssertionError("El baseline de vintages contiene fuentes activas duplicadas.")
    for record in baseline_manifest["files"]:
        path = ROOT / record["raw_path"]
        if sha256_file(path) != record["sha256"]:
            raise AssertionError(
                f"No concilia el baseline de vintages: {record.get('id', record.get('source_id'))}."
            )

    alfred_manifest_path = (
        VINTAGES / "historical" / "alfred_factores_pronostico.manifest.json"
    )
    if alfred_manifest_path.exists():
        alfred_manifest = json.loads(alfred_manifest_path.read_text(encoding="utf-8"))
        alfred_path = ROOT / alfred_manifest["path"]
        if sha256_file(alfred_path) != alfred_manifest["sha256"]:
            raise AssertionError("No concilia la huella del histórico ALFRED.")
        alfred = pd.read_csv(alfred_path)
        if alfred["origen_vintage"].nunique() != 48 or alfred["serie_id"].nunique() != 6:
            raise AssertionError("El histórico ALFRED no cubre 48 orígenes y seis series.")
        alfred["origen_vintage"] = pd.to_datetime(alfred["origen_vintage"])
        alfred["fecha_observacion"] = pd.to_datetime(alfred["fecha_observacion"])
        if (alfred["fecha_observacion"] >= alfred["origen_vintage"]).any():
            raise AssertionError("El histórico ALFRED contiene observaciones futuras al origen.")

    fiscal_history = json.loads(
        (VINTAGES / "historical" / "minhacienda" / "version_history.json").read_text(
            encoding="utf-8"
        )
    )
    if len(fiscal_history["versions"]) != 8:
        raise AssertionError("El catálogo fiscal no contiene las ocho versiones verificadas.")

    vintage_coverage = pd.read_csv(RESULTS / "pronostico/cobertura_vintages_pronostico.csv")
    if len(vintage_coverage) != 14:
        raise AssertionError("La cobertura de vintages debe reportar los 14 factores.")
    from model.config import FORECAST_AVAILABILITY

    expected_vintage_factors = {row[0] for row in FORECAST_AVAILABILITY}
    actual_vintage_factors = set(vintage_coverage["factor"])
    if actual_vintage_factors != expected_vintage_factors:
        raise AssertionError(
            "La cobertura de vintages no concilia con los factores de pronóstico: "
            f"missing={sorted(expected_vintage_factors - actual_vintage_factors)}, "
            f"extra={sorted(actual_vintage_factors - expected_vintage_factors)}"
        )
    complete_vintage = vintage_coverage["apto_backtest_genuino"].astype("string").str.lower().eq("true")
    alfred_csv = DATA / "vintages" / "historical" / "alfred_factores_pronostico.csv"
    expected_complete = 3 if alfred_csv.exists() else 0
    if int(complete_vintage.sum()) != expected_complete:
        raise AssertionError(
            f"Se esperan {expected_complete} factores aptos, hay {int(complete_vintage.sum())}."
        )
    if bool(metadata["backtest_genuino_disponible"]):
        raise AssertionError("El backtest no puede rotularse como genuino con cobertura parcial.")

    comparison = pd.read_csv(RESULTS / "explicacion/comparacion_especificaciones.csv")
    if set(comparison["modelo"]) != {"Controles externos y financieros", "Marco macroeconómico integral"}:
        raise AssertionError("La comparación no contiene las dos agrupaciones descriptivas esperadas.")
    if comparison["observaciones"].nunique() != 1:
        raise AssertionError("Las dos agrupaciones no usan la misma muestra efectiva.")

    expanded_coefficients = pd.read_csv(
        RESULTS / "explicacion/coeficientes_marco_macro_integral.csv"
    )
    expanded_terms = set(expanded_coefficients["termino"])
    required_active_terms = {
        "D.ln_terminos_intercambio.L0",
        "D.embig_colombia_pp.L0",
        "D.asinh_balanza_comercial.L1",
        "D.asinh_flujos_capital.L1",
        "D.diferencial_bei_5y_pp.L1",
        "D.ln_ise_total_dane.L0",
        "D.ln_ipc_colombia.L0",
        "factor_monedas_regionales_4.L0",
        "D.yield_real_10y_tips_pct.L0",
        "D.yield_real_5y_us_pct.L0",
        "D.breakeven_5y_us_pct.L0",
        "D.breakeven_10y_us_pct.L0",
        "D.nfci_chicago.L0",
        "D.anfci_chicago.L0",
        "D.desempleo_us_pct.L0",
        "D.ln_fletes_transporte_us.L0",
        "D.ln_brent_global.L0",
        "D.ln_produccion_industrial_us.L0",
    }
    missing_active_terms = required_active_terms.difference(expanded_terms)
    if missing_active_terms:
        raise AssertionError(
            "Faltan términos activos en el marco macroeconómico integral: "
            f"{sorted(missing_active_terms)}"
        )
    retired_terms = {
        "D.ln_brent.L0",
        "D.spread_tes_ust_10y_pp.L0",
        "diferencial_inflacion_pp.L1",
    }
    unexpected_retired_terms = retired_terms.intersection(expanded_terms)
    if unexpected_retired_terms:
        raise AssertionError(
            "Persisten términos sustituidos en el marco macroeconómico integral: "
            f"{sorted(unexpected_retired_terms)}"
        )

    integration = pd.read_csv(RESULTS / "explicacion/pruebas_integracion.csv")
    for variable in ["asinh_balanza_comercial", "asinh_flujos_capital"]:
        transformations = set(
            integration.loc[integration["variable"].eq(variable), "transformacion"]
        )
        if transformations != {"nivel", "primera_diferencia"}:
            raise AssertionError(
                f"Faltan pruebas de integración completas para {variable}."
            )

    monthly_path = DATA / "modelo_trm_datos_mensuales.csv"
    monthly = pd.read_csv(monthly_path)
    required_columns = {
        "ise_total_dane",
        "ipc_colombia_indice",
        "ln_ise_total_dane",
        "ln_ipc_colombia",
        "terminos_intercambio",
        "ln_terminos_intercambio",
        "embig_colombia_pb",
        "embig_colombia_pp",
        "tes_5y_pesos_colombia_pct",
        "tes_5y_uvr_colombia_pct",
        "bei_colombia_5y_pct",
        "bei_eeuu_5y_pct",
        "diferencial_bei_5y_pp",
        "diferencial_bei_5y_comun_pp",
        "diferencia_comun_menos_separada_pp",
        "dias_comunes",
        "tes_5y_pesos_comun_pct",
        "tes_5y_uvr_comun_pct",
        "bei_eeuu_5y_comun_pct",
        "ln_reservas_netas_sin_flar",
        "asinh_balanza_comercial",
        "asinh_flujos_capital",
        "pen_por_usd",
        "factor_monedas_regionales_3",
        "factor_monedas_regionales_4",
        "factor_monedas_regionales",
        "yield_real_10y_tips_pct",
        "yield_real_5y_us_pct",
        "yield_2y_us_pct",
        "yield_10y_us_pct",
        "spread_10y_2y_us_pct",
        "breakeven_5y_us_pct",
        "breakeven_10y_us_pct",
        "ln_brent_global",
        "ln_commodities_global",
        "epu_global",
        "estres_financiero_stl",
        "nfci_chicago",
        "anfci_chicago",
        "desempleo_us_pct",
        "ln_empleo_manufactura_us",
        "ln_produccion_industrial_us",
        "ln_fletes_transporte_us",
    }
    missing_columns = required_columns.difference(monthly.columns)
    if missing_columns:
        raise AssertionError(f"Faltan variables ampliadas: {sorted(missing_columns)}")

    identities = {
        "ln(términos de intercambio)": (
            monthly["ln_terminos_intercambio"],
            np.log(monthly["terminos_intercambio"]),
        ),
        "EMBIG en puntos porcentuales": (
            monthly["embig_colombia_pp"],
            monthly["embig_colombia_pb"] / 100.0,
        ),
        "BEI Colombia a 5 años": (
            monthly["bei_colombia_5y_pct"],
            monthly["tes_5y_pesos_colombia_pct"]
            - monthly["tes_5y_uvr_colombia_pct"],
        ),
        "diferencial BEI Colombia-EE. UU.": (
            monthly["diferencial_bei_5y_pp"],
            monthly["bei_colombia_5y_pct"] - monthly["bei_eeuu_5y_pct"],
        ),
        "diferencial BEI sobre fechas comunes": (
            monthly["diferencial_bei_5y_comun_pp"],
            monthly["tes_5y_pesos_comun_pct"]
            - monthly["tes_5y_uvr_comun_pct"]
            - monthly["bei_eeuu_5y_comun_pct"],
        ),
        "ln ISE total DANE": (
            monthly["ln_ise_total_dane"],
            np.log(monthly["ise_total_dane"]),
        ),
        "ln IPC Colombia": (
            monthly["ln_ipc_colombia"],
            np.log(monthly["ipc_colombia_indice"]),
        ),
    }
    for label, (actual, expected) in identities.items():
        comparable = actual.notna() & expected.notna()
        if not comparable.any() or not np.allclose(
            actual.loc[comparable], expected.loc[comparable], rtol=0.0, atol=1e-8
        ):
            raise AssertionError(f"No concilia la construcción de {label}.")

    bei_aggregation = pd.read_csv(RESULTS / "robustez/comparacion_agregacion_bei_5y.csv")
    if len(bei_aggregation) != 244 or bei_aggregation["dias_comunes"].isna().any():
        raise AssertionError("La comparación BEI debe cubrir 244 meses con fechas comunes.")
    if bei_aggregation["dias_comunes"].min() < 1:
        raise AssertionError("Existe un mes sin cruce diario para las tres curvas BEI.")
    if bei_aggregation["diferencial_bei_5y_pp"].corr(
        bei_aggregation["diferencial_bei_5y_comun_pp"]
    ) < 0.99:
        raise AssertionError("Las dos agregaciones BEI divergen de forma inesperada.")

    bei_stationarity = pd.read_csv(RESULTS / "robustez/pruebas_estacionariedad_bei_5y.csv")
    if len(bei_stationarity) != 24:
        raise AssertionError("Las pruebas BEI no cubren agregaciones, transformaciones y determinísticos.")
    differenced_tests = bei_stationarity.loc[
        bei_stationarity["transformacion"].eq("primera_diferencia")
    ]
    if not differenced_tests.loc[differenced_tests["prueba"].eq("ADF"), "p_valor"].lt(0.05).all():
        raise AssertionError("ADF no respalda la primera diferencia BEI en todas las variantes.")
    if not differenced_tests.loc[differenced_tests["prueba"].eq("KPSS"), "p_valor"].ge(0.05).all():
        raise AssertionError("KPSS rechaza estacionariedad de una primera diferencia BEI.")

    bei_trends = pd.read_csv(RESULTS / "robustez/tendencias_quiebres_bei_5y.csv")
    if len(bei_trends) != 6 or bei_trends["fecha_quiebre_za"].nunique() != 1:
        raise AssertionError("La comparación de tendencias/quiebres BEI está incompleta.")

    bei_specs = pd.read_csv(RESULTS / "robustez/comparacion_especificaciones_bei_5y.csv")
    if len(bei_specs) != 6 or bei_specs["observaciones"].nunique() != 1:
        raise AssertionError("Las seis especificaciones BEI no usan una muestra común.")
    active_bei = bei_specs.loc[bei_specs["especificacion"].str.contains("vigente")].iloc[0]
    if active_bei["transformacion_bei"] != "primera_diferencia":
        raise AssertionError("La especificación BEI activa debe usar la primera diferencia.")
    # El nivel puede obtener un BIC ligeramente menor, pero la especificación
    # vigente prioriza una transformación estable frente a una comparación
    # puramente mecánica de BIC; metadata.json conserva cuál ganó cada criterio.

    monthly["fecha"] = pd.to_datetime(monthly["fecha"])
    currency_levels = monthly.set_index("fecha")[[
        "brl_por_usd",
        "clp_por_usd",
        "mxn_por_usd",
        "pen_por_usd",
    ]]
    currency_changes = np.log(currency_levels).diff()
    calibration = currency_changes.loc["2006-01-01":"2019-12-01"]
    standardized = (currency_changes - calibration.mean()) / calibration.std(ddof=0)
    expected_factor_3 = standardized[[
        "brl_por_usd", "clp_por_usd", "mxn_por_usd"
    ]].mean(axis=1, skipna=False)
    expected_factor_4 = standardized.mean(axis=1, skipna=False)
    regional_identities = {
        "factor regional de tres monedas": expected_factor_3,
        "factor regional de cuatro monedas": expected_factor_4,
        "alias regional activo": expected_factor_4,
    }
    monthly_indexed = monthly.set_index("fecha")
    for label, expected in regional_identities.items():
        column = {
            "factor regional de tres monedas": "factor_monedas_regionales_3",
            "factor regional de cuatro monedas": "factor_monedas_regionales_4",
            "alias regional activo": "factor_monedas_regionales",
        }[label]
        actual = monthly_indexed[column]
        comparable = actual.notna() & expected.notna()
        if not comparable.any() or not np.allclose(
            actual.loc[comparable], expected.loc[comparable], rtol=0.0, atol=2e-8
        ):
            raise AssertionError(f"No concilia la construcción de {label}.")

    regional_comparison = pd.read_csv(RESULTS / "explicacion/comparacion_factor_regional.csv")
    expected_uses = {
        "Explicación histórica",
        "Pronóstico con rezagos de publicación",
    }
    expected_compositions = {"BRL, CLP y MXN", "BRL, CLP, MXN y PEN"}
    if len(regional_comparison) != 4:
        raise AssertionError("La comparación regional debe contener exactamente cuatro variantes.")
    if set(regional_comparison["uso"]) != expected_uses:
        raise AssertionError("La comparación regional no separa explicación y pronóstico.")
    if set(regional_comparison["monedas"]) != expected_compositions:
        raise AssertionError("La comparación regional no contiene factores de tres y cuatro monedas.")
    if regional_comparison["correlacion_factores_3_4"].nunique() != 1:
        raise AssertionError("La correlación de factores regionales no es consistente.")

    historical_variants = regional_comparison.loc[
        regional_comparison["uso"].eq("Explicación histórica")
    ].set_index("monedas")
    forecast_variants = regional_comparison.loc[
        regional_comparison["uso"].eq("Pronóstico con rezagos de publicación")
    ].set_index("monedas")
    if historical_variants.loc["BRL, CLP, MXN y PEN", "bic"] >= historical_variants.loc[
        "BRL, CLP y MXN", "bic"
    ]:
        raise AssertionError("PEN no mejora el BIC de la explicación histórica.")
    if forecast_variants.loc["BRL, CLP y MXN", "bic"] >= forecast_variants.loc[
        "BRL, CLP, MXN y PEN", "bic"
    ]:
        raise AssertionError("La selección de tres monedas no minimiza el BIC del pronóstico.")

    forecast_coefficients = pd.read_csv(
        RESULTS / "pronostico/coeficientes_modelo_pronostico.csv"
    )
    forecast_terms = set(forecast_coefficients["termino"])
    required_forecast_internal_terms = {
        "D.ln_ise_total_dane.L2",
        "D.ln_ipc_colombia.L2",
    }
    if not required_forecast_internal_terms.issubset(forecast_terms):
        raise AssertionError(
            "Faltan términos internos rezagados en el pronóstico: "
            f"{sorted(required_forecast_internal_terms.difference(forecast_terms))}"
        )
    for term in forecast_terms.difference({"const", "dummy_pandemia_2020"}):
        match = re.search(r"\.L(\d+)$", term)
        if match is None or int(match.group(1)) < 1:
            raise AssertionError(
                f"El pronóstico contiene información contemporánea o sin rezago: {term}."
            )

    forecast_metrics = pd.read_csv(RESULTS / "pronostico/validacion_metricas_pronostico.csv")
    forecast_predictions = pd.read_csv(
        RESULTS / "pronostico/validacion_predicciones_pronostico.csv"
    )
    if set(forecast_metrics["modelo"]) != {
        "Pronóstico con rezagos de publicación",
        "Caminata aleatoria",
    }:
        raise AssertionError("Faltan los dos modelos de la validación de pronóstico.")
    if len(forecast_predictions) != 48 or not forecast_metrics["observaciones"].eq(48).all():
        raise AssertionError("La validación de pronóstico debe cubrir exactamente 48 meses.")

    sample = pd.read_csv(DATA / "modelo_trm_muestra_estimacion.csv")
    required_sample_columns = {
        "ln_terminos_intercambio",
        "ln_ise_total_dane",
        "ln_ipc_colombia",
        "ise_total_dane",
        "ipc_colombia_indice",
        "embig_colombia_pp",
        "diferencial_bei_5y_pp",
        "diferencial_bei_5y_comun_pp",
        "factor_monedas_regionales_3",
        "factor_monedas_regionales_4",
        "yield_real_10y_tips_pct",
        "ln_brent_global",
        "ln_commodities_global",
        "epu_global",
        "estres_financiero_stl",
        "nfci_chicago",
        "anfci_chicago",
        "desempleo_us_pct",
        "ln_empleo_manufactura_us",
        "ln_produccion_industrial_us",
        "ln_fletes_transporte_us",
    }
    missing_sample_columns = required_sample_columns.difference(sample.columns)
    if missing_sample_columns:
        raise AssertionError(
            "Faltan sustituciones activas en la muestra de estimación: "
            f"{sorted(missing_sample_columns)}"
        )

    workbook = load_workbook(EXCEL, data_only=True)
    required_sheets = {
        "Resumen",
        "Datos_fuente",
        "Transformaciones",
        "Controles_externos",
        "Marco_macro_integral",
        "Pesos_explicativos",
        "Robustez",
        "BEI_robustez",
        "Validacion",
        "Pronostico",
        "ECM_exploratorio",
        "Diagnosticos",
        "Variables",
        "Fuentes",
    }
    if workbook.sheetnames != list(required_sheets):
        expected_sheet_order = [
            "Resumen",
            "Datos_fuente",
            "Transformaciones",
            "Controles_externos",
            "Marco_macro_integral",
            "Pesos_explicativos",
            "Robustez",
            "BEI_robustez",
            "Validacion",
            "Pronostico",
            "ECM_exploratorio",
            "Diagnosticos",
            "Variables",
            "Fuentes",
        ]
        if workbook.sheetnames != expected_sheet_order:
            raise AssertionError(f"El archivo Excel debe contener exactamente las 14 hojas descriptivas: {workbook.sheetnames}")

    source_headers = {workbook["Datos_fuente"].cell(5, column).value for column in range(1, workbook["Datos_fuente"].max_column + 1)}
    if not {"ISE total DANE (índice)", "IPC Colombia (índice)"}.issubset(source_headers):
        raise AssertionError("Datos_fuente no contiene ISE total DANE e IPC Colombia.")
    transformation_headers = {workbook["Transformaciones"].cell(5, column).value for column in range(1, workbook["Transformaciones"].max_column + 1)}
    if not {"ln ISE total DANE", "Δln ISE total DANE", "ln IPC Colombia", "Δln IPC Colombia"}.issubset(transformation_headers):
        raise AssertionError("Transformaciones no contiene los logaritmos y diferencias de ISE/IPC.")

    # El CI no reconstruye el archivo Excel. Estas conciliaciones impiden que un
    # cambio de resultados deje versionado un archivo Excel desactualizado.
    weights_sheet = workbook["Pesos_explicativos"]
    header_row = next(
        (
            row
            for row in range(1, weights_sheet.max_row + 1)
            if weights_sheet.cell(row, 1).value == "Factor"
            and weights_sheet.cell(row, 4).value == "Peso entre factores"
        ),
        None,
    )
    if header_row is None:
        raise AssertionError("No se encontró la tabla de pesos en el archivo Excel.")
    excel_weights = {}
    for row in range(header_row + 1, header_row + 1 + len(weights)):
        factor = weights_sheet.cell(row, 1).value
        if factor:
            excel_weights[str(factor)] = {
                "shapley_r2": weights_sheet.cell(row, 3).value,
                "peso": weights_sheet.cell(row, 4).value,
                "peso_total": weights_sheet.cell(row, 5).value,
            }
    if set(excel_weights) != set(weights["factor"]):
        raise AssertionError("Los factores del archivo Excel no coinciden con el CSV Shapley.")
    for record in weights.to_dict("records"):
        excel_record = excel_weights[record["factor"]]
        assert_close(excel_record["shapley_r2"], record["shapley_r2"], record["factor"])
        assert_close(
            excel_record["peso"],
            record["peso_entre_factores_pct"] / 100.0,
            f"Peso de {record['factor']}",
        )
        assert_close(
            excel_record["peso_total"],
            record["peso_r2_total_pct"] / 100.0,
            f"Peso total de {record['factor']}",
        )

    robustness_sheet = workbook["Robustez"]
    bootstrap_header = next(
        (
            row
            for row in range(1, robustness_sheet.max_row + 1)
            if robustness_sheet.cell(row, 1).value == "Factor"
            and robustness_sheet.cell(row, 2).value == "Peso puntual"
        ),
        None,
    )
    if bootstrap_header is None:
        raise AssertionError("No se encontró la tabla bootstrap en Robustez.")
    excel_bootstrap = {
        str(robustness_sheet.cell(row, 1).value): [
            robustness_sheet.cell(row, column).value for column in range(2, 8)
        ]
        for row in range(bootstrap_header + 1, bootstrap_header + 1 + len(bootstrap))
    }
    for record in bootstrap.to_dict("records"):
        values = excel_bootstrap.get(record["factor"])
        if values is None:
            raise AssertionError(f"Falta {record['factor']} en Robustez.")
        expected = [
            record["peso_puntual_pct"] / 100.0,
            record["peso_bootstrap_mediana_pct"] / 100.0,
            record["ic_95_inferior_pct"] / 100.0,
            record["ic_95_superior_pct"] / 100.0,
            (record["ic_95_superior_pct"] - record["ic_95_inferior_pct"]) / 100.0,
            record["probabilidad_top3_pct"] / 100.0,
        ]
        for label, actual, expected_value in zip(
            ["peso", "mediana", "límite inferior", "límite superior", "ancho", "top 3"],
            values,
            expected,
        ):
            assert_close(actual, expected_value, f"Robustez {record['factor']} — {label}")

    bei_sheet = workbook["BEI_robustez"]
    bei_header = next(
        (
            row
            for row in range(1, bei_sheet.max_row + 1)
            if bei_sheet.cell(row, 1).value == "Especificación"
            and bei_sheet.cell(row, 2).value == "Agregación"
        ),
        None,
    )
    if bei_header is None:
        raise AssertionError("No se encontró la comparación de especificaciones en BEI_robustez.")
    excel_bei = {
        str(bei_sheet.cell(row, 1).value): [
            bei_sheet.cell(row, column).value for column in range(5, 11)
        ]
        for row in range(bei_header + 1, bei_header + 1 + len(bei_specs))
    }
    for record in bei_specs.to_dict("records"):
        values = excel_bei.get(record["especificacion"])
        if values is None:
            raise AssertionError(f"Falta {record['especificacion']} en BEI_robustez.")
        expected = [
            record["r_cuadrado_ajustado"],
            record["bic"],
            record["coeficiente_bei_pre_quiebre"],
            record["p_valor_hac_bei_pre_quiebre"],
            record["mape_condicional_pct"] / 100.0,
            record["r2_validacion_condicional_vs_caminata"],
        ]
        for label, actual, expected_value in zip(
            ["R² ajustado", "BIC", "coeficiente", "p HAC", "MAPE", "R² validación"],
            values,
            expected,
        ):
            assert_close(
                actual,
                expected_value,
                f"BEI_robustez {record['especificacion']} — {label}",
            )

    summary = workbook["Resumen"]
    comparison_header = next(
        (
            row
            for row in range(1, summary.max_row + 1)
            if summary.cell(row, 8).value == "Modelo"
            and summary.cell(row, 9).value == "Obs."
        ),
        None,
    )
    if comparison_header is None:
        raise AssertionError("No se encontró la comparación de modelos en el archivo Excel.")
    excel_comparison = {
        str(summary.cell(row, 8).value): [
            summary.cell(row, column).value for column in range(9, 15)
        ]
        for row in range(comparison_header + 1, comparison_header + 3)
    }
    for record in comparison.to_dict("records"):
        values = excel_comparison.get(record["modelo"])
        if values is None:
            raise AssertionError(f"Falta {record['modelo']} en la comparación del archivo Excel.")
        expected = [
            record["observaciones"],
            record["r_cuadrado_ajustado"],
            record["aic"],
            record["bic"],
            record["mape_pct"] / 100.0,
            record["acierto_direccion_pct"] / 100.0,
        ]
        for label, actual, expected_value in zip(
            ["observaciones", "R² ajustado", "AIC", "BIC", "MAPE", "dirección"],
            values,
            expected,
        ):
            assert_close(actual, expected_value, f"{record['modelo']} — {label}")

    forecast_sheet = workbook["Pronostico"]
    forecast_header = next(
        (
            row
            for row in range(1, forecast_sheet.max_row + 1)
            if forecast_sheet.cell(row, 1).value == "Modelo"
            and forecast_sheet.cell(row, 2).value == "Observaciones"
        ),
        None,
    )
    if forecast_header is None:
        raise AssertionError("No se encontró la tabla de pronóstico en el archivo Excel.")
    excel_forecast = {
        str(forecast_sheet.cell(row, 1).value): [
            forecast_sheet.cell(row, column).value for column in range(2, 7)
        ]
        for row in range(forecast_header + 1, forecast_header + 3)
    }
    for record in forecast_metrics.to_dict("records"):
        values = excel_forecast.get(record["modelo"])
        if values is None:
            raise AssertionError(
                f"Falta {record['modelo']} en la tabla de pronóstico del archivo Excel."
            )
        expected = [
            record["observaciones"],
            record["mae_log"],
            record["rmse_log"],
            record["mape_pct"] / 100.0,
            None
            if pd.isna(record["acierto_direccion_pct"])
            else record["acierto_direccion_pct"] / 100.0,
        ]
        for label, actual, expected_value in zip(
            ["observaciones", "MAE", "RMSE", "MAPE", "dirección"],
            values,
            expected,
        ):
            if expected_value is None:
                if actual is not None:
                    raise AssertionError(
                        f"{record['modelo']} — {label}: se esperaba una celda vacía."
                    )
            else:
                assert_close(
                    actual,
                    expected_value,
                    f"{record['modelo']} — {label}",
                )

    interpretation, factor_contributions_integrated = _validate_factor_explanation_outputs(
        weights,
        bootstrap,
        stability_detail,
        stability_summary,
    )
    _validate_factor_workbook(
        workbook,
        interpretation,
        factor_contributions_integrated,
    )

    print(
        "OK: robustez BEI, vintages, bootstrap Shapley, submuestras por término, "
        "explicación por factor, contribuciones mensuales y archivo Excel sincronizado."
    )


if __name__ == "__main__":
    main()
