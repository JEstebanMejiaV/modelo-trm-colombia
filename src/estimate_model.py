from __future__ import annotations

import hashlib
import json
import math
import warnings
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
from scipy import stats
from statsmodels.stats.diagnostic import (
    acorr_breusch_godfrey,
    acorr_ljungbox,
    breaks_cusumolsresid,
    het_arch,
    linear_reset,
)
from statsmodels.stats.stattools import durbin_watson, jarque_bera
from statsmodels.tsa.ardl import ARDL, UECM
from statsmodels.tsa.stattools import adfuller, kpss, zivot_andrews


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
RESULTS = ROOT / "results"
DATA = ROOT / "data"
SAMPLE_START = pd.Timestamp("2006-01-01")
SAMPLE_END = pd.Timestamp("2026-04-01")
SHAPLEY_BOOTSTRAP_REPLICATIONS = 200
SHAPLEY_BOOTSTRAP_BLOCK_MONTHS = 12
SHAPLEY_BOOTSTRAP_PERMUTATIONS = 64
SHAPLEY_BOOTSTRAP_SEED = 20260823

MONTH_NUMBERS_ES = {
    "Ene": 1,
    "Feb": 2,
    "Mar": 3,
    "Abr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Ago": 8,
    "Sep": 9,
    "Set": 9,
    "Oct": 10,
    "Nov": 11,
    "Dic": 12,
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


ECM_LEVEL_VARIABLES = [
    "ln_terminos_intercambio",
    "ln_remesas_12m",
    "diferencial_tasas_pp",
    "deficit_fiscal_12m_pct_pib",
    "ln_dolar_amplio",
]


# Cada factor es un jugador de la descomposicion Shapley. Todos sus terminos
# (transformaciones y rezagos) entran o salen juntos al calcular el R2 marginal.
BASE_FACTOR_SPECS = {
    "Términos de intercambio": {
        "grupo": "Sector externo Colombia",
        "terminos": [("D.ln_terminos_intercambio", 0)],
    },
    "Remesas": {
        "grupo": "Sector externo Colombia",
        "terminos": [("D.ln_remesas_12m", 1)],
    },
    "Diferencial de tasas": {
        "grupo": "Política doméstica",
        "terminos": [("D.diferencial_tasas_pp", 1)],
    },
    "Déficit fiscal": {
        "grupo": "Política doméstica",
        "terminos": [("D.deficit_fiscal_12m_pct_pib", 1)],
    },
    "Dólar amplio": {
        "grupo": "Global",
        "terminos": [("D.ln_dolar_amplio", 0)],
    },
    "VIX": {
        "grupo": "Global",
        "terminos": [("D.ln_vix", 0)],
    },
}


def expanded_factor_specs(regional_component: str) -> dict[str, dict[str, object]]:
    return {
    **BASE_FACTOR_SPECS,
    "Riesgo soberano EMBIG Colombia": {
        "grupo": "Riesgo local",
        # Contemporaneo: esta version es contabilidad historica/nowcast, no causal.
        "terminos": [("D.embig_colombia_pp", 0)],
    },
    "Reservas internacionales": {
        "grupo": "Sector externo Colombia",
        "terminos": [("D.ln_reservas_netas_sin_flar", 1)],
    },
    "Balanza comercial cambiaria": {
        "grupo": "Sector externo Colombia",
        "terminos": [("D.asinh_balanza_comercial", 1)],
    },
    "Flujos netos de capital": {
        "grupo": "Sector externo Colombia",
        "terminos": [("D.asinh_flujos_capital", 1)],
    },
    "Diferencial de compensación inflacionaria 5 años": {
        "grupo": "Política doméstica",
        "terminos": [("D.diferencial_bei_5y_pp", 1)],
    },
    "Monedas regionales": {
        "grupo": "Regional",
        "terminos": [(regional_component, 0)],
    },
}


EXPANDED_FACTOR_SPECS_3 = expanded_factor_specs("factor_monedas_regionales_3")
EXPANDED_FACTOR_SPECS_4 = expanded_factor_specs("factor_monedas_regionales_4")
EXPANDED_FACTOR_SPECS = EXPANDED_FACTOR_SPECS_4


def forecast_factor_specs(regional_component: str) -> dict[str, dict[str, object]]:
    """Especificación ex ante: ninguna variable del mes objetivo entra contemporánea."""
    return {
        "Términos de intercambio": {
            "grupo": "Sector externo Colombia",
            "terminos": [("D.ln_terminos_intercambio", 3)],
        },
        "Remesas": {
            "grupo": "Sector externo Colombia",
            "terminos": [("D.ln_remesas_12m", 2)],
        },
        "Diferencial de tasas": {
            "grupo": "Política doméstica",
            "terminos": [("D.diferencial_tasas_pp", 1)],
        },
        "Déficit fiscal": {
            "grupo": "Política doméstica",
            "terminos": [("D.deficit_fiscal_12m_pct_pib", 3)],
        },
        "Dólar amplio": {
            "grupo": "Global",
            "terminos": [("D.ln_dolar_amplio", 1)],
        },
        "VIX": {
            "grupo": "Global",
            "terminos": [("D.ln_vix", 1)],
        },
        "Riesgo soberano EMBIG Colombia": {
            "grupo": "Riesgo local",
            "terminos": [("D.embig_colombia_pp", 1)],
        },
        "Reservas internacionales": {
            "grupo": "Sector externo Colombia",
            "terminos": [("D.ln_reservas_netas_sin_flar", 2)],
        },
        "Balanza comercial cambiaria": {
            "grupo": "Sector externo Colombia",
            "terminos": [("D.asinh_balanza_comercial", 2)],
        },
        "Flujos netos de capital": {
            "grupo": "Sector externo Colombia",
            "terminos": [("D.asinh_flujos_capital", 2)],
        },
        "Diferencial de compensación inflacionaria 5 años": {
            "grupo": "Política doméstica",
            "terminos": [("D.diferencial_bei_5y_pp", 1)],
        },
        "Monedas regionales": {
            "grupo": "Regional",
            "terminos": [(regional_component, 1)],
        },
    }


FORECAST_FACTOR_SPECS_3 = forecast_factor_specs("factor_monedas_regionales_3")
FORECAST_FACTOR_SPECS_4 = forecast_factor_specs("factor_monedas_regionales_4")


FORECAST_AVAILABILITY = [
    ("Términos de intercambio", 3, "Mensual; publicación aproximada t+2", "Último cambio utilizable al inicio de t: t-3"),
    ("Remesas", 2, "Mensual; publicación posterior al mes de referencia", "Supuesto conservador: t-2"),
    ("Diferencial de tasas", 1, "Tasas observables durante el mes", "Promedios completos conocidos para t-1"),
    ("Déficit fiscal", 3, "Mensual con rezago y posibles revisiones", "Supuesto conservador: t-3"),
    ("Dólar amplio", 1, "Diaria", "Promedio completo conocido para t-1"),
    ("VIX", 1, "Diaria", "Promedio completo conocido para t-1"),
    ("Riesgo soberano EMBIG Colombia", 1, "Diaria", "Promedio completo conocido para t-1"),
    ("Reservas internacionales", 2, "Mensual; publicación posterior al cierre", "Supuesto conservador: t-2"),
    ("Balanza comercial cambiaria", 2, "Mensual; publicación posterior al cierre", "Supuesto conservador: t-2"),
    ("Flujos netos de capital", 2, "Mensual; publicación posterior al cierre", "Supuesto conservador: t-2"),
    ("Diferencial de compensación inflacionaria 5 años", 1, "Curvas diarias", "Promedios completos conocidos para t-1"),
    ("Monedas regionales", 1, "Tipos de cambio mensuales", "Promedios completos conocidos para t-1"),
]


DIFFERENCED_COMPONENTS = [
    "ln_terminos_intercambio",
    "ln_remesas_12m",
    "diferencial_tasas_pp",
    "deficit_fiscal_12m_pct_pib",
    "ln_dolar_amplio",
    "ln_vix",
    "embig_colombia_pp",
    "ln_reservas_netas_sin_flar",
    "asinh_balanza_comercial",
    "asinh_flujos_capital",
    "diferencial_bei_5y_pp",
    "diferencial_bei_5y_comun_pp",
]


LEVEL_COMPONENTS = [
    "diferencial_bei_5y_pp",
    "diferencial_bei_5y_comun_pp",
    "factor_monedas_regionales_3",
    "factor_monedas_regionales_4",
]


@dataclass
class SelectedModel:
    p: int
    q: int
    result: object


@dataclass
class SelectedDifferenceModel:
    p: int
    q: int
    result: object
    y: pd.Series
    x: pd.DataFrame


def month_start(values: pd.Series | pd.DatetimeIndex) -> pd.DatetimeIndex:
    return pd.DatetimeIndex(values).to_period("M").to_timestamp()


def read_fred(path: Path, output_name: str, daily: bool = False) -> pd.Series:
    raw = pd.read_csv(path)
    frame = pd.DataFrame(
        {
            "fecha": pd.to_datetime(raw.iloc[:, 0].astype("string"), errors="coerce"),
            "valor": pd.to_numeric(raw.iloc[:, 1], errors="coerce"),
        }
    ).dropna()
    series = frame.set_index("fecha")["valor"].sort_index()
    if daily:
        series = series.resample("MS").mean()
    else:
        series.index = month_start(series.index)
        series = series.groupby(level=0).mean()
    series.name = output_name
    return series


def load_fed_gsw_breakeven_daily(path: Path) -> pd.Series:
    """Lee BKEVEN05 diario del archivo Gürkaynak-Sack-Wright."""
    raw = pd.read_csv(
        path,
        skiprows=18,
        usecols=["Date", "BKEVEN05"],
        na_values=["NA"],
    )
    frame = pd.DataFrame(
        {
            "fecha": pd.to_datetime(raw["Date"], errors="coerce"),
            "bei_eeuu_5y_pct": pd.to_numeric(raw["BKEVEN05"], errors="coerce"),
        }
    ).dropna()
    series = frame.set_index("fecha")["bei_eeuu_5y_pct"].sort_index()
    series.index = series.index.normalize()
    series = series.groupby(level=0).mean()
    series.name = "bei_eeuu_5y_pct"
    return series


def load_fed_gsw_breakeven(path: Path) -> pd.Series:
    daily = load_fed_gsw_breakeven_daily(path)
    monthly = daily.resample("MS").mean()
    monthly.name = "bei_eeuu_5y_pct"
    return monthly


def load_banrep_series(path: Path, output_name: str, daily: bool = False) -> pd.Series:
    """Lee el JSON publico del graficador de BanRep y lo lleva a frecuencia mensual."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    item = payload[0] if isinstance(payload, list) else payload
    data = pd.DataFrame(item["data"], columns=["timestamp_ms", output_name])
    dates = pd.to_datetime(data.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    data.index = pd.DatetimeIndex(dates)
    series = pd.to_numeric(data[output_name], errors="coerce").dropna().sort_index()
    if daily:
        # Se promedia cada mercado por separado; no se cruzan calendarios diarios.
        series = series.resample("MS").mean()
    else:
        series.index = month_start(series.index)
        series = series.groupby(level=0).mean()
    series.name = output_name
    return series


def load_banrep_observations(path: Path, output_name: str) -> pd.Series:
    """Lee observaciones BanRep sin agregarlas para comparar calendarios diarios."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    item = payload[0] if isinstance(payload, list) else payload
    frame = pd.DataFrame(item["data"], columns=["timestamp_ms", output_name])
    dates = pd.to_datetime(frame.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    frame.index = pd.DatetimeIndex(dates).normalize()
    series = pd.to_numeric(frame[output_name], errors="coerce").dropna().sort_index()
    series = series.groupby(level=0).mean()
    series.name = output_name
    return series


def build_bei_aggregations() -> pd.DataFrame:
    """Construye BEI mensual con medias separadas y con fechas diarias comunes."""
    nominal = load_banrep_observations(
        RAW / "tes_5y_pesos_banrep.json", "tes_5y_pesos_colombia_pct"
    )
    real = load_banrep_observations(
        RAW / "tes_5y_uvr_banrep.json", "tes_5y_uvr_colombia_pct"
    )
    us = load_fed_gsw_breakeven_daily(RAW / "bei_5y_eeuu_diario_fed.csv")

    separate = pd.concat(
        [
            nominal.resample("MS").mean(),
            real.resample("MS").mean(),
            us.resample("MS").mean(),
        ],
        axis=1,
        sort=True,
    )
    separate["bei_colombia_5y_pct"] = (
        separate["tes_5y_pesos_colombia_pct"]
        - separate["tes_5y_uvr_colombia_pct"]
    )
    separate["diferencial_bei_5y_pp"] = (
        separate["bei_colombia_5y_pct"] - separate["bei_eeuu_5y_pct"]
    )

    common_daily = pd.concat([nominal, real, us], axis=1, join="inner").dropna()
    common_daily["diferencial_bei_5y_comun_pp"] = (
        common_daily["tes_5y_pesos_colombia_pct"]
        - common_daily["tes_5y_uvr_colombia_pct"]
        - common_daily["bei_eeuu_5y_pct"]
    )
    common = common_daily.resample("MS").agg(
        tes_5y_pesos_comun_pct=("tes_5y_pesos_colombia_pct", "mean"),
        tes_5y_uvr_comun_pct=("tes_5y_uvr_colombia_pct", "mean"),
        bei_eeuu_5y_comun_pct=("bei_eeuu_5y_pct", "mean"),
        diferencial_bei_5y_comun_pp=("diferencial_bei_5y_comun_pp", "mean"),
        dias_comunes=("diferencial_bei_5y_comun_pp", "count"),
    )
    counts = pd.concat(
        [
            nominal.resample("MS").count().rename("dias_tes_pesos"),
            real.resample("MS").count().rename("dias_tes_uvr"),
            us.resample("MS").count().rename("dias_bei_eeuu"),
        ],
        axis=1,
        sort=True,
    )
    out = separate.join(common, how="outer").join(counts, how="outer")
    out["diferencia_comun_menos_separada_pp"] = (
        out["diferencial_bei_5y_comun_pp"] - out["diferencial_bei_5y_pp"]
    )
    out.index.name = "fecha"
    return out


def load_remittances(path: Path) -> pd.Series:
    payload = json.loads(path.read_text(encoding="utf-8"))[0]
    data = pd.DataFrame(payload["data"], columns=["timestamp_ms", "remesas_usd_millones"])
    dates = pd.to_datetime(data.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    data.index = month_start(dates)
    return data["remesas_usd_millones"].astype(float).sort_index()


def load_banrep_daily(path: Path, output_name: str) -> pd.Series:
    return load_banrep_series(path, output_name, daily=True)


def load_terms_of_trade(path: Path) -> pd.Series:
    payload = json.loads(path.read_text(encoding="utf-8"))
    series_payload = next(item for item in payload if item.get("id") == 15360)
    data = pd.DataFrame(series_payload["data"], columns=["timestamp_ms", "terminos_intercambio"])
    dates = pd.to_datetime(data.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    data.index = month_start(dates)
    return data["terminos_intercambio"].astype(float).sort_index()


def load_embig_bcrp(path: Path) -> pd.Series:
    """Lee EMBIG Colombia del JSON público del BCRP y promedia por mes."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[tuple[pd.Timestamp, float]] = []
    for observation in payload.get("periods", []):
        parts = str(observation.get("name", "")).strip().split(".")
        values = observation.get("values") or []
        if len(parts) != 3 or not values or parts[1] not in MONTH_NUMBERS_ES:
            continue
        year = int(parts[2])
        year += 2000 if year < 70 else 1900
        date = pd.Timestamp(year=year, month=MONTH_NUMBERS_ES[parts[1]], day=int(parts[0]))
        value = pd.to_numeric(str(values[0]).replace(",", "."), errors="coerce")
        if pd.notna(value):
            rows.append((date, float(value)))
    if not rows:
        raise ValueError(f"No se encontraron observaciones EMBIG válidas en {path}.")
    daily = pd.Series(
        (value for _, value in rows),
        index=pd.DatetimeIndex(date for date, _ in rows),
        name="embig_colombia_pb",
    ).sort_index()
    daily = daily.groupby(level=0).mean()
    monthly = daily.resample("MS").mean()
    monthly.name = "embig_colombia_pb"
    return monthly


def load_bcrp_monthly(path: Path, output_name: str) -> pd.Series:
    """Lee una serie mensual de BCRPData con periodos como Ene.2006."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[tuple[pd.Timestamp, float]] = []
    for observation in payload.get("periods", []):
        parts = str(observation.get("name", "")).strip().split(".")
        values = observation.get("values") or []
        if len(parts) != 2 or not values or parts[0] not in MONTH_NUMBERS_ES:
            continue
        value = pd.to_numeric(str(values[0]).replace(",", "."), errors="coerce")
        if pd.notna(value):
            rows.append(
                (
                    pd.Timestamp(year=int(parts[1]), month=MONTH_NUMBERS_ES[parts[0]], day=1),
                    float(value),
                )
            )
    if not rows:
        raise ValueError(f"No se encontraron observaciones mensuales válidas en {path}.")
    series = pd.Series(
        (value for _, value in rows),
        index=pd.DatetimeIndex(date for date, _ in rows),
        name=output_name,
    ).sort_index()
    return series.groupby(level=0).mean()


def _row_values(ws, row_number: int) -> tuple[list[object], list[object]]:
    dates = [cell.value for cell in ws[6]][1:]
    values = [cell.value for cell in ws[row_number]][1:]
    return dates, values


def load_fiscal(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    monthly_amounts = workbook.worksheets[0]
    monthly_pct = workbook.worksheets[3]

    dates, balance_values = _row_values(monthly_amounts, 31)
    _, income_values = _row_values(monthly_amounts, 8)
    pct_dates, income_pct_values = _row_values(monthly_pct, 8)

    fiscal = pd.DataFrame(
        {
            "fecha": pd.to_datetime(dates, errors="coerce"),
            "balance_fiscal_miles_millones_cop": pd.to_numeric(balance_values, errors="coerce"),
            "ingresos_totales_miles_millones_cop": pd.to_numeric(income_values, errors="coerce"),
        }
    ).dropna(subset=["fecha"])
    fiscal["fecha"] = month_start(fiscal["fecha"])
    fiscal = fiscal.set_index("fecha").sort_index()

    fiscal_pct = pd.DataFrame(
        {
            "fecha": pd.to_datetime(pct_dates, errors="coerce"),
            "ingresos_totales_pct_pib": pd.to_numeric(income_pct_values, errors="coerce"),
        }
    ).dropna(subset=["fecha"])
    fiscal_pct["fecha"] = month_start(fiscal_pct["fecha"])
    fiscal_pct = fiscal_pct.set_index("fecha").sort_index()

    fiscal = fiscal.join(fiscal_pct, how="left")
    valid = fiscal["ingresos_totales_pct_pib"].abs() > 1e-9
    fiscal.loc[valid, "pib_anual_miles_millones_cop_observado"] = (
        100.0
        * fiscal.loc[valid, "ingresos_totales_miles_millones_cop"]
        / fiscal.loc[valid, "ingresos_totales_pct_pib"]
    )
    year_gdp = fiscal.groupby(fiscal.index.year)["pib_anual_miles_millones_cop_observado"].median()
    fiscal["pib_anual_miles_millones_cop"] = fiscal.index.year.map(year_gdp)
    fiscal["balance_fiscal_12m_miles_millones_cop"] = fiscal[
        "balance_fiscal_miles_millones_cop"
    ].rolling(12, min_periods=12).sum()
    fiscal["deficit_fiscal_12m_pct_pib"] = (
        -100.0
        * fiscal["balance_fiscal_12m_miles_millones_cop"]
        / fiscal["pib_anual_miles_millones_cop"]
    )
    return fiscal


def build_dataset() -> pd.DataFrame:
    series = [
        load_banrep_daily(RAW / "trm_diaria_banrep.json", "trm_cop_usd"),
        load_banrep_daily(
            RAW / "tasa_politica_diaria_banrep.json", "tasa_politica_colombia_pct"
        ),
        read_fred(RAW / "fed_funds_mensual_fred.csv", "fed_funds_eeuu_pct"),
        read_fred(RAW / "dolar_amplio_diario_fred.csv", "indice_dolar_amplio", daily=True),
        read_fred(RAW / "vix_diario_fred.csv", "vix", daily=True),
        load_remittances(RAW / "remesas_mensuales_banrep.json"),
        load_terms_of_trade(RAW / "series_15360_15368.json"),
        load_banrep_series(
            RAW / "reservas_netas_sin_flar_banrep.json",
            "reservas_netas_sin_flar_usd_millones",
        ),
        build_bei_aggregations(),
        load_embig_bcrp(RAW / "embig_colombia_diario_bcrp.json"),
        load_banrep_series(
            RAW / "balanza_comercial_cambiaria_banrep.json",
            "balanza_comercial_cambiaria_usd_millones",
        ),
        load_banrep_series(
            RAW / "flujos_capital_totales_banrep.json",
            "flujos_capital_usd_millones",
        ),
        read_fred(RAW / "brl_usd_mensual_fred.csv", "brl_por_usd"),
        read_fred(RAW / "clp_usd_mensual_fred.csv", "clp_por_usd"),
        read_fred(RAW / "mxn_usd_mensual_fred.csv", "mxn_por_usd"),
        load_bcrp_monthly(RAW / "pen_usd_mensual_bcrp.json", "pen_por_usd"),
    ]
    data = pd.concat(series, axis=1, sort=True).sort_index()
    data = data.join(load_fiscal(RAW / "balance_fiscal_gnc_mensual_trimestral.xlsx"), how="outer")

    data["remesas_12m_usd_millones"] = data["remesas_usd_millones"].rolling(
        12, min_periods=12
    ).sum()
    data["diferencial_tasas_pp"] = (
        data["tasa_politica_colombia_pct"] - data["fed_funds_eeuu_pct"]
    )
    data["embig_colombia_pp"] = data["embig_colombia_pb"] / 100.0
    data["asinh_balanza_comercial"] = np.arcsinh(
        data["balanza_comercial_cambiaria_usd_millones"] / 1000.0
    )
    data["asinh_flujos_capital"] = np.arcsinh(
        data["flujos_capital_usd_millones"] / 1000.0
    )

    regional_returns = pd.DataFrame(
        {
            currency: np.log(data[currency].where(data[currency] > 0)).diff()
            for currency in ["brl_por_usd", "clp_por_usd", "mxn_por_usd", "pen_por_usd"]
        }
    )
    calibration = regional_returns.loc["2006-01-01":"2019-12-01"]
    regional_z = (regional_returns - calibration.mean()) / calibration.std(ddof=0)
    data["factor_monedas_regionales_3"] = regional_z[
        ["brl_por_usd", "clp_por_usd", "mxn_por_usd"]
    ].mean(axis=1, skipna=False)
    data["factor_monedas_regionales_4"] = regional_z[
        ["brl_por_usd", "clp_por_usd", "mxn_por_usd", "pen_por_usd"]
    ].mean(axis=1, skipna=False)
    # Alias explícito del modelo ampliado activo: composición de cuatro monedas.
    data["factor_monedas_regionales"] = data["factor_monedas_regionales_4"]

    positive_logs = {
        "ln_trm": "trm_cop_usd",
        "ln_remesas_12m": "remesas_12m_usd_millones",
        "ln_dolar_amplio": "indice_dolar_amplio",
        "ln_vix": "vix",
        "ln_terminos_intercambio": "terminos_intercambio",
        "ln_reservas_netas_sin_flar": "reservas_netas_sin_flar_usd_millones",
    }
    for target, source in positive_logs.items():
        data[target] = np.log(data[source].where(data[source] > 0))
    data["dln_vix"] = data["ln_vix"].diff()
    data["dummy_pandemia_2020"] = (
        (data.index >= pd.Timestamp("2020-03-01"))
        & (data.index <= pd.Timestamp("2020-05-01"))
    ).astype(int)
    data.index.name = "fecha"
    return data


def integration_tests(data: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for column in columns:
        for transform, series in [
            ("nivel", data[column].dropna()),
            ("primera_diferencia", data[column].diff().dropna()),
        ]:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                adf = adfuller(series, regression="c", autolag="BIC")
                try:
                    kpss_result = kpss(series, regression="c", nlags="auto")
                    kpss_stat, kpss_p = float(kpss_result[0]), float(kpss_result[1])
                except Exception:
                    kpss_stat, kpss_p = math.nan, math.nan
            rows.append(
                {
                    "variable": column,
                    "transformacion": transform,
                    "n": int(series.shape[0]),
                    "adf_estadistico": float(adf[0]),
                    "adf_p": float(adf[1]),
                    "adf_rezagos": int(adf[2]),
                    "kpss_estadistico": kpss_stat,
                    "kpss_p": kpss_p,
                }
            )
    return pd.DataFrame(rows)


def bei_stationarity_tests(data: pd.DataFrame) -> pd.DataFrame:
    """Contrasta nivel/diferencia con constante, tendencia y un quiebre endógeno."""
    definitions = {
        "Medias mensuales separadas": "diferencial_bei_5y_pp",
        "Fechas diarias comunes": "diferencial_bei_5y_comun_pp",
    }
    rows: list[dict[str, object]] = []
    for aggregation, column in definitions.items():
        original = data[column].loc[SAMPLE_START:SAMPLE_END].dropna()
        for transformation, series in [
            ("nivel", original),
            ("primera_diferencia", original.diff().dropna()),
        ]:
            for deterministic, regression in [
                ("constante", "c"),
                ("constante_tendencia", "ct"),
            ]:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    adf = adfuller(series, regression=regression, autolag="BIC")
                    kpss_result = kpss(series, regression=regression, nlags="auto")
                rows.extend(
                    [
                        {
                            "agregacion": aggregation,
                            "variable": column,
                            "transformacion": transformation,
                            "prueba": "ADF",
                            "deterministico": deterministic,
                            "hipotesis_nula": "raiz_unitaria",
                            "n": len(series),
                            "estadistico": float(adf[0]),
                            "p_valor": float(adf[1]),
                            "rezagos": int(adf[2]),
                            "fecha_quiebre": "",
                            "critico_5_pct": float(adf[4]["5%"]),
                        },
                        {
                            "agregacion": aggregation,
                            "variable": column,
                            "transformacion": transformation,
                            "prueba": "KPSS",
                            "deterministico": deterministic,
                            "hipotesis_nula": "estacionariedad",
                            "n": len(series),
                            "estadistico": float(kpss_result[0]),
                            "p_valor": float(kpss_result[1]),
                            "rezagos": int(kpss_result[2]),
                            "fecha_quiebre": "",
                            "critico_5_pct": float(kpss_result[3]["5%"]),
                        },
                    ]
                )
            for deterministic, regression in [
                ("constante_con_quiebre", "c"),
                ("constante_tendencia_con_quiebre", "ct"),
            ]:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    za = zivot_andrews(series, regression=regression, autolag="BIC")
                break_date = series.index[int(za[4])]
                rows.append(
                    {
                        "agregacion": aggregation,
                        "variable": column,
                        "transformacion": transformation,
                        "prueba": "Zivot-Andrews",
                        "deterministico": deterministic,
                        "hipotesis_nula": "raiz_unitaria_con_quiebre",
                        "n": len(series),
                        "estadistico": float(za[0]),
                        "p_valor": float(za[1]),
                        "rezagos": int(za[3]),
                        "fecha_quiebre": break_date.strftime("%Y-%m-%d"),
                        "critico_5_pct": float(za[2]["5%"]),
                    }
                )
    return pd.DataFrame(rows)


def bei_trend_break_models(
    data: pd.DataFrame, stationarity: pd.DataFrame
) -> pd.DataFrame:
    """Compara tendencia lineal y tendencia segmentada en el propio diferencial BEI."""
    definitions = {
        "Medias mensuales separadas": "diferencial_bei_5y_pp",
        "Fechas diarias comunes": "diferencial_bei_5y_comun_pp",
    }
    rows: list[dict[str, object]] = []
    for aggregation, column in definitions.items():
        series = data[column].loc[SAMPLE_START:SAMPLE_END].dropna()
        break_record = stationarity.loc[
            stationarity["agregacion"].eq(aggregation)
            & stationarity["transformacion"].eq("nivel")
            & stationarity["prueba"].eq("Zivot-Andrews")
            & stationarity["deterministico"].eq("constante_tendencia_con_quiebre")
        ].iloc[0]
        break_date = pd.Timestamp(break_record["fecha_quiebre"])
        trend_years = pd.Series(
            np.arange(len(series), dtype=float) / 12.0,
            index=series.index,
            name="tendencia_anual",
        )
        post = (series.index >= break_date).astype(float)
        break_year = float(trend_years.loc[break_date])
        post_slope = np.maximum(0.0, trend_years.to_numpy() - break_year)
        designs = {
            "Sin tendencia": pd.DataFrame(index=series.index),
            "Tendencia lineal": pd.DataFrame(
                {"tendencia_anual": trend_years}, index=series.index
            ),
            "Tendencia segmentada con quiebre ZA": pd.DataFrame(
                {
                    "tendencia_anual": trend_years,
                    "cambio_nivel_post_quiebre": post,
                    "cambio_pendiente_post_quiebre": post_slope,
                },
                index=series.index,
            ),
        }
        for model_name, design in designs.items():
            x = sm.add_constant(design, has_constant="add")
            result = sm.OLS(series, x).fit()
            _, coefficients = tidy_robust_ols(result, maxlags=6)
            lookup = coefficients.set_index("termino")

            def value(term: str, field: str) -> float:
                if term not in lookup.index:
                    return math.nan
                return float(lookup.loc[term, field])

            rows.append(
                {
                    "agregacion": aggregation,
                    "variable": column,
                    "modelo_deterministico": model_name,
                    "fecha_quiebre_za": break_date.strftime("%Y-%m-%d"),
                    "observaciones": int(result.nobs),
                    "r_cuadrado_ajustado": float(result.rsquared_adj),
                    "aic": float(result.aic),
                    "bic": float(result.bic),
                    "tendencia_pp_por_ano": value("tendencia_anual", "coeficiente"),
                    "p_valor_hac_tendencia": value("tendencia_anual", "p_valor"),
                    "cambio_nivel_quiebre_pp": value(
                        "cambio_nivel_post_quiebre", "coeficiente"
                    ),
                    "p_valor_hac_cambio_nivel": value(
                        "cambio_nivel_post_quiebre", "p_valor"
                    ),
                    "cambio_pendiente_pp_por_ano": value(
                        "cambio_pendiente_post_quiebre", "coeficiente"
                    ),
                    "p_valor_hac_cambio_pendiente": value(
                        "cambio_pendiente_post_quiebre", "p_valor"
                    ),
                }
            )
    return pd.DataFrame(rows)


def select_ardl(y: pd.Series, exog: pd.DataFrame, fixed: pd.DataFrame) -> tuple[SelectedModel, pd.DataFrame]:
    candidates: list[dict[str, float | int]] = []
    selected: SelectedModel | None = None
    for p in range(1, 5):
        for q in range(1, 3):
            result = ARDL(
                y,
                lags=p,
                exog=exog,
                order=q,
                trend="c",
                fixed=fixed,
                causal=False,
                missing="raise",
            ).fit()
            candidates.append(
                {
                    "p_trm": p,
                    "q_explicativas": q,
                    "aic": float(result.aic),
                    "bic": float(result.bic),
                    "hqic": float(result.hqic),
                    "loglik": float(result.llf),
                }
            )
            if selected is None or result.bic < selected.result.bic:
                selected = SelectedModel(p=p, q=q, result=result)
    assert selected is not None
    return selected, pd.DataFrame(candidates).sort_values("bic").reset_index(drop=True)


def difference_components(model_data: pd.DataFrame) -> pd.DataFrame:
    diff = pd.DataFrame(index=model_data.index)
    diff["D.ln_trm"] = model_data["ln_trm"].diff()
    for variable in DIFFERENCED_COMPONENTS:
        if variable in model_data:
            diff[f"D.{variable}"] = model_data[variable].diff()
    for variable in LEVEL_COMPONENTS:
        if variable in model_data:
            diff[variable] = model_data[variable]
    diff["dummy_pandemia_2020"] = model_data["dummy_pandemia_2020"]
    return diff


def make_difference_design(
    components: pd.DataFrame, p: int, q: int, index: pd.Index | None = None
) -> tuple[pd.Series, pd.DataFrame]:
    y = components["D.ln_trm"].rename("D.ln_trm")
    x = pd.DataFrame(index=components.index)
    for lag in range(1, p + 1):
        x[f"D.ln_trm.L{lag}"] = components["D.ln_trm"].shift(lag)
    drivers = [f"D.{variable}" for variable in ECM_LEVEL_VARIABLES] + ["D.ln_vix"]
    for driver in drivers:
        for lag in range(0, q + 1):
            x[f"{driver}.L{lag}"] = components[driver].shift(lag)
    x["dummy_pandemia_2020"] = components["dummy_pandemia_2020"]
    x = sm.add_constant(x, has_constant="add")
    combined = pd.concat([y, x], axis=1).dropna()
    if index is not None:
        combined = combined.reindex(index).dropna()
    return combined["D.ln_trm"], combined.drop(columns="D.ln_trm")


def select_difference_model(model_data: pd.DataFrame) -> tuple[SelectedDifferenceModel, pd.DataFrame]:
    components = difference_components(model_data)
    common_index = make_difference_design(components, p=3, q=2)[0].index
    candidates: list[dict[str, float | int]] = []
    selected: SelectedDifferenceModel | None = None
    for p in range(0, 4):
        for q in range(0, 3):
            y, x = make_difference_design(components, p=p, q=q, index=common_index)
            result = sm.OLS(y, x).fit()
            candidates.append(
                {
                    "p_cambio_trm": p,
                    "q_cambios_explicativas": q,
                    "aic": float(result.aic),
                    "bic": float(result.bic),
                    "r_cuadrado_ajustado": float(result.rsquared_adj),
                }
            )
            if selected is None or result.bic < selected.result.bic:
                selected = SelectedDifferenceModel(p=p, q=q, result=result, y=y, x=x)
    assert selected is not None
    grid = pd.DataFrame(candidates).sort_values("bic").reset_index(drop=True)
    return selected, grid


def design_term_name(component: str, lag: int) -> str:
    return f"{component}.L{lag}"


def make_timed_difference_design(
    components: pd.DataFrame,
    p: int,
    factor_specs: dict[str, dict[str, object]],
    index: pd.Index | None = None,
) -> tuple[pd.Series, pd.DataFrame]:
    y = components["D.ln_trm"].rename("D.ln_trm")
    x = pd.DataFrame(index=components.index)
    for lag in range(1, p + 1):
        x[f"D.ln_trm.L{lag}"] = components["D.ln_trm"].shift(lag)

    for factor in factor_specs.values():
        for component, lag in factor["terminos"]:
            x[design_term_name(component, lag)] = components[component].shift(lag)
    x["dummy_pandemia_2020"] = components["dummy_pandemia_2020"]
    x = sm.add_constant(x, has_constant="add")
    combined = pd.concat([y, x], axis=1).dropna()
    if index is not None:
        combined = combined.reindex(index).dropna()
    return combined["D.ln_trm"], combined.drop(columns="D.ln_trm")


def select_timed_difference_model(
    model_data: pd.DataFrame,
    factor_specs: dict[str, dict[str, object]],
    common_index: pd.Index | None = None,
) -> tuple[SelectedDifferenceModel, pd.DataFrame]:
    components = difference_components(model_data)
    if common_index is None:
        common_index = make_timed_difference_design(
            components, p=3, factor_specs=factor_specs
        )[0].index
    candidates: list[dict[str, float | int]] = []
    selected: SelectedDifferenceModel | None = None
    for p in range(0, 4):
        y, x = make_timed_difference_design(
            components, p=p, factor_specs=factor_specs, index=common_index
        )
        result = sm.OLS(y, x).fit()
        candidates.append(
            {
                "p_cambio_trm": p,
                "aic": float(result.aic),
                "bic": float(result.bic),
                "r_cuadrado_ajustado": float(result.rsquared_adj),
            }
        )
        if selected is None or result.bic < selected.result.bic:
            selected = SelectedDifferenceModel(p=p, q=0, result=result, y=y, x=x)
    assert selected is not None
    return selected, pd.DataFrame(candidates).sort_values("bic").reset_index(drop=True)


def tidy_robust_ols(result, maxlags: int = 6) -> tuple[object, pd.DataFrame]:
    robust = result.get_robustcov_results(
        cov_type="HAC", maxlags=maxlags, use_correction=True, use_t=True
    )
    names = result.model.exog_names
    confidence = robust.conf_int(alpha=0.05)
    table = pd.DataFrame(
        {
            "termino": names,
            "coeficiente": robust.params,
            "error_estandar_hac": robust.bse,
            "estadistico_t": robust.tvalues,
            "p_valor": robust.pvalues,
            "ic_95_inferior": confidence[:, 0],
            "ic_95_superior": confidence[:, 1],
        }
    )
    return robust, table


def tidy_result(result) -> pd.DataFrame:
    confidence = result.conf_int(alpha=0.05)
    return pd.DataFrame(
        {
            "termino": result.params.index,
            "coeficiente": result.params.values,
            "error_estandar_hac": result.bse.values,
            "estadistico_t": result.tvalues.values,
            "p_valor": result.pvalues.values,
            "ic_95_inferior": confidence.iloc[:, 0].values,
            "ic_95_superior": confidence.iloc[:, 1].values,
        }
    )


def tidy_long_run(result) -> pd.DataFrame:
    confidence = result.ci_conf_int(alpha=0.05)
    return pd.DataFrame(
        {
            "termino": result.ci_params.index,
            "coeficiente_largo_plazo": result.ci_params.values,
            "error_estandar": result.ci_bse.values,
            "estadistico_t": result.ci_tvalues.values,
            "p_valor": result.ci_pvalues.values,
            "ic_95_inferior": confidence.iloc[:, 0].values,
            "ic_95_superior": confidence.iloc[:, 1].values,
        }
    )


def diagnostics(result) -> pd.DataFrame:
    residuals = pd.Series(result.resid).dropna()
    lb = acorr_ljungbox(residuals, lags=[6, 12], return_df=True)
    arch = het_arch(residuals, nlags=12)
    jb = jarque_bera(residuals)
    if hasattr(result.model, "_y") and hasattr(result.model, "_x"):
        ols_proxy = sm.OLS(result.model._y, result.model._x).fit()
    else:
        ols_proxy = result
    bg = acorr_breusch_godfrey(ols_proxy, nlags=12)
    reset = linear_reset(ols_proxy, power=2, use_f=True)
    cusum = breaks_cusumolsresid(residuals, ddof=int(result.df_model) + 1)
    return pd.DataFrame(
        [
            {"prueba": "Ljung-Box (6)", "estadistico": lb.loc[6, "lb_stat"], "p_valor": lb.loc[6, "lb_pvalue"]},
            {"prueba": "Ljung-Box (12)", "estadistico": lb.loc[12, "lb_stat"], "p_valor": lb.loc[12, "lb_pvalue"]},
            {"prueba": "Breusch-Godfrey (12)", "estadistico": bg[0], "p_valor": bg[1]},
            {"prueba": "ARCH-LM (12)", "estadistico": arch[0], "p_valor": arch[1]},
            {"prueba": "Jarque-Bera", "estadistico": jb[0], "p_valor": jb[1]},
            {"prueba": "Ramsey RESET", "estadistico": float(reset.fvalue), "p_valor": float(reset.pvalue)},
            {"prueba": "CUSUM estabilidad", "estadistico": cusum[0], "p_valor": cusum[1]},
            {"prueba": "Durbin-Watson", "estadistico": durbin_watson(residuals), "p_valor": np.nan},
        ]
    )


def expanding_validation(
    y: pd.Series,
    exog: pd.DataFrame,
    fixed: pd.DataFrame,
    p: int,
    q: int,
    holdout: int = 48,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    predictions: list[dict[str, object]] = []
    split = len(y) - holdout
    for i in range(split, len(y)):
        train_y = y.iloc[:i]
        train_x = exog.iloc[:i]
        train_fixed = fixed.iloc[:i]
        model = ARDL(
            train_y,
            lags=p,
            exog=train_x,
            order=q,
            trend="c",
            fixed=train_fixed,
            causal=False,
            missing="raise",
        ).fit()
        forecast_log = float(model.forecast(1, exog=exog.iloc[[i]], fixed=fixed.iloc[[i]]).iloc[0])
        predictions.append(
            {
                "fecha": y.index[i],
                "ln_trm_observada": float(y.iloc[i]),
                "ln_trm_modelo_condicional": forecast_log,
                "ln_trm_caminata_aleatoria": float(y.iloc[i - 1]),
            }
        )
    pred = pd.DataFrame(predictions).set_index("fecha")
    pred["trm_observada"] = np.exp(pred["ln_trm_observada"])
    pred["trm_modelo_condicional"] = np.exp(pred["ln_trm_modelo_condicional"])
    pred["trm_caminata_aleatoria"] = np.exp(pred["ln_trm_caminata_aleatoria"])
    pred["cambio_observado"] = pred["ln_trm_observada"].diff()
    pred["cambio_modelo"] = pred["ln_trm_modelo_condicional"] - pred["ln_trm_caminata_aleatoria"]
    pred["cambio_caminata"] = 0.0

    metrics: list[dict[str, object]] = []
    for label, forecast_col in [
        ("ARDL condicional", "ln_trm_modelo_condicional"),
        ("Caminata aleatoria", "ln_trm_caminata_aleatoria"),
    ]:
        errors = pred[forecast_col] - pred["ln_trm_observada"]
        if label == "ARDL condicional":
            direction = np.sign(pred["cambio_modelo"])
        else:
            direction = np.sign(pred["cambio_caminata"])
        observed_direction = np.sign(pred["ln_trm_observada"] - pred["ln_trm_caminata_aleatoria"])
        direction_hit = float((direction == observed_direction).mean())
        metrics.append(
            {
                "modelo": label,
                "observaciones": int(errors.shape[0]),
                "mae_log": float(errors.abs().mean()),
                "rmse_log": float(np.sqrt(np.mean(np.square(errors)))),
                "mape_pct": float(
                    100
                    * np.mean(
                        np.abs(
                            np.exp(pred[forecast_col]) - np.exp(pred["ln_trm_observada"])
                        )
                        / np.exp(pred["ln_trm_observada"])
                    )
                ),
                "acierto_direccion_pct": 100 * direction_hit,
            }
        )
    return pred, pd.DataFrame(metrics)


def difference_validation(
    model_data: pd.DataFrame,
    selected: SelectedDifferenceModel,
    holdout: int = 48,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    y, x = selected.y, selected.x
    split = len(y) - holdout
    rows: list[dict[str, object]] = []
    for i in range(split, len(y)):
        train = sm.OLS(y.iloc[:i], x.iloc[:i]).fit()
        forecast_change = float(train.predict(x.iloc[[i]]).iloc[0])
        date = y.index[i]
        previous_log = float(model_data["ln_trm"].shift(1).loc[date])
        actual_log = float(model_data.loc[date, "ln_trm"])
        rows.append(
            {
                "fecha": date,
                "ln_trm_observada": actual_log,
                "ln_trm_modelo_condicional": previous_log + forecast_change,
                "ln_trm_caminata_aleatoria": previous_log,
                "cambio_log_observado": float(y.iloc[i]),
                "cambio_log_modelo": forecast_change,
            }
        )
    pred = pd.DataFrame(rows).set_index("fecha")
    pred["trm_observada"] = np.exp(pred["ln_trm_observada"])
    pred["trm_modelo_condicional"] = np.exp(pred["ln_trm_modelo_condicional"])
    pred["trm_caminata_aleatoria"] = np.exp(pred["ln_trm_caminata_aleatoria"])

    metrics: list[dict[str, object]] = []
    for label, forecast_col in [
        ("ADL diferencias condicional", "ln_trm_modelo_condicional"),
        ("Caminata aleatoria", "ln_trm_caminata_aleatoria"),
    ]:
        errors = pred[forecast_col] - pred["ln_trm_observada"]
        direction_hit = np.nan
        if label.startswith("ADL"):
            direction_hit = float(
                (
                    np.sign(pred["cambio_log_modelo"])
                    == np.sign(pred["cambio_log_observado"])
                ).mean()
            )
        metrics.append(
            {
                "modelo": label,
                "observaciones": int(errors.shape[0]),
                "mae_log": float(errors.abs().mean()),
                "rmse_log": float(np.sqrt(np.mean(np.square(errors)))),
                "mape_pct": float(
                    100
                    * np.mean(
                        np.abs(
                            np.exp(pred[forecast_col]) - np.exp(pred["ln_trm_observada"])
                        )
                        / np.exp(pred["ln_trm_observada"])
                    )
                ),
                "acierto_direccion_pct": 100 * direction_hit,
            }
        )
    return pred, pd.DataFrame(metrics)


def bei_factor_specs(component: str) -> dict[str, dict[str, object]]:
    """Copia la especificación ampliada y sustituye únicamente el término BEI."""
    specs = {
        name: {**spec, "terminos": list(spec["terminos"])}
        for name, spec in EXPANDED_FACTOR_SPECS_4.items()
    }
    specs["Diferencial de compensación inflacionaria 5 años"]["terminos"] = [
        (component, 1)
    ]
    return specs


def bei_model_specification_comparison(
    model_data: pd.DataFrame,
    selected_expanded: SelectedDifferenceModel,
    break_date: pd.Timestamp,
) -> pd.DataFrame:
    """Compara transformaciones y calendarios BEI sobre una muestra idéntica."""
    components = difference_components(model_data)
    common_index = selected_expanded.y.index
    variants = [
        (
            "Nivel — medias separadas (referencia)",
            "Medias mensuales separadas",
            "nivel",
            "diferencial_bei_5y_pp",
            "ninguno",
        ),
        (
            "Primera diferencia — medias separadas (vigente)",
            "Medias mensuales separadas",
            "primera_diferencia",
            "D.diferencial_bei_5y_pp",
            "ninguno",
        ),
        (
            "Nivel — fechas diarias comunes",
            "Fechas diarias comunes",
            "nivel",
            "diferencial_bei_5y_comun_pp",
            "ninguno",
        ),
        (
            "Primera diferencia — fechas diarias comunes",
            "Fechas diarias comunes",
            "primera_diferencia",
            "D.diferencial_bei_5y_comun_pp",
            "ninguno",
        ),
        (
            "Nivel separado + tendencia lineal",
            "Medias mensuales separadas",
            "nivel",
            "diferencial_bei_5y_pp",
            "tendencia",
        ),
        (
            "Nivel separado + quiebre de coeficiente ZA",
            "Medias mensuales separadas",
            "nivel",
            "diferencial_bei_5y_pp",
            "quiebre_coeficiente",
        ),
    ]
    rows: list[dict[str, object]] = []
    full_trend = pd.Series(
        np.arange(len(model_data), dtype=float) / 12.0,
        index=model_data.index,
        name="tendencia_anual",
    )
    for name, aggregation, transformation, component, extension in variants:
        specs = bei_factor_specs(component)
        y, x = make_timed_difference_design(
            components,
            p=selected_expanded.p,
            factor_specs=specs,
            index=common_index,
        )
        bei_term = design_term_name(component, 1)
        if extension == "tendencia":
            x["tendencia_anual"] = full_trend.reindex(x.index)
        elif extension == "quiebre_coeficiente":
            x["post_quiebre_za"] = (x.index >= break_date).astype(float)
            x[f"{bei_term}_x_post_quiebre"] = x[bei_term] * x["post_quiebre_za"]

        result = sm.OLS(y, x).fit()
        robust, coefficients = tidy_robust_ols(result, maxlags=6)
        lookup = coefficients.set_index("termino")
        predictions, metrics = difference_validation(
            model_data,
            SelectedDifferenceModel(
                p=selected_expanded.p,
                q=0,
                result=result,
                y=y,
                x=x,
            ),
            holdout=48,
        )
        metric = metrics.loc[~metrics["modelo"].str.contains("Caminata")].iloc[0]
        model_error = (
            predictions["ln_trm_modelo_condicional"]
            - predictions["ln_trm_observada"]
        )
        benchmark_error = (
            predictions["ln_trm_caminata_aleatoria"]
            - predictions["ln_trm_observada"]
        )
        r2_validation = 1.0 - float(np.square(model_error).sum()) / float(
            np.square(benchmark_error).sum()
        )

        pre_coefficient = float(lookup.loc[bei_term, "coeficiente"])
        pre_p_value = float(lookup.loc[bei_term, "p_valor"])
        change_term = f"{bei_term}_x_post_quiebre"
        change_coefficient = math.nan
        change_p_value = math.nan
        post_coefficient = math.nan
        post_p_value = math.nan
        if change_term in lookup.index:
            change_coefficient = float(lookup.loc[change_term, "coeficiente"])
            change_p_value = float(lookup.loc[change_term, "p_valor"])
            restriction = np.zeros(len(result.params), dtype=float)
            restriction[list(x.columns).index(bei_term)] = 1.0
            restriction[list(x.columns).index(change_term)] = 1.0
            post_test = robust.t_test(restriction)
            post_coefficient = float(np.asarray(post_test.effect).reshape(-1)[0])
            post_p_value = float(np.asarray(post_test.pvalue).reshape(-1)[0])

        rows.append(
            {
                "especificacion": name,
                "agregacion_bei": aggregation,
                "transformacion_bei": transformation,
                "extension_deterministica": extension,
                "fecha_quiebre_za": (
                    break_date.strftime("%Y-%m-%d")
                    if extension == "quiebre_coeficiente"
                    else ""
                ),
                "observaciones": int(result.nobs),
                "p_cambio_trm": int(selected_expanded.p),
                "r_cuadrado_ajustado": float(result.rsquared_adj),
                "aic": float(result.aic),
                "bic": float(result.bic),
                "coeficiente_bei_pre_quiebre": pre_coefficient,
                "p_valor_hac_bei_pre_quiebre": pre_p_value,
                "cambio_coeficiente_post_quiebre": change_coefficient,
                "p_valor_hac_cambio_coeficiente": change_p_value,
                "coeficiente_bei_post_quiebre": post_coefficient,
                "p_valor_hac_bei_post_quiebre": post_p_value,
                "mape_condicional_pct": float(metric["mape_pct"]),
                "rmse_log_condicional": float(metric["rmse_log"]),
                "acierto_direccion_condicional_pct": float(
                    metric["acierto_direccion_pct"]
                ),
                "r2_validacion_condicional_vs_caminata": r2_validation,
                "quiebre_elegido_con_muestra_completa": extension
                == "quiebre_coeficiente",
            }
        )
    comparison = pd.DataFrame(rows)
    current = comparison.loc[
        comparison["especificacion"].eq(
            "Primera diferencia — medias separadas (vigente)"
        )
    ].iloc[0]
    if not np.isclose(current["bic"], selected_expanded.result.bic, atol=1e-8):
        raise AssertionError("La especificación BEI vigente no concilia con el modelo ampliado.")
    return comparison


def difference_fit_and_contributions(
    model_data: pd.DataFrame, selected: SelectedDifferenceModel
) -> tuple[pd.DataFrame, pd.DataFrame]:
    fitted = pd.DataFrame(index=selected.y.index)
    fitted["cambio_log_observado"] = selected.y
    fitted["cambio_log_ajustado"] = selected.result.fittedvalues
    fitted["ln_trm_mes_anterior"] = model_data["ln_trm"].shift(1).reindex(fitted.index)
    fitted["trm_observada"] = np.exp(model_data["ln_trm"].reindex(fitted.index))
    fitted["trm_ajustada_un_paso"] = np.exp(
        fitted["ln_trm_mes_anterior"] + fitted["cambio_log_ajustado"]
    )
    fitted["residuo_cambio_log"] = fitted["cambio_log_observado"] - fitted[
        "cambio_log_ajustado"
    ]

    contributions = selected.x.multiply(selected.result.params, axis=1)
    contributions.index.name = "fecha"
    contributions["ajuste_total"] = contributions.sum(axis=1)
    return fitted, contributions


def exact_shapley_r2(
    selected: SelectedDifferenceModel,
    factor_specs: dict[str, dict[str, object]],
    robust_coefficients: pd.DataFrame,
) -> pd.DataFrame:
    """Descompone exactamente el incremento del R2 mediante Shapley/LMG."""
    factor_columns = {
        name: [design_term_name(component, lag) for component, lag in spec["terminos"]]
        for name, spec in factor_specs.items()
    }
    assigned = {column for columns in factor_columns.values() for column in columns}
    base_columns = [column for column in selected.x.columns if column not in assigned]
    missing = assigned.difference(selected.x.columns)
    if missing:
        raise ValueError(f"Faltan terminos para Shapley: {sorted(missing)}")

    y = selected.y.to_numpy(dtype=float)
    total_ss = float(np.square(y - y.mean()).sum())

    def r_squared(columns: list[str]) -> float:
        matrix = selected.x[columns].to_numpy(dtype=float)
        beta, *_ = np.linalg.lstsq(matrix, y, rcond=None)
        residual = y - matrix @ beta
        return 1.0 - float(np.square(residual).sum()) / total_ss

    names = list(factor_specs)
    player_count = len(names)
    cache: dict[int, float] = {}
    for mask in range(1 << player_count):
        columns = list(base_columns)
        for player, factor_name in enumerate(names):
            if mask & (1 << player):
                columns.extend(factor_columns[factor_name])
        cache[mask] = r_squared(columns)

    empty_r2 = cache[0]
    full_r2 = cache[(1 << player_count) - 1]
    incremental_r2 = full_r2 - empty_r2
    coefficient_lookup = robust_coefficients.set_index("termino")
    rows: list[dict[str, object]] = []
    for player, factor_name in enumerate(names):
        shapley = 0.0
        for mask in range(1 << player_count):
            if mask & (1 << player):
                continue
            subset_size = mask.bit_count()
            weight = (
                math.factorial(subset_size)
                * math.factorial(player_count - subset_size - 1)
                / math.factorial(player_count)
            )
            shapley += weight * (cache[mask | (1 << player)] - cache[mask])

        terms = factor_columns[factor_name]
        coefficient = math.nan
        p_value = math.nan
        if len(terms) == 1 and terms[0] in coefficient_lookup.index:
            coefficient = float(coefficient_lookup.loc[terms[0], "coeficiente"])
            p_value = float(coefficient_lookup.loc[terms[0], "p_valor"])
        rows.append(
            {
                "factor": factor_name,
                "grupo": factor_specs[factor_name]["grupo"],
                "terminos": ", ".join(terms),
                "coeficiente_modelo": coefficient,
                "p_valor_hac": p_value,
                "shapley_r2": shapley,
                "aporte_r2_puntos_porcentuales": 100.0 * shapley,
                "peso_entre_factores_pct": 100.0 * shapley / incremental_r2,
                "peso_r2_total_pct": 100.0 * shapley / full_r2,
                "r2_base": empty_r2,
                "r2_completo": full_r2,
                "r2_incremental": incremental_r2,
            }
        )

    result = pd.DataFrame(rows).sort_values("shapley_r2", ascending=False).reset_index(drop=True)
    if not np.isclose(result["shapley_r2"].sum(), incremental_r2, atol=1e-10):
        raise AssertionError("La suma Shapley no cierra contra el incremento del R2.")
    if not np.isclose(result["peso_entre_factores_pct"].sum(), 100.0, atol=1e-8):
        raise AssertionError("Los pesos Shapley no suman 100%.")
    return result


def factor_columns_from_specs(
    factor_specs: dict[str, dict[str, object]],
) -> dict[str, list[str]]:
    return {
        name: [design_term_name(component, lag) for component, lag in spec["terminos"]]
        for name, spec in factor_specs.items()
    }


def moving_block_indices(
    observations: int, block_months: int, rng: np.random.Generator
) -> np.ndarray:
    """Bootstrap circular de bloques móviles para conservar dependencia mensual local."""
    blocks = math.ceil(observations / block_months)
    starts = rng.integers(0, observations, size=blocks)
    offsets = np.arange(block_months)
    return ((starts[:, None] + offsets[None, :]) % observations).ravel()[:observations]


def permutation_shapley_weights(
    y: np.ndarray,
    x: pd.DataFrame,
    factor_specs: dict[str, dict[str, object]],
    permutations: int,
    rng: np.random.Generator,
) -> np.ndarray:
    """Aproxima Shapley con permutaciones antitéticas dentro de una réplica."""
    factor_columns = factor_columns_from_specs(factor_specs)
    assigned = {column for columns in factor_columns.values() for column in columns}
    base_columns = [column for column in x.columns if column not in assigned]
    names = list(factor_specs)
    total_ss = float(np.square(y - y.mean()).sum())
    matrix = x.to_numpy(dtype=float)
    gram = matrix.T @ matrix
    cross = matrix.T @ y
    y_square = float(y @ y)
    column_positions = {column: position for position, column in enumerate(x.columns)}
    base_positions = [column_positions[column] for column in base_columns]
    factor_positions = [
        [column_positions[column] for column in factor_columns[name]] for name in names
    ]
    cache: dict[int, float] = {}

    def r_squared(mask: int) -> float:
        if mask in cache:
            return cache[mask]
        positions = list(base_positions)
        for player, columns in enumerate(factor_positions):
            if mask & (1 << player):
                positions.extend(columns)
        sub_gram = gram[np.ix_(positions, positions)]
        sub_cross = cross[positions]
        beta, *_ = np.linalg.lstsq(sub_gram, sub_cross, rcond=None)
        rss = max(0.0, y_square - float(beta @ sub_cross))
        value = 1.0 - rss / total_ss
        cache[mask] = value
        return value

    contributions = np.zeros(len(names), dtype=float)
    evaluated = 0
    for _ in range(math.ceil(permutations / 2)):
        random_order = rng.permutation(len(names))
        for order in (random_order, random_order[::-1]):
            if evaluated >= permutations:
                break
            mask = 0
            previous_r2 = r_squared(mask)
            for player in order:
                mask |= 1 << int(player)
                current_r2 = r_squared(mask)
                contributions[int(player)] += current_r2 - previous_r2
                previous_r2 = current_r2
            evaluated += 1
    contributions /= evaluated
    incremental = float(contributions.sum())
    if incremental <= 0:
        raise ValueError("El R² incremental bootstrap no es positivo.")
    return 100.0 * contributions / incremental


def block_bootstrap_shapley(
    selected: SelectedDifferenceModel,
    factor_specs: dict[str, dict[str, object]],
    point_shapley: pd.DataFrame,
) -> pd.DataFrame:
    """Intervalos percentiles de pesos Shapley mediante bootstrap por bloques."""
    rng = np.random.default_rng(SHAPLEY_BOOTSTRAP_SEED)
    bootstrap_weights: list[np.ndarray] = []
    for _ in range(SHAPLEY_BOOTSTRAP_REPLICATIONS):
        sample_positions = moving_block_indices(
            len(selected.y), SHAPLEY_BOOTSTRAP_BLOCK_MONTHS, rng
        )
        y_bootstrap = selected.y.to_numpy(dtype=float)[sample_positions]
        x_bootstrap = selected.x.iloc[sample_positions].reset_index(drop=True)
        bootstrap_weights.append(
            permutation_shapley_weights(
                y_bootstrap,
                x_bootstrap,
                factor_specs,
                SHAPLEY_BOOTSTRAP_PERMUTATIONS,
                rng,
            )
        )
    draws = np.vstack(bootstrap_weights)
    point_lookup = point_shapley.set_index("factor")
    names = list(factor_specs)
    top_three = np.argsort(-draws, axis=1)[:, :3]
    rows = []
    for player, factor in enumerate(names):
        values = draws[:, player]
        rows.append(
            {
                "factor": factor,
                "grupo": factor_specs[factor]["grupo"],
                "peso_puntual_pct": float(
                    point_lookup.loc[factor, "peso_entre_factores_pct"]
                ),
                "peso_bootstrap_media_pct": float(values.mean()),
                "peso_bootstrap_mediana_pct": float(np.median(values)),
                "ic_95_inferior_pct": float(np.quantile(values, 0.025)),
                "ic_95_superior_pct": float(np.quantile(values, 0.975)),
                "probabilidad_top3_pct": float(
                    100.0 * np.mean(np.any(top_three == player, axis=1))
                ),
                "replicas_validas": SHAPLEY_BOOTSTRAP_REPLICATIONS,
                "bloque_meses": SHAPLEY_BOOTSTRAP_BLOCK_MONTHS,
                "permutaciones_por_replica": SHAPLEY_BOOTSTRAP_PERMUTATIONS,
                "semilla": SHAPLEY_BOOTSTRAP_SEED,
            }
        )
    return pd.DataFrame(rows).sort_values("peso_puntual_pct", ascending=False).reset_index(drop=True)


def subsample_stability(
    selected: SelectedDifferenceModel,
    factor_specs: dict[str, dict[str, object]],
    point_shapley: pd.DataFrame,
    full_coefficients: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Reestima coeficientes y Shapley en cortes temporales predefinidos."""
    midpoint = len(selected.y) // 2
    masks = [
        ("Muestra completa", np.ones(len(selected.y), dtype=bool)),
        ("Primera mitad", np.arange(len(selected.y)) < midpoint),
        ("Segunda mitad", np.arange(len(selected.y)) >= midpoint),
        ("Prepandemia", selected.y.index <= pd.Timestamp("2019-12-01")),
        ("2020 en adelante", selected.y.index >= pd.Timestamp("2020-01-01")),
    ]
    full_weights = point_shapley.set_index("factor")["peso_entre_factores_pct"]
    full_signs = (
        full_coefficients.set_index("termino")["coeficiente"]
        .apply(np.sign)
        .to_dict()
    )
    factor_columns = factor_columns_from_specs(factor_specs)
    detail_rows: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    for label, mask in masks:
        y_sub = selected.y.loc[mask]
        x_sub = selected.x.loc[mask]
        result_sub = sm.OLS(y_sub, x_sub).fit()
        sub_selected = SelectedDifferenceModel(
            p=selected.p, q=selected.q, result=result_sub, y=y_sub, x=x_sub
        )
        _, coefficients_sub = tidy_robust_ols(result_sub, maxlags=6)
        if label == "Muestra completa":
            shapley_sub = point_shapley.copy()
            coefficients_sub = full_coefficients.copy()
        else:
            shapley_sub = exact_shapley_r2(
                sub_selected, factor_specs, coefficients_sub
            )
        weights_sub = shapley_sub.set_index("factor")["peso_entre_factores_pct"]
        ranks_sub = weights_sub.rank(ascending=False, method="min")
        differences = weights_sub - full_weights
        coefficient_lookup = coefficients_sub.set_index("termino")
        sign_matches = []
        for factor in factor_specs:
            term = factor_columns[factor][0]
            coefficient = float(coefficient_lookup.loc[term, "coeficiente"])
            p_value = float(coefficient_lookup.loc[term, "p_valor"])
            sign_match = bool(np.sign(coefficient) == full_signs[term])
            sign_matches.append(sign_match)
            detail_rows.append(
                {
                    "submuestra": label,
                    "inicio": y_sub.index.min().strftime("%Y-%m-%d"),
                    "fin": y_sub.index.max().strftime("%Y-%m-%d"),
                    "observaciones": len(y_sub),
                    "r2": float(result_sub.rsquared),
                    "r2_ajustado": float(result_sub.rsquared_adj),
                    "factor": factor,
                    "grupo": factor_specs[factor]["grupo"],
                    "coeficiente": coefficient,
                    "p_valor_hac": p_value,
                    "shapley_r2": float(shapley_sub.set_index("factor").loc[factor, "shapley_r2"]),
                    "peso_entre_factores_pct": float(weights_sub[factor]),
                    "rango_peso": int(ranks_sub[factor]),
                    "signo_coincide_muestra_completa": sign_match,
                    "diferencia_peso_vs_completa_pp": float(differences[factor]),
                }
            )
        full_ranks = full_weights.rank(ascending=False, method="min")
        rank_correlation = float(
            stats.spearmanr(
                full_ranks.to_numpy(),
                ranks_sub.reindex(full_ranks.index).to_numpy(),
            ).statistic
        )
        summary_rows.append(
            {
                "submuestra": label,
                "inicio": y_sub.index.min().strftime("%Y-%m-%d"),
                "fin": y_sub.index.max().strftime("%Y-%m-%d"),
                "observaciones": len(y_sub),
                "r2_ajustado": float(result_sub.rsquared_adj),
                "correlacion_spearman_rangos_vs_completa": rank_correlation,
                "mediana_diferencia_abs_peso_pp": float(differences.abs().median()),
                "max_diferencia_abs_peso_pp": float(differences.abs().max()),
                "factores_mismo_signo_de_12": int(sum(sign_matches)),
            }
        )
    return pd.DataFrame(detail_rows), pd.DataFrame(summary_rows)


def bounds_to_frames(bounds_result) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary = pd.DataFrame(
        [
            {
                "estadistico_f": float(bounds_result.stat),
                "p_valor_i0": float(bounds_result.p_values.loc["lower"]),
                "p_valor_i1": float(bounds_result.p_values.loc["upper"]),
            }
        ]
    )
    critical = bounds_result.crit_vals.reset_index().rename(columns={"index": "percentil"})
    return summary, critical


def main() -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    DATA.mkdir(parents=True, exist_ok=True)

    data = build_dataset()
    model_columns = [
        "ln_trm",
        *ECM_LEVEL_VARIABLES,
        "ln_vix",
        "dln_vix",
        "embig_colombia_pp",
        "ln_reservas_netas_sin_flar",
        "asinh_balanza_comercial",
        "asinh_flujos_capital",
        *LEVEL_COMPONENTS,
        "dummy_pandemia_2020",
    ]
    expected_index = pd.date_range(SAMPLE_START, SAMPLE_END, freq="MS")
    model_data = data.reindex(expected_index)[model_columns].copy()
    if model_data.isna().any().any():
        missing = {
            column: model_data.index[model_data[column].isna()].strftime("%Y-%m").tolist()
            for column in model_data.columns
            if model_data[column].isna().any()
        }
        raise ValueError(f"La muestra balanceada contiene meses faltantes: {missing}")
    model_data.index.name = "fecha"
    bei_stationarity = bei_stationarity_tests(model_data)
    bei_trend_breaks = bei_trend_break_models(model_data, bei_stationarity)
    bei_break_date = pd.Timestamp(
        bei_stationarity.loc[
            bei_stationarity["agregacion"].eq("Medias mensuales separadas")
            & bei_stationarity["transformacion"].eq("nivel")
            & bei_stationarity["prueba"].eq("Zivot-Andrews")
            & bei_stationarity["deterministico"].eq(
                "constante_tendencia_con_quiebre"
            ),
            "fecha_quiebre",
        ].iloc[0]
    )

    components = difference_components(model_data)
    common_index = make_timed_difference_design(
        components, p=3, factor_specs=EXPANDED_FACTOR_SPECS
    )[0].index

    selected_diff, lag_grid_diff = select_timed_difference_model(
        model_data, BASE_FACTOR_SPECS, common_index=common_index
    )
    _, coefficients_diff = tidy_robust_ols(selected_diff.result, maxlags=6)
    diagnostics_diff = diagnostics(selected_diff.result)
    predictions, validation = difference_validation(
        model_data, selected_diff, holdout=min(48, len(selected_diff.y) // 4)
    )
    fitted_diff, contributions_diff = difference_fit_and_contributions(
        model_data, selected_diff
    )

    selected_expanded_3, _ = select_timed_difference_model(
        model_data, EXPANDED_FACTOR_SPECS_3, common_index=common_index
    )
    _, coefficients_expanded_3 = tidy_robust_ols(selected_expanded_3.result, maxlags=6)
    predictions_expanded_3, validation_expanded_3 = difference_validation(
        model_data, selected_expanded_3, holdout=min(48, len(selected_expanded_3.y) // 4)
    )

    selected_expanded, lag_grid_expanded = select_timed_difference_model(
        model_data, EXPANDED_FACTOR_SPECS_4, common_index=common_index
    )
    _, coefficients_expanded = tidy_robust_ols(selected_expanded.result, maxlags=6)
    diagnostics_expanded = diagnostics(selected_expanded.result)
    bei_model_comparison = bei_model_specification_comparison(
        model_data, selected_expanded, bei_break_date
    )
    predictions_expanded, validation_expanded = difference_validation(
        model_data, selected_expanded, holdout=min(48, len(selected_expanded.y) // 4)
    )
    fitted_expanded, contributions_expanded = difference_fit_and_contributions(
        model_data, selected_expanded
    )
    shapley_expanded = exact_shapley_r2(
        selected_expanded, EXPANDED_FACTOR_SPECS_4, coefficients_expanded
    )
    shapley_bootstrap = block_bootstrap_shapley(
        selected_expanded, EXPANDED_FACTOR_SPECS_4, shapley_expanded
    )
    stability_detail, stability_summary = subsample_stability(
        selected_expanded,
        EXPANDED_FACTOR_SPECS_4,
        shapley_expanded,
        coefficients_expanded,
    )

    forecast_common_index = make_timed_difference_design(
        components, p=3, factor_specs=FORECAST_FACTOR_SPECS_4
    )[0].index
    selected_forecast_3, lag_grid_forecast_3 = select_timed_difference_model(
        model_data, FORECAST_FACTOR_SPECS_3, common_index=forecast_common_index
    )
    _, coefficients_forecast_3 = tidy_robust_ols(selected_forecast_3.result, maxlags=6)
    diagnostics_forecast_3 = diagnostics(selected_forecast_3.result)
    predictions_forecast_3, validation_forecast_3 = difference_validation(
        model_data, selected_forecast_3, holdout=min(48, len(selected_forecast_3.y) // 4)
    )

    selected_forecast_4, lag_grid_forecast_4 = select_timed_difference_model(
        model_data, FORECAST_FACTOR_SPECS_4, common_index=forecast_common_index
    )
    _, coefficients_forecast_4 = tidy_robust_ols(selected_forecast_4.result, maxlags=6)
    diagnostics_forecast_4 = diagnostics(selected_forecast_4.result)
    predictions_forecast_4, validation_forecast_4 = difference_validation(
        model_data, selected_forecast_4, holdout=min(48, len(selected_forecast_4.y) // 4)
    )

    # La composición del pronóstico se elige por BIC, antes de mirar la métrica
    # de la ventana de validación. La explicación histórica conserva cuatro monedas.
    if selected_forecast_3.result.bic <= selected_forecast_4.result.bic:
        forecast_currencies = "BRL, CLP y MXN"
        selected_forecast = selected_forecast_3
        lag_grid_forecast = lag_grid_forecast_3
        coefficients_forecast = coefficients_forecast_3
        diagnostics_forecast = diagnostics_forecast_3
        predictions_forecast = predictions_forecast_3.copy()
        validation_forecast = validation_forecast_3.copy()
    else:
        forecast_currencies = "BRL, CLP, MXN y PEN"
        selected_forecast = selected_forecast_4
        lag_grid_forecast = lag_grid_forecast_4
        coefficients_forecast = coefficients_forecast_4
        diagnostics_forecast = diagnostics_forecast_4
        predictions_forecast = predictions_forecast_4.copy()
        validation_forecast = validation_forecast_4.copy()

    def out_of_sample_r2(
        predictions_frame: pd.DataFrame,
        forecast_column: str = "ln_trm_modelo_condicional",
    ) -> float:
        model_error = (
            predictions_frame[forecast_column]
            - predictions_frame["ln_trm_observada"]
        )
        benchmark_error = (
            predictions_frame["ln_trm_caminata_aleatoria"]
            - predictions_frame["ln_trm_observada"]
        )
        return 1.0 - float(np.square(model_error).sum()) / float(
            np.square(benchmark_error).sum()
        )

    base_validation_row = validation.iloc[0]
    expanded_validation_row = validation_expanded.iloc[0]
    comparison = pd.DataFrame(
        [
            {
                "modelo": "Base",
                "observaciones": int(selected_diff.result.nobs),
                "r_cuadrado": float(selected_diff.result.rsquared),
                "r_cuadrado_ajustado": float(selected_diff.result.rsquared_adj),
                "aic": float(selected_diff.result.aic),
                "bic": float(selected_diff.result.bic),
                "mape_pct": float(base_validation_row["mape_pct"]),
                "acierto_direccion_pct": float(
                    base_validation_row["acierto_direccion_pct"]
                ),
                "r2_validacion_condicional_vs_caminata": out_of_sample_r2(predictions),
            },
            {
                "modelo": "Ampliado historico",
                "observaciones": int(selected_expanded.result.nobs),
                "r_cuadrado": float(selected_expanded.result.rsquared),
                "r_cuadrado_ajustado": float(selected_expanded.result.rsquared_adj),
                "aic": float(selected_expanded.result.aic),
                "bic": float(selected_expanded.result.bic),
                "mape_pct": float(expanded_validation_row["mape_pct"]),
                "acierto_direccion_pct": float(
                    expanded_validation_row["acierto_direccion_pct"]
                ),
                "r2_validacion_condicional_vs_caminata": out_of_sample_r2(
                    predictions_expanded
                ),
            },
        ]
    )

    regional_correlation = float(
        model_data["factor_monedas_regionales_3"].corr(
            model_data["factor_monedas_regionales_4"]
        )
    )

    def coefficient_value(table: pd.DataFrame, term: str, column: str) -> float:
        return float(table.loc[table["termino"].eq(term), column].iloc[0])

    regional_comparison_rows: list[dict[str, object]] = []
    regional_variants = [
        (
            "Explicación histórica",
            "BRL, CLP y MXN",
            selected_expanded_3,
            coefficients_expanded_3,
            predictions_expanded_3,
            validation_expanded_3,
            "factor_monedas_regionales_3.L0",
        ),
        (
            "Explicación histórica",
            "BRL, CLP, MXN y PEN",
            selected_expanded,
            coefficients_expanded,
            predictions_expanded,
            validation_expanded,
            "factor_monedas_regionales_4.L0",
        ),
        (
            "Pronóstico con rezagos de publicación",
            "BRL, CLP y MXN",
            selected_forecast_3,
            coefficients_forecast_3,
            predictions_forecast_3,
            validation_forecast_3,
            "factor_monedas_regionales_3.L1",
        ),
        (
            "Pronóstico con rezagos de publicación",
            "BRL, CLP, MXN y PEN",
            selected_forecast_4,
            coefficients_forecast_4,
            predictions_forecast_4,
            validation_forecast_4,
            "factor_monedas_regionales_4.L1",
        ),
    ]
    for use, currencies, selected_variant, coefficient_table, predictions_variant, validation_variant, term in regional_variants:
        metric = validation_variant.loc[
            ~validation_variant["modelo"].str.contains("Caminata", case=False)
        ].iloc[0]
        regional_comparison_rows.append(
            {
                "uso": use,
                "monedas": currencies,
                "observaciones": int(selected_variant.result.nobs),
                "p_cambio_trm": int(selected_variant.p),
                "r_cuadrado": float(selected_variant.result.rsquared),
                "r_cuadrado_ajustado": float(selected_variant.result.rsquared_adj),
                "aic": float(selected_variant.result.aic),
                "bic": float(selected_variant.result.bic),
                "mape_pct": float(metric["mape_pct"]),
                "acierto_direccion_pct": float(metric["acierto_direccion_pct"]),
                "r2_validacion_vs_caminata": out_of_sample_r2(predictions_variant),
                "coeficiente_factor_regional": coefficient_value(
                    coefficient_table, term, "coeficiente"
                ),
                "p_valor_hac_factor_regional": coefficient_value(
                    coefficient_table, term, "p_valor"
                ),
                "correlacion_factores_3_4": regional_correlation,
            }
        )
    regional_comparison = pd.DataFrame(regional_comparison_rows)

    availability = pd.DataFrame(
        FORECAST_AVAILABILITY,
        columns=[
            "factor",
            "rezago_meses_modelo",
            "frecuencia_y_publicacion",
            "regla_disponibilidad_al_inicio_del_mes_t",
        ],
    )

    validation_forecast.loc[
        ~validation_forecast["modelo"].str.contains("Caminata", case=False), "modelo"
    ] = "Pronóstico con rezagos de publicación"
    predictions_forecast = predictions_forecast.rename(
        columns={
            "ln_trm_modelo_condicional": "ln_trm_pronostico_publicacion",
            "cambio_log_modelo": "cambio_log_pronostico",
            "trm_modelo_condicional": "trm_pronostico_publicacion",
        }
    )

    y = model_data["ln_trm"]
    exog = model_data[ECM_LEVEL_VARIABLES]
    fixed = model_data[["dln_vix", "dummy_pandemia_2020"]]
    selected_ecm, lag_grid_ecm = select_ardl(y, exog, fixed)
    uecm_model = UECM.from_ardl(selected_ecm.result.model)
    uecm_result = uecm_model.fit(
        cov_type="HAC", cov_kwds={"maxlags": 6, "use_correction": True}, use_t=True
    )
    bounds = uecm_result.bounds_test(case=3, cov_type="nonrobust")
    bounds_summary, bounds_critical = bounds_to_frames(bounds)

    tests = integration_tests(
        model_data,
        [
            "ln_trm",
            *ECM_LEVEL_VARIABLES,
            "ln_vix",
            "embig_colombia_pp",
            "ln_reservas_netas_sin_flar",
            "asinh_balanza_comercial",
            "asinh_flujos_capital",
            "diferencial_bei_5y_pp",
            "diferencial_bei_5y_comun_pp",
        ],
    )
    short_run_ecm = tidy_result(uecm_result)
    long_run_ecm = tidy_long_run(uecm_result)
    diagnostics_ecm = diagnostics(selected_ecm.result)

    data.to_csv(DATA / "modelo_trm_datos_mensuales.csv", encoding="utf-8-sig", float_format="%.10g")
    model_data.to_csv(DATA / "modelo_trm_muestra_estimacion.csv", encoding="utf-8-sig", float_format="%.10g")
    bei_aggregation_columns = [
        "tes_5y_pesos_colombia_pct",
        "tes_5y_uvr_colombia_pct",
        "bei_eeuu_5y_pct",
        "bei_colombia_5y_pct",
        "diferencial_bei_5y_pp",
        "tes_5y_pesos_comun_pct",
        "tes_5y_uvr_comun_pct",
        "bei_eeuu_5y_comun_pct",
        "diferencial_bei_5y_comun_pp",
        "diferencia_comun_menos_separada_pp",
        "dias_tes_pesos",
        "dias_tes_uvr",
        "dias_bei_eeuu",
        "dias_comunes",
    ]
    data.loc[SAMPLE_START:SAMPLE_END, bei_aggregation_columns].to_csv(
        RESULTS / "comparacion_agregacion_bei_5y.csv",
        encoding="utf-8-sig",
        float_format="%.10g",
    )
    bei_stationarity.to_csv(
        RESULTS / "pruebas_estacionariedad_bei_5y.csv",
        index=False,
        encoding="utf-8-sig",
        float_format="%.10g",
    )
    bei_trend_breaks.to_csv(
        RESULTS / "tendencias_quiebres_bei_5y.csv",
        index=False,
        encoding="utf-8-sig",
        float_format="%.10g",
    )
    bei_model_comparison.to_csv(
        RESULTS / "comparacion_especificaciones_bei_5y.csv",
        index=False,
        encoding="utf-8-sig",
        float_format="%.10g",
    )
    lag_grid_diff.to_csv(
        RESULTS / "seleccion_rezagos_adl_diferencias.csv", index=False, encoding="utf-8-sig"
    )
    coefficients_diff.to_csv(
        RESULTS / "coeficientes_modelo_principal.csv", index=False, encoding="utf-8-sig"
    )
    diagnostics_diff.to_csv(
        RESULTS / "diagnosticos_modelo_principal.csv", index=False, encoding="utf-8-sig"
    )
    fitted_diff.to_csv(
        RESULTS / "ajuste_historico_modelo_principal.csv", encoding="utf-8-sig"
    )
    contributions_diff.to_csv(
        RESULTS / "contribuciones_modelo_principal.csv", encoding="utf-8-sig"
    )
    lag_grid_expanded.to_csv(
        RESULTS / "seleccion_rezagos_modelo_ampliado.csv",
        index=False,
        encoding="utf-8-sig",
    )
    coefficients_expanded.to_csv(
        RESULTS / "coeficientes_modelo_ampliado.csv", index=False, encoding="utf-8-sig"
    )
    diagnostics_expanded.to_csv(
        RESULTS / "diagnosticos_modelo_ampliado.csv", index=False, encoding="utf-8-sig"
    )
    fitted_expanded.to_csv(
        RESULTS / "ajuste_historico_modelo_ampliado.csv", encoding="utf-8-sig"
    )
    contributions_expanded.to_csv(
        RESULTS / "contribuciones_modelo_ampliado.csv", encoding="utf-8-sig"
    )
    shapley_expanded.to_csv(
        RESULTS / "pesos_explicativos_modelo_ampliado.csv",
        index=False,
        encoding="utf-8-sig",
    )
    shapley_bootstrap.to_csv(
        RESULTS / "intervalos_bootstrap_pesos_shapley.csv",
        index=False,
        encoding="utf-8-sig",
    )
    stability_detail.to_csv(
        RESULTS / "estabilidad_submuestras_modelo_ampliado.csv",
        index=False,
        encoding="utf-8-sig",
    )
    stability_summary.to_csv(
        RESULTS / "estabilidad_submuestras_resumen.csv",
        index=False,
        encoding="utf-8-sig",
    )
    comparison.to_csv(
        RESULTS / "comparacion_modelos.csv", index=False, encoding="utf-8-sig"
    )
    regional_comparison.to_csv(
        RESULTS / "comparacion_factor_regional.csv", index=False, encoding="utf-8-sig"
    )
    availability.to_csv(
        RESULTS / "calendario_disponibilidad_pronostico.csv",
        index=False,
        encoding="utf-8-sig",
    )
    lag_grid_forecast.to_csv(
        RESULTS / "seleccion_rezagos_modelo_pronostico.csv",
        index=False,
        encoding="utf-8-sig",
    )
    coefficients_forecast.to_csv(
        RESULTS / "coeficientes_modelo_pronostico.csv",
        index=False,
        encoding="utf-8-sig",
    )
    diagnostics_forecast.to_csv(
        RESULTS / "diagnosticos_modelo_pronostico.csv",
        index=False,
        encoding="utf-8-sig",
    )
    predictions_forecast.to_csv(
        RESULTS / "validacion_predicciones_pronostico.csv", encoding="utf-8-sig"
    )
    validation_forecast.to_csv(
        RESULTS / "validacion_metricas_pronostico.csv",
        index=False,
        encoding="utf-8-sig",
    )
    lag_grid_ecm.to_csv(
        RESULTS / "seleccion_rezagos_ecm.csv", index=False, encoding="utf-8-sig"
    )
    tests.to_csv(RESULTS / "pruebas_integracion.csv", index=False, encoding="utf-8-sig")
    short_run_ecm.to_csv(
        RESULTS / "coeficientes_corto_plazo_ecm.csv", index=False, encoding="utf-8-sig"
    )
    long_run_ecm.to_csv(
        RESULTS / "coeficientes_largo_plazo_ecm.csv", index=False, encoding="utf-8-sig"
    )
    bounds_summary.to_csv(RESULTS / "bounds_resumen.csv", index=False, encoding="utf-8-sig")
    bounds_critical.to_csv(RESULTS / "bounds_criticos.csv", index=False, encoding="utf-8-sig")
    diagnostics_ecm.to_csv(
        RESULTS / "diagnosticos_ecm.csv", index=False, encoding="utf-8-sig"
    )
    predictions.to_csv(RESULTS / "validacion_predicciones_modelo_principal.csv", encoding="utf-8-sig")
    validation.to_csv(RESULTS / "validacion_metricas_modelo_principal.csv", index=False, encoding="utf-8-sig")
    predictions_expanded.to_csv(
        RESULTS / "validacion_predicciones_modelo_ampliado.csv", encoding="utf-8-sig"
    )
    validation_expanded.to_csv(
        RESULTS / "validacion_metricas_modelo_ampliado.csv",
        index=False,
        encoding="utf-8-sig",
    )

    bei_level_test = tests.loc[
        (tests["variable"] == "diferencial_bei_5y_pp")
        & (tests["transformacion"] == "nivel")
    ].iloc[0]
    bounds_p_i0 = float(bounds.p_values.loc["lower"])
    bounds_p_i1 = float(bounds.p_values.loc["upper"])
    if bounds_p_i1 < 0.05:
        cointegration_5pct = "evidencia de cointegracion"
    elif bounds_p_i0 > 0.05:
        cointegration_5pct = "sin evidencia de cointegracion"
    else:
        cointegration_5pct = "no concluyente"

    vintage_coverage = pd.read_csv(RESULTS / "cobertura_vintages_pronostico.csv")
    complete_vintage_factors = int(
        vintage_coverage["apto_backtest_genuino"].astype("string").str.lower().eq("true").sum()
    )
    bootstrap_widest = shapley_bootstrap.assign(
        ancho=lambda frame: frame["ic_95_superior_pct"] - frame["ic_95_inferior_pct"]
    ).sort_values("ancho", ascending=False).iloc[0]
    recent_stability = stability_summary.loc[
        stability_summary["submuestra"].eq("2020 en adelante")
    ].iloc[0]
    bei_best_bic = bei_model_comparison.sort_values("bic").iloc[0]
    bei_best_validation = bei_model_comparison.sort_values(
        "rmse_log_condicional"
    ).iloc[0]
    bei_aggregation_sample = data.loc[
        SAMPLE_START:SAMPLE_END,
        [
            "diferencial_bei_5y_pp",
            "diferencial_bei_5y_comun_pp",
            "diferencia_comun_menos_separada_pp",
            "dias_comunes",
        ],
    ].dropna()
    bei_adf_trend = bei_stationarity.loc[
        bei_stationarity["agregacion"].eq("Medias mensuales separadas")
        & bei_stationarity["transformacion"].eq("nivel")
        & bei_stationarity["prueba"].eq("ADF")
        & bei_stationarity["deterministico"].eq("constante_tendencia")
    ].iloc[0]
    bei_za_trend = bei_stationarity.loc[
        bei_stationarity["agregacion"].eq("Medias mensuales separadas")
        & bei_stationarity["transformacion"].eq("nivel")
        & bei_stationarity["prueba"].eq("Zivot-Andrews")
        & bei_stationarity["deterministico"].eq(
            "constante_tendencia_con_quiebre"
        )
    ].iloc[0]

    metadata = {
        "muestra_inicio": model_data.index.min().strftime("%Y-%m-%d"),
        "muestra_fin": model_data.index.max().strftime("%Y-%m-%d"),
        "observaciones": int(model_data.shape[0]),
        "modelo_principal": "Modelo mensual en primeras diferencias con temporización económica y errores HAC",
        "adl_p_cambio_trm": selected_diff.p,
        "temporizacion": "Términos de intercambio, dólar amplio y VIX contemporáneos; remesas, diferencial de tasas y déficit rezagados un mes",
        "adl_observaciones": int(selected_diff.result.nobs),
        "adl_aic": float(selected_diff.result.aic),
        "adl_bic": float(selected_diff.result.bic),
        "adl_r_cuadrado": float(selected_diff.result.rsquared),
        "adl_r_cuadrado_ajustado": float(selected_diff.result.rsquared_adj),
        "modelo_ampliado": "Contabilidad historica mensual en primeras diferencias con 12 factores, cuatro monedas regionales y errores HAC",
        "ampliado_p_cambio_trm": selected_expanded.p,
        "ampliado_observaciones": int(selected_expanded.result.nobs),
        "ampliado_aic": float(selected_expanded.result.aic),
        "ampliado_bic": float(selected_expanded.result.bic),
        "ampliado_r_cuadrado": float(selected_expanded.result.rsquared),
        "ampliado_r_cuadrado_ajustado": float(
            selected_expanded.result.rsquared_adj
        ),
        "ampliado_temporizacion": "Términos de intercambio, dólar amplio, VIX, EMBIG Colombia y monedas regionales contemporáneos; cambios de remesas, tasas, déficit, reservas, balanza, flujos de capital y diferencial BEI rezagados un mes",
        "pesos_metodo": "Shapley/LMG exacto del incremento del R2 sobre intercepto, dinamica de TRM y dummy de pandemia",
        "pesos_suma_pct": float(shapley_expanded["peso_entre_factores_pct"].sum()),
        "shapley_r2_base": float(shapley_expanded["r2_base"].iloc[0]),
        "shapley_r2_completo": float(shapley_expanded["r2_completo"].iloc[0]),
        "shapley_r2_incremental": float(
            shapley_expanded["r2_incremental"].iloc[0]
        ),
        "shapley_bootstrap_metodo": "Bootstrap circular de bloques mensuales; pesos de cada réplica aproximados con permutaciones antitéticas",
        "shapley_bootstrap_replicas": SHAPLEY_BOOTSTRAP_REPLICATIONS,
        "shapley_bootstrap_bloque_meses": SHAPLEY_BOOTSTRAP_BLOCK_MONTHS,
        "shapley_bootstrap_permutaciones": SHAPLEY_BOOTSTRAP_PERMUTATIONS,
        "shapley_bootstrap_semilla": SHAPLEY_BOOTSTRAP_SEED,
        "shapley_bootstrap_factor_intervalo_mas_ancho": str(bootstrap_widest["factor"]),
        "shapley_bootstrap_intervalo_mas_ancho_pp": float(bootstrap_widest["ancho"]),
        "estabilidad_submuestras_cortes": int(len(stability_summary)),
        "estabilidad_2020_spearman_rangos": float(
            recent_stability["correlacion_spearman_rangos_vs_completa"]
        ),
        "estabilidad_2020_factores_mismo_signo_de_12": int(
            recent_stability["factores_mismo_signo_de_12"]
        ),
        "factor_regional": "Modelo activo: promedio de cambios log estandarizados de BRL, CLP, MXN y PEN por USD; comparación contra BRL, CLP y MXN; parámetros calibrados 2006-2019",
        "factor_regional_correlacion_3_4": regional_correlation,
        "pronostico_modelo": f"Modelo mensual de un paso con todos los factores rezagados conforme a un calendario conservador de disponibilidad al inicio del mes objetivo; composición regional seleccionada por BIC: {forecast_currencies}",
        "pronostico_factor_regional_monedas": forecast_currencies,
        "pronostico_advertencia_vintages": "El backtest respeta rezagos de publicación, pero usa la última versión disponible de las series. Es pseudo-tiempo-real hasta contar con vintages históricos archivados; no debe rotularse como backtest genuino en tiempo real.",
        "vintages_archivo_inicio": "2026-08-23",
        "vintages_origenes_alfred_recuperados": 0,
        "vintages_factores_completos_de_12": complete_vintage_factors,
        "backtest_genuino_disponible": complete_vintage_factors == len(FORECAST_FACTOR_SPECS_3),
        "pronostico_p_cambio_trm": selected_forecast.p,
        "pronostico_observaciones": int(selected_forecast.result.nobs),
        "pronostico_r_cuadrado": float(selected_forecast.result.rsquared),
        "pronostico_r_cuadrado_ajustado": float(selected_forecast.result.rsquared_adj),
        "pronostico_aic": float(selected_forecast.result.aic),
        "pronostico_bic": float(selected_forecast.result.bic),
        "pronostico_mape_pct": float(validation_forecast.iloc[0]["mape_pct"]),
        "pronostico_acierto_direccion_pct": float(
            validation_forecast.iloc[0]["acierto_direccion_pct"]
        ),
        "pronostico_r2_vs_caminata": out_of_sample_r2(
            predictions_forecast, "ln_trm_pronostico_publicacion"
        ),
        "terminos_intercambio": "BanRep serie 15360; índice encadenado mensual, base geométrica 2000=100",
        "riesgo_soberano": "EMBIG Colombia del BCRP; promedio mensual de puntos base y conversión a puntos porcentuales",
        "bei_colombia_5y": "Diferencia entre promedios mensuales separados de TES COP 5 años BanRep 15273 y TES UVR 5 años BanRep 15276",
        "bei_eeuu_5y": "Federal Reserve Board Gürkaynak-Sack-Wright BKEVEN05; compensación inflacionaria cero cupón a 5 años, capitalización continua, promedio mensual",
        "bei_advertencia": "El BEI es compensación inflacionaria y no una expectativa pura: incorpora primas de riesgo de inflación y diferencias de liquidez",
        "proxies_snapshot_fecha_descarga": "2026-08-23",
        "proxies_snapshot_sha256": {
            filename: sha256_file(RAW / filename)
            for filename in [
                "embig_colombia_diario_bcrp.json",
                "tes_5y_pesos_banrep.json",
                "tes_5y_uvr_banrep.json",
                "bei_5y_eeuu_diario_fed.csv",
                "pen_usd_mensual_bcrp.json",
            ]
        },
        "diferencial_bei_5y_transformacion": "Modelo vigente: primera diferencia rezagada un mes y promedios mensuales separados; nivel, fechas comunes, tendencia y quiebre se reportan como robustez",
        "diferencial_bei_5y_advertencia_estacionariedad": "Las conclusiones cambian al permitir tendencia o quiebre; ninguna prueba aislada determina la transformación económica correcta",
        "diferencial_bei_5y_adf_p_nivel": float(bei_level_test["adf_p"]),
        "diferencial_bei_5y_kpss_p_nivel": float(bei_level_test["kpss_p"]),
        "diferencial_bei_5y_adf_p_nivel_con_tendencia": float(
            bei_adf_trend["p_valor"]
        ),
        "diferencial_bei_5y_za_p_nivel_con_tendencia_quiebre": float(
            bei_za_trend["p_valor"]
        ),
        "diferencial_bei_5y_quiebre_za": str(bei_za_trend["fecha_quiebre"]),
        "diferencial_bei_5y_correlacion_agregaciones": float(
            bei_aggregation_sample["diferencial_bei_5y_pp"].corr(
                bei_aggregation_sample["diferencial_bei_5y_comun_pp"]
            )
        ),
        "diferencial_bei_5y_diferencia_media_comun_menos_separada_pp": float(
            bei_aggregation_sample["diferencia_comun_menos_separada_pp"].mean()
        ),
        "diferencial_bei_5y_max_diferencia_abs_agregacion_pp": float(
            bei_aggregation_sample[
                "diferencia_comun_menos_separada_pp"
            ].abs().max()
        ),
        "diferencial_bei_5y_min_dias_comunes_mes": int(
            bei_aggregation_sample["dias_comunes"].min()
        ),
        "diferencial_bei_5y_mejor_bic_especificacion": str(
            bei_best_bic["especificacion"]
        ),
        "diferencial_bei_5y_mejor_validacion_especificacion": str(
            bei_best_validation["especificacion"]
        ),
        "flujos_capital": "Movimientos netos de capital de la balanza cambiaria, BanRep serie 16706",
        "validacion_base_mape_pct": float(base_validation_row["mape_pct"]),
        "validacion_ampliado_mape_pct": float(expanded_validation_row["mape_pct"]),
        "validacion_base_acierto_direccion_pct": float(
            base_validation_row["acierto_direccion_pct"]
        ),
        "validacion_ampliado_acierto_direccion_pct": float(
            expanded_validation_row["acierto_direccion_pct"]
        ),
        "ecm_p": selected_ecm.p,
        "ecm_q_comun": selected_ecm.q,
        "bounds_f": float(bounds.stat),
        "bounds_p_i0": bounds_p_i0,
        "bounds_p_i1": bounds_p_i1,
        "cointegracion_5pct": cointegration_5pct,
        "velocidad_ajuste": float(uecm_result.params.get("ln_trm.L1", np.nan)),
    }
    (RESULTS / "metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(json.dumps(metadata, indent=2, ensure_ascii=False))
    print("\nCoeficientes del modelo principal en diferencias")
    print(coefficients_diff.to_string(index=False))
    print("\nDiagnósticos del modelo principal")
    print(diagnostics_diff.to_string(index=False))
    print("\nValidación")
    print(validation.to_string(index=False))
    print("\nComparacion base vs. ampliado")
    print(comparison.to_string(index=False))
    print("\nPesos explicativos Shapley del modelo ampliado")
    print(
        shapley_expanded[
            ["factor", "grupo", "shapley_r2", "peso_entre_factores_pct"]
        ].to_string(index=False)
    )
    print("\nComparación del factor regional de tres y cuatro monedas")
    print(regional_comparison.to_string(index=False))
    print("\nValidación del pronóstico con rezagos de publicación")
    print(validation_forecast.to_string(index=False))
    print("\nECM exploratorio: coeficientes de largo plazo")
    print(long_run_ecm.to_string(index=False))


if __name__ == "__main__":
    main()
