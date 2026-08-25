from __future__ import annotations

"""Sincroniza el workbook versionado sin depender de @oai/artifact-tool.

El generador de workbook usa ``src/build_workbook.mjs`` y el paquete privado
``@oai/artifact-tool``. Este fallback conserva el workbook existente, pero
reemplaza las tablas auditables con los CSV actuales para que el entregable
pueda reconstruirse en entornos que solo tienen Python y openpyxl.
"""

from copy import copy
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import range_boundaries


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
DATA = ROOT / "data"
WORKBOOK_PATH = ROOT / "deliverables" / "modelo_trm_colombia.xlsx"


TERM_LABELS = {
    "const": "Intercepto",
    "D.ln_terminos_intercambio.L0": "Δ ln términos de intercambio (t)",
    "D.ln_dolar_amplio.L0": "Δ ln índice dólar amplio (t)",
    "D.ln_vix.L0": "Δ ln VIX (t)",
    "D.ln_remesas_12m.L1": "Δ ln remesas 12m (t−1)",
    "D.diferencial_tasas_pp.L1": "Δ diferencial tasas (t−1)",
    "D.deficit_fiscal_12m_pct_pib.L1": "Δ déficit fiscal 12m/PIB (t−1)",
    "D.embig_colombia_pp.L0": "Δ EMBIG Colombia (pp, t)",
    "D.ln_reservas_netas_sin_flar.L1": "Δ ln reservas netas sin FLAR (t−1)",
    "D.asinh_balanza_comercial.L1": "Δ asinh balanza comercial (t−1)",
    "D.asinh_flujos_capital.L1": "Δ asinh flujos de capital (t−1)",
    "D.diferencial_bei_5y_pp.L1": "Δ diferencial BEI 5 años (t−1)",
    "factor_monedas_regionales_4.L0": "Factor regional BRL, CLP, MXN y PEN (t)",
    "factor_monedas_regionales_3.L1": "Factor regional BRL, CLP y MXN (t−1)",
    "D.yield_real_10y_tips_pct.L0": "Δ rendimiento real TIPS 10 años (t)",
    "D.yield_real_5y_us_pct.L0": "Δ rendimiento real EE. UU. 5 años (t)",
    "D.yield_2y_us_pct.L0": "Δ Treasury EE. UU. 2 años (t)",
    "D.yield_10y_us_pct.L0": "Δ Treasury EE. UU. 10 años (t)",
    "D.spread_10y_2y_us_pct.L0": "Δ pendiente Treasury 10Y−2Y (t)",
    "D.breakeven_5y_us_pct.L0": "Δ BEI EE. UU. 5 años (t)",
    "D.breakeven_10y_us_pct.L0": "Δ BEI EE. UU. 10 años (t)",
    "D.epu_global.L0": "Δ incertidumbre de política económica global (t)",
    "D.estres_financiero_stl.L0": "Δ estrés financiero STLFSI (t)",
    "D.nfci_chicago.L0": "Δ NFCI Chicago (t)",
    "D.anfci_chicago.L0": "Δ ANFCI Chicago (t)",
    "D.ln_brent_global.L0": "Δ ln Brent global (t)",
    "D.ln_commodities_global.L0": "Δ ln commodities globales (t)",
    "D.desempleo_us_pct.L0": "Δ desempleo EE. UU. (t)",
    "D.ln_empleo_manufactura_us.L0": "Δ ln empleo manufacturero EE. UU. (t)",
    "D.ln_produccion_industrial_us.L0": "Δ ln producción industrial EE. UU. (t)",
    "D.ln_fletes_transporte_us.L0": "Δ ln fletes/logística EE. UU. (t)",
    "D.yield_real_10y_tips_pct.L1": "Δ rendimiento real TIPS 10 años (t−1)",
    "D.yield_real_5y_us_pct.L1": "Δ rendimiento real EE. UU. 5 años (t−1)",
    "D.yield_2y_us_pct.L1": "Δ Treasury EE. UU. 2 años (t−1)",
    "D.yield_10y_us_pct.L1": "Δ Treasury EE. UU. 10 años (t−1)",
    "D.spread_10y_2y_us_pct.L1": "Δ pendiente Treasury 10Y−2Y (t−1)",
    "D.breakeven_5y_us_pct.L1": "Δ BEI EE. UU. 5 años (t−1)",
    "D.breakeven_10y_us_pct.L1": "Δ BEI EE. UU. 10 años (t−1)",
    "D.epu_global.L1": "Δ incertidumbre de política económica global (t−1)",
    "D.estres_financiero_stl.L1": "Δ estrés financiero STLFSI (t−1)",
    "D.nfci_chicago.L1": "Δ NFCI Chicago (t−1)",
    "D.anfci_chicago.L1": "Δ ANFCI Chicago (t−1)",
    "D.ln_brent_global.L1": "Δ ln Brent global (t−1)",
    "D.ln_commodities_global.L1": "Δ ln commodities globales (t−1)",
    "D.desempleo_us_pct.L2": "Δ desempleo EE. UU. (t−2)",
    "D.ln_empleo_manufactura_us.L2": "Δ ln empleo manufacturero EE. UU. (t−2)",
    "D.ln_produccion_industrial_us.L2": "Δ ln producción industrial EE. UU. (t−2)",
    "D.ln_fletes_transporte_us.L2": "Δ ln fletes/logística EE. UU. (t−2)",
    "D.ln_ise_total_dane.L0": "Δ ln ISE total DANE (t)",
    "D.ln_ipc_colombia.L0": "Δ ln IPC Colombia (t)",
    "D.ln_ise_total_dane.L2": "Δ ln ISE total DANE (t−2)",
    "D.ln_ipc_colombia.L2": "Δ ln IPC Colombia (t−2)",
    "dummy_pandemia_2020": "Dummy pandemia 2020",
}


def csv(relative_path: str) -> pd.DataFrame:
    return pd.read_csv(RESULTS / relative_path)


def value(value: object) -> object:
    """Convert pandas/numpy missing values to Excel-compatible values."""
    if value is None:
        return None
    if isinstance(value, (float, np.floating)) and not np.isfinite(value):
        return None
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def rows_from_frame(frame: pd.DataFrame, columns: Sequence[str] | None = None) -> list[list[object]]:
    selected = frame if columns is None else frame.loc[:, list(columns)]
    return [[value(item) for item in row] for row in selected.itertuples(index=False, name=None)]


def table_object(ws, name: str) -> Table | None:
    table = ws.tables.get(name)
    return table if table is not None else None


def copy_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def write_table(
    ws,
    name: str,
    start_row: int,
    start_col: int,
    headers: Sequence[object],
    data_rows: Iterable[Sequence[object]],
) -> tuple[int, int]:
    """Replace a table's values and reference while retaining its visual style."""
    rows = [list(row) for row in data_rows]
    width = len(headers)
    end_row = start_row + len(rows)
    end_col = start_col + width - 1
    # Los workbooks heredados pueden conservar celdas combinadas dentro del
    # rango de una tabla que ahora crece (por ejemplo, al pasar de 13 a 14
    # factores). Se descombinan únicamente los rangos que se solapan con la
    # tabla objetivo para que openpyxl permita escribir cada valor.
    for merged_range in list(ws.merged_cells.ranges):
        if not (
            merged_range.max_row < start_row
            or merged_range.min_row > end_row
            or merged_range.max_col < start_col
            or merged_range.min_col > end_col
        ):
            ws.unmerge_cells(str(merged_range))

    existing = table_object(ws, name)

    if existing is not None:
        old_min_col, old_min_row, old_max_col, old_max_row = range_boundaries(existing.ref)
        style_rows = max(old_max_row, end_row)
        style_cols = max(old_max_col, end_col)
        header_styles = {
            col: copy(ws.cell(start_row, min(col, old_max_col))._style)
            for col in range(start_col, end_col + 1)
        }
        data_styles = {
            col: copy(ws.cell(min(start_row + 1, old_max_row), min(col, old_max_col))._style)
            for col in range(start_col, end_col + 1)
        }
        for row in range(old_min_row, style_rows + 1):
            for col in range(old_min_col, style_cols + 1):
                if not (start_row <= row <= end_row and start_col <= col <= end_col):
                    ws.cell(row, col).value = None
        column_count_matches = len(existing.tableColumns) == width
    else:
        old_max_col, old_max_row = end_col, end_row
        header_styles = {}
        data_styles = {}
        column_count_matches = False

    for offset, item in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset)
        cell.value = value(item)
        if start_col + offset in header_styles:
            cell._style = copy(header_styles[start_col + offset])

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, item in enumerate(row_values):
            col = start_col + col_offset
            cell = ws.cell(start_row + row_offset, col)
            cell.value = value(item)
            if col in data_styles:
                cell._style = copy(data_styles[col])

    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    if existing is None or not column_count_matches:
        if existing is not None:
            del ws.tables[name]
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        existing.ref = ref
        if existing.autoFilter is not None:
            existing.autoFilter.ref = ref
    return end_row, end_col


def term_label(term: object) -> str:
    text = str(term)
    if text in TERM_LABELS:
        return TERM_LABELS[text]
    return (
        text.replace("D.", "Δ ")
        .replace(".L0", " (t)")
        .replace(".L1", " (t−1)")
        .replace(".L2", " (t−2)")
        .replace(".L3", " (t−3)")
        .replace("_", " ")
    )


def diagnostic_reading(test: str, p_value: object) -> str:
    p = value(p_value)
    if p is None:
        return "Cercano a 2" if "Durbin" in test else "Referencia"
    p_float = float(p)
    if "Jarque" in test:
        return "Residuos no normales" if p_float < 0.05 else "No se rechaza normalidad"
    if "ARCH" in test:
        return "Volatilidad condicional pendiente" if p_float < 0.05 else "Sin ARCH detectado"
    if "CUSUM" in test:
        return "Sin inestabilidad detectada" if p_float >= 0.05 else "Posible inestabilidad"
    if "RESET" in test:
        return "Sin evidencia de mala forma funcional" if p_float >= 0.05 else "Revisar forma funcional"
    if "Durbin" in test:
        return "Cercano a 2"
    return "Sin autocorrelación detectada" if p_float >= 0.05 else "Autocorrelación residual"


def pct(number: float, digits: int = 1) -> str:
    return f"{number * 100:.{digits}f}%"


def effect_text(coefficient: float, proportional_change: float | None = None) -> str:
    if proportional_change is None:
        effect = 100 * (np.exp(coefficient) - 1)
    else:
        effect = 100 * (np.exp(coefficient * np.log1p(proportional_change)) - 1)
    sign = "+" if effect > 0 else "−" if effect < 0 else ""
    return f"{sign}{abs(effect):.2f}%".replace(".", ",")


def update_summary(wb, metadata: dict[str, object]) -> None:
    ws = wb["Resumen"]
    comparison = csv("explicacion/comparacion_especificaciones.csv")
    forecast_metrics = csv("pronostico/validacion_metricas_pronostico.csv")
    weights = csv("explicacion/pesos_explicativos_marco_macro_integral.csv")
    reference = csv("explicacion/coeficientes_controles_externos.csv").set_index("termino")

    ws["A3"] = (
        "Se separan dos usos: explicación histórica con información contemporánea y "
        "pronóstico de un mes con rezagos de publicación. Muestra común: enero de 2006 a abril de 2026."
    )
    ws["A6"] = int(metadata["marco_macro_integral_observaciones"])
    ws["C6"] = pct(float(metadata["adl_r_cuadrado_ajustado"]))
    ws["E6"] = pct(float(metadata["marco_macro_integral_r_cuadrado_ajustado"]))
    ws["A10"] = f"{float(metadata['validacion_marco_macro_integral_mape_pct']):.2f}%"
    ws["C10"] = f"{float(metadata['pronostico_mape_pct']):.2f}%"
    ws["E10"] = pct(float(metadata["pronostico_r2_vs_caminata"]))

    model_rows = rows_from_frame(comparison[[
        "modelo", "observaciones", "r_cuadrado_ajustado", "aic", "bic", "mape_pct", "acierto_direccion_pct"
    ]].rename(columns={"modelo": "Modelo", "observaciones": "Obs.", "r_cuadrado_ajustado": "R² ajustado", "aic": "AIC", "bic": "BIC", "mape_pct": "MAPE", "acierto_direccion_pct": "Dirección"}))
    model_rows = [
        [row[0], row[1], row[2], row[3], row[4], float(row[5]) / 100, float(row[6]) / 100]
        for row in model_rows
    ]
    model_rows.append([
        "Pronóstico publicación",
        int(metadata["pronostico_observaciones"]),
        float(metadata["pronostico_r_cuadrado_ajustado"]),
        float(metadata["pronostico_aic"]),
        float(metadata["pronostico_bic"]),
        float(metadata["pronostico_mape_pct"]) / 100,
        float(metadata["pronostico_acierto_direccion_pct"]) / 100,
    ])
    write_table(ws, "ComparacionModelosResumenTable", 30, 8,
                ["Modelo", "Obs.", "R² ajustado", "AIC", "BIC", "MAPE", "Dirección"], model_rows)

    ordered_weights = weights.sort_values("peso_entre_factores_pct", ascending=False).head(8)
    write_table(
        ws,
        "PesosResumenTable",
        39,
        1,
        ["Factor", "Grupo", "Contribución al R²", "Peso entre factores", "Peso del R² total", "Lectura"],
        [
            [
                row.factor,
                row.grupo,
                row.shapley_r2,
                row.peso_entre_factores_pct / 100,
                row.peso_r2_total_pct / 100,
                "Participación explicativa, no efecto causal",
            ]
            for row in ordered_weights.itertuples(index=False)
        ],
    )

    labels = {
        "D.ln_terminos_intercambio.L0": ("Términos de intercambio", "+10%", 0.10, "Poder de compra externo; el dato se publica con rezago"),
        "D.ln_dolar_amplio.L0": ("Índice amplio del dólar", "+1%", 0.01, "Fortaleza global del dólar"),
        "D.ln_vix.L0": ("VIX", "+10%", 0.10, "Captura episodios globales de aversión al riesgo"),
        "D.ln_remesas_12m.L1": ("Remesas, acumulado 12m (t−1)", "+10%", 0.10, "Puede reflejar endogeneidad y shocks simultáneos"),
        "D.diferencial_tasas_pp.L1": ("Diferencial tasas CO−EE. UU. (t−1)", "+1 pp", None, "Retorno nominal relativo; no es un shock exógeno"),
        "D.deficit_fiscal_12m_pct_pib.L1": ("Déficit fiscal 12m/PIB (t−1)", "+1 pp", None, "Prima fiscal; revise el intervalo antes de concluir"),
        "dummy_pandemia_2020": ("Pandemia, mar–may 2020", "Dummy = 1", None, "Control de episodio extraordinario"),
    }
    output_rows = [["Variable", "Movimiento ilustrativo", "Efecto aproximado en TRM", "Signo", "p-valor", "Lectura"]]
    for term, (label, movement, proportional, reading) in labels.items():
        record = reference.loc[term]
        coefficient = float(record["coeficiente"])
        output_rows.append([
            label,
            movement,
            effect_text(coefficient, proportional),
            "Deprecia el COP" if coefficient >= 0 else "Aprecia el COP",
            float(record["p_valor"]),
            reading,
        ])
    output_rows.append(["Conclusión", "—", "—", "—", None, "La magnitud, la precisión y el peso Shapley responden preguntas distintas"])
    for row_number, row_values in enumerate(output_rows, start=19):
        for column, item in enumerate(row_values, start=1):
            ws.cell(row_number, column).value = value(item)


def update_weights_and_robustness(wb) -> None:
    weights = csv("explicacion/pesos_explicativos_marco_macro_integral.csv")
    write_table(
        wb["Pesos_explicativos"],
        "PesosExplicativosTable",
        10,
        1,
        ["Factor", "Grupo", "Shapley R²", "Peso entre factores", "Peso del R² total", "R² base", "R² completo", "R² incremental"],
        [
            [row.factor, row.grupo, row.shapley_r2, row.peso_entre_factores_pct / 100,
             row.peso_r2_total_pct / 100, row.r2_base, row.r2_completo, row.r2_incremental]
            for row in weights.itertuples(index=False)
        ],
    )
    ws = wb["Pesos_explicativos"]
    first, last = 11, 10 + len(weights)
    ws["B6"] = float(weights["shapley_r2"].sum())
    ws["B7"] = float(weights["r2_incremental"].iloc[0])
    ws["B8"] = float(ws["B6"].value) - float(ws["B7"].value)
    ws["E6"] = float(weights["r2_base"].iloc[0]) + float(weights["r2_incremental"].iloc[0])
    ws["E7"] = float(weights["r2_completo"].iloc[0])
    ws["E8"] = float(ws["E6"].value) - float(ws["E7"].value)
    ws["H6"] = float(weights["peso_entre_factores_pct"].sum()) / 100
    ws["H7"] = float(weights["peso_r2_total_pct"].sum()) / 100
    ws["H8"] = float(ws["H6"].value) - 1
    ws["A2"] = (
        "La descomposición reparte el R² incremental entre factores promediando todos los órdenes de entrada. "
        "El bloque de condiciones financieras, commodities y actividad internacional permanece como un solo jugador "
        "para controlar colinealidad y conservar 14 factores."
    )
    ws["B6"].number_format = ws["B7"].number_format = ws["B8"].number_format = "0.0000%"
    ws["E6"].number_format = ws["E7"].number_format = ws["E8"].number_format = "0.0000%"
    ws["H6"].number_format = ws["H7"].number_format = ws["H8"].number_format = "0.0000%"

    bootstrap = csv("explicacion/intervalos_bootstrap_pesos_shapley.csv")
    write_table(
        wb["Robustez"],
        "IntervalosShapleyTable",
        6,
        1,
        ["Factor", "Peso puntual", "Mediana bootstrap", "IC 95% inferior", "IC 95% superior", "Ancho IC", "Prob. top 3", "Réplicas"],
        [
            [row.factor, row.peso_puntual_pct / 100, row.peso_bootstrap_mediana_pct / 100,
             row.ic_95_inferior_pct / 100, row.ic_95_superior_pct / 100,
             (row.ic_95_superior_pct - row.ic_95_inferior_pct) / 100,
             row.probabilidad_top3_pct / 100, row.replicas_validas]
            for row in bootstrap.itertuples(index=False)
        ],
    )

    stability_summary = csv("explicacion/estabilidad_submuestras_resumen.csv")
    write_table(
        wb["Robustez"],
        "EstabilidadResumenTable",
        6,
        10,
        ["Submuestra", "Obs.", "R² ajustado", "Spearman rangos", "Mediana |Δ peso|", "Máx. |Δ peso|", "Mismo signo / 14"],
        [
            [row.submuestra, row.observaciones, row.r2_ajustado,
             row.correlacion_spearman_rangos_vs_completa,
             row.mediana_diferencia_abs_peso_pp / 100,
             row.max_diferencia_abs_peso_pp / 100,
             row.factores_mismo_signo_de_14]
            for row in stability_summary.itertuples(index=False)
        ],
    )

    stability_detail = csv("explicacion/estabilidad_submuestras_marco_macro_integral.csv")
    subsamples = list(stability_summary["submuestra"])
    factor_order = list(weights["factor"])
    stability_rows: list[list[object]] = []
    for factor in factor_order:
        factor_rows = stability_detail[stability_detail["factor"].eq(factor)]
        by_subsample = factor_rows.set_index("submuestra")["peso_entre_factores_pct"].to_dict()
        group = str(factor_rows["grupo"].iloc[0]) if not factor_rows.empty else ""
        stability_rows.append([factor, group, *[by_subsample.get(label, None) / 100 for label in subsamples]])
    write_table(
        wb["Robustez"],
        "PesosSubmuestrasTable",
        22,
        1,
        ["Factor", "Grupo", *subsamples],
        stability_rows,
    )

    coverage = pd.read_csv(DATA / "base_global_cobertura.csv")
    vintage = csv("pronostico/cobertura_vintages_pronostico.csv")
    write_table(
        wb["Robustez"],
        "CoberturaVintagesTable",
        22,
        9,
        ["Factor", "Estado", "Orígenes", "Cobertura", "Apto", "Archivo desde", "Fuentes", "Detalle"],
        [
            [row.factor, row.estado_vintages_2022_05_a_2026_04, row.origenes_completos_de_48,
             row.cobertura_pct / 100,
             "Sí" if str(row.apto_backtest_genuino).lower() == "true" else "No",
             row.archivo_hacia_adelante_desde, row.fuentes, row.detalle]
            for row in vintage.itertuples(index=False)
        ],
    )
    ws = wb["Robustez"]
    ws["A2"] = (
        "Los intervalos Shapley usan 200 réplicas de bloques circulares de 12 meses. "
        "Las submuestras revelan cuánto cambian pesos, rangos y signos. La cobertura de vintages sigue incompleta: "
        "el ejercicio de pronóstico continúa siendo pseudo-tiempo-real y las series candidatas globales no se imputan."
    )
    # Se deja visible el registro de cobertura global en la hoja de fuentes; leerlo aquí evita que el fallback
    # pierda la evidencia de que los candidatos incompletos siguen documentados.
    _ = coverage


def update_bei(wb, metadata: dict[str, object]) -> None:
    specifications = csv("robustez/comparacion_especificaciones_bei_5y.csv")
    write_table(
        wb["BEI_robustez"],
        "EspecificacionesBeiTable",
        6,
        1,
        ["Especificación", "Agregación", "Transformación", "Extensión", "R² ajustado", "BIC", "Coef. BEI", "p HAC", "MAPE cond.", "R² validación", "Quiebre", "Cautela"],
        [
            [row.especificacion, row.agregacion_bei, row.transformacion_bei,
             row.extension_deterministica, row.r_cuadrado_ajustado, row.bic,
             row.coeficiente_bei_pre_quiebre, row.p_valor_hac_bei_pre_quiebre,
             row.mape_condicional_pct / 100, row.r2_validacion_condicional_vs_caminata,
             row.fecha_quiebre_za if value(row.fecha_quiebre_za) is not None else "—",
             "Quiebre elegido ex post" if str(row.quiebre_elegido_con_muestra_completa).lower() == "true" else "Comparación regular"]
            for row in specifications.itertuples(index=False)
        ],
    )

    aggregation = csv("robustez/comparacion_agregacion_bei_5y.csv")
    common_days = sorted(aggregation["dias_comunes"].dropna().astype(float).tolist())
    median_days = common_days[len(common_days) // 2] if len(common_days) % 2 else (common_days[len(common_days) // 2 - 1] + common_days[len(common_days) // 2]) / 2
    aggregation_rows = [
        ["Correlación entre agregaciones", float(metadata["diferencial_bei_5y_correlacion_agregaciones"]), "Prácticamente idénticas en el conjunto de la muestra"],
        ["Diferencia media común − separada", float(metadata["diferencial_bei_5y_diferencia_media_comun_menos_separada_pp"]), "Puntos porcentuales"],
        ["Máxima diferencia absoluta", float(metadata["diferencial_bei_5y_max_diferencia_abs_agregacion_pp"]), "Puntos porcentuales; ocurre en un mes con pocos cruces"],
        ["Mínimo de días comunes", float(metadata["diferencial_bei_5y_min_dias_comunes_mes"]), "La intersección puede perder gran parte del mes"],
        ["Mediana de días comunes", median_days, "Días con las tres curvas observadas"],
    ]
    write_table(wb["BEI_robustez"], "AgregacionBeiResumenTable", 16, 1, ["Métrica", "Valor", "Lectura"], aggregation_rows)

    stationarity = csv("robustez/pruebas_estacionariedad_bei_5y.csv")
    stationarity = stationarity.loc[
        ((stationarity["prueba"] == "ADF") & stationarity["deterministico"].isin(["constante", "constante_tendencia"]))
        | ((stationarity["prueba"] == "KPSS") & (stationarity["deterministico"] == "constante_tendencia"))
        | ((stationarity["prueba"] == "Zivot-Andrews") & (stationarity["deterministico"] == "constante_tendencia_con_quiebre"))
    ]
    write_table(
        wb["BEI_robustez"],
        "EstacionariedadBeiTable",
        25,
        1,
        ["Agregación", "Transformación", "Prueba", "Determinístico", "H₀", "N", "Estadístico", "p-valor", "Rezagos", "Fecha quiebre", "Crítico 5%"],
        [
            [row.agregacion, row.transformacion, row.prueba, row.deterministico, row.hipotesis_nula,
             row.n, row.estadistico, row.p_valor, row.rezagos,
             row.fecha_quiebre if value(row.fecha_quiebre) is not None else "—", row.critico_5_pct]
            for row in stationarity.itertuples(index=False)
        ],
    )

    trends = csv("robustez/tendencias_quiebres_bei_5y.csv")
    trend_start = 27 + len(stationarity)
    write_table(
        wb["BEI_robustez"],
        "TendenciasBeiTable",
        trend_start + 1,
        1,
        ["Agregación", "Modelo", "Quiebre ZA", "N", "R² ajustado", "BIC", "Tendencia pp/año", "p tendencia", "Cambio nivel", "p nivel", "Cambio pendiente", "p pendiente"],
        [
            [row.agregacion, row.modelo_deterministico, row.fecha_quiebre_za, row.observaciones,
             row.r_cuadrado_ajustado, row.bic, row.tendencia_pp_por_ano, row.p_valor_hac_tendencia,
             row.cambio_nivel_quiebre_pp, row.p_valor_hac_cambio_nivel,
             row.cambio_pendiente_pp_por_ano, row.p_valor_hac_cambio_pendiente]
            for row in trends.itertuples(index=False)
        ],
    )
    active = specifications.loc[specifications["especificacion"].str.contains("vigente")].iloc[0]
    level = specifications.loc[specifications["especificacion"].str.contains("Nivel")].iloc[0]
    wb["BEI_robustez"]["A2"] = (
        "Se adopta Δ diferencial BEI con promedios mensuales separados y rezago de un mes: "
        f"BIC {float(active.bic):.2f} frente a {float(level.bic):.2f} en nivel. "
        "La primera diferencia es la especificación vigente; nivel, tendencia, quiebre y fechas comunes quedan como robustez."
    )


def update_forecast(wb, metadata: dict[str, object]) -> None:
    ws = wb["Pronostico"]
    metrics = csv("pronostico/validacion_metricas_pronostico.csv")
    write_table(
        ws,
        "MetricasPronosticoTable",
        6,
        1,
        ["Modelo", "Observaciones", "MAE (log)", "RMSE (log)", "MAPE", "Acierto dirección"],
        [
            [row.modelo, row.observaciones, row.mae_log, row.rmse_log, row.mape_pct / 100,
             None if pd.isna(row.acierto_direccion_pct) else row.acierto_direccion_pct / 100]
            for row in metrics.itertuples(index=False)
        ],
    )
    regional = csv("explicacion/comparacion_factor_regional.csv")
    write_table(
        ws,
        "ComparacionFactorRegionalTable",
        13,
        1,
        ["Uso", "Monedas", "R² ajustado", "BIC", "MAPE", "Dirección", "R² vs. caminata", "Coef. regional", "p-valor HAC", "Correlación 3–4"],
        [
            [row.uso, row.monedas, row.r_cuadrado_ajustado, row.bic, row.mape_pct / 100,
             row.acierto_direccion_pct / 100, row.r2_validacion_vs_caminata,
             row.coeficiente_factor_regional, row.p_valor_hac_factor_regional,
             row.correlacion_factores_3_4]
            for row in regional.itertuples(index=False)
        ],
    )
    availability = csv("pronostico/calendario_disponibilidad_pronostico.csv")
    write_table(
        ws,
        "CalendarioDisponibilidadTable",
        20,
        1,
        ["Factor", "Rezago (meses)", "Frecuencia/publicación", "Regla utilizada"],
        [
            [row.factor, row.rezago_meses_modelo, row.frecuencia_y_publicacion,
             row.regla_disponibilidad_al_inicio_del_mes_t]
            for row in availability.itertuples(index=False)
        ],
    )
    coefficients = csv("pronostico/coeficientes_modelo_pronostico.csv")
    predictions = csv("pronostico/validacion_predicciones_pronostico.csv")
    forecast_header = max(38, 22 + len(coefficients))
    # La tabla histórica de predicciones comenzaba en A38. Se mueve primero
    # para dejar espacio a los 31 coeficientes actuales (F20:L51).
    if "A37:I37" in {str(item) for item in ws.merged_cells.ranges}:
        ws.unmerge_cells("A37:I37")
    write_table(
        ws,
        "PrediccionesPronosticoTable",
        forecast_header,
        1,
        ["Mes", "ln TRM observada", "ln TRM pronóstico", "ln TRM caminata", "Δln observado", "Δln pronóstico", "TRM observada", "TRM pronóstico", "TRM caminata"],
        [
            [str(row.fecha)[:7], row.ln_trm_observada, row.ln_trm_pronostico_publicacion,
             row.ln_trm_caminata_aleatoria, row.cambio_log_observado,
             row.cambio_log_pronostico, row.trm_observada,
             row.trm_pronostico_publicacion, row.trm_caminata_aleatoria]
            for row in predictions.itertuples(index=False)
        ],
    )
    section_row = forecast_header - 1
    ws.cell(section_row, 1).value = "Predicciones de un mes con información disponible al origen"
    ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=9)
    write_table(
        ws,
        "CoeficientesPronosticoTable",
        20,
        6,
        ["Variable", "Coeficiente", "EE HAC", "t", "p-valor", "IC 95% inferior", "IC 95% superior"],
        [
            [term_label(row.termino), row.coeficiente, row.error_estandar_hac, row.estadistico_t,
             row.p_valor, row.ic_95_inferior, row.ic_95_superior]
            for row in coefficients.itertuples(index=False)
        ],
    )
    diagnostics = csv("pronostico/diagnosticos_modelo_pronostico.csv")
    write_table(
        ws,
        "DiagnosticosPronosticoTable",
        20,
        14,
        ["Prueba", "Estadístico", "p-valor"],
        [[row.prueba, row.estadistico, row.p_valor] for row in diagnostics.itertuples(index=False)],
    )
    ws["A2"] = (
        "Objetivo: pronosticar la TRM promedio del mes t al inicio de t. Ningún factor económico del mes objetivo entra "
        "contemporáneamente. El backtest respeta un calendario conservador, pero usa el último vintage disponible; por eso es pseudo-tiempo-real."
    )
    ws["J8"] = float(metrics.loc[metrics["modelo"].str.contains("Pronóstico")].iloc[0].mape_pct) / 100
    ws["J9"] = float(metadata["pronostico_r2_vs_caminata"])


def update_models(wb) -> None:
    principal = csv("explicacion/coeficientes_controles_externos.csv")
    write_table(
        wb["Controles_externos"],
        "CoeficientesModeloTable",
        5,
        1,
        ["Variable", "Coeficiente", "EE HAC", "t", "p-valor", "IC 95% inferior", "IC 95% superior"],
        [[term_label(row.termino), row.coeficiente, row.error_estandar_hac, row.estadistico_t,
          row.p_valor, row.ic_95_inferior, row.ic_95_superior] for row in principal.itertuples(index=False)],
    )

    integrated = csv("explicacion/coeficientes_marco_macro_integral.csv")
    ws = wb["Marco_macro_integral"]
    write_table(
        ws,
        "CoeficientesMarcoMacroIntegralTable",
        5,
        1,
        ["Variable", "Coeficiente", "EE HAC", "t", "p-valor", "IC 95% inferior", "IC 95% superior"],
        [[term_label(row.termino), row.coeficiente, row.error_estandar_hac, row.estadistico_t,
          row.p_valor, row.ic_95_inferior, row.ic_95_superior] for row in integrated.itertuples(index=False)],
    )
    integrated_diagnostics = csv("explicacion/diagnosticos_marco_macro_integral.csv")
    write_table(
        ws,
        "DiagnosticosMarcoMacroIntegralTable",
        13,
        9,
        ["Prueba", "Estadístico", "p-valor", "Lectura"],
        [[row.prueba, row.estadistico, row.p_valor, diagnostic_reading(row.prueba, row.p_valor)] for row in integrated_diagnostics.itertuples(index=False)],
    )

    integrated_fit = csv("explicacion/ajuste_historico_marco_macro_integral.csv")
    contributions = csv("explicacion/contribuciones_marco_macro_integral.csv")
    contribution_keys = [column for column in contributions.columns if column != "fecha"]
    contribution_by_date = contributions.set_index("fecha").to_dict(orient="index")
    fit_header = max(22, 5 + len(integrated) + 3, 13 + len(integrated_diagnostics) + 3)
    headers = [
        "Mes", "Δln TRM observado", "Δln TRM ajustado", "Residuo", "TRM observada", "TRM ajustada 1 paso",
        *["Ajuste total de contribuciones" if key == "ajuste_total" else term_label(key) for key in contribution_keys],
        "Diferencia de control",
    ]
    fit_rows: list[list[object]] = []
    for row in integrated_fit.itertuples(index=False):
        date = str(row.fecha)
        contribution = contribution_by_date.get(date, {})
        adjustment = contribution.get("ajuste_total")
        fit_rows.append([
            date[:7], row.cambio_log_observado, row.cambio_log_ajustado, row.residuo_cambio_log,
            row.trm_observada, row.trm_ajustada_un_paso,
            *[contribution.get(key) for key in contribution_keys],
            None if value(adjustment) is None else row.cambio_log_ajustado - adjustment,
        ])
    write_table(ws, "AjusteMarcoMacroIntegralTable", fit_header, 1, headers, fit_rows)
    ws["A1"] = "Marco macroeconómico integral: condiciones externas, internas y regionales"
    ws["A2"] = (
        "Explicación histórica ex post. Integra 14 factores, incluido un bloque de 17 términos globales agrupados: "
        "expectativas y rendimientos de EE. UU., condiciones financieras, commodities, desempleo, empleo industrial y fletes. "
        "Las series candidatas incompletas se documentan, pero no se imputan ni entran al modelo balanceado."
    )
    # El gráfico anterior apuntaba a la tabla vieja. Se sustituye por un gráfico pequeño con la tabla vigente.
    ws._charts = []
    if len(fit_rows) > 0:
        chart = LineChart()
        chart.title = "TRM observada y ajuste del marco macroeconómico integral"
        chart.y_axis.title = "COP por USD"
        chart.x_axis.title = "Mes"
        data = Reference(ws, min_col=5, max_col=6, min_row=fit_header, max_row=fit_header + len(fit_rows))
        categories = Reference(ws, min_col=1, min_row=fit_header + 1, max_row=fit_header + len(fit_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 13
        ws.add_chart(chart, "O5")

    controles_externos_diagnostics = csv("explicacion/diagnosticos_controles_externos.csv")
    diagnostic_ws = wb["Diagnosticos"]
    write_table(
        diagnostic_ws,
        "DiagnosticosModeloTable",
        6,
        1,
        ["Prueba", "Estadístico", "p-valor", "Lectura"],
        [[row.prueba, row.estadistico, row.p_valor, diagnostic_reading(row.prueba, row.p_valor)] for row in controles_externos_diagnostics.itertuples(index=False)],
    )
    write_table(
        diagnostic_ws,
        "DiagnosticosMarcoMacroIntegralResumenTable",
        15,
        10,
        ["Prueba", "Estadístico", "p-valor", "Lectura"],
        [[row.prueba, row.estadistico, row.p_valor, diagnostic_reading(row.prueba, row.p_valor)] for row in integrated_diagnostics.itertuples(index=False)],
    )
    diagnostic_ws["J14"] = "Diagnósticos del marco macroeconómico integral"


def update_source_sheet(wb, metadata: dict[str, object]) -> None:
    ws = wb["Datos_fuente"]
    monthly = pd.read_csv(DATA / "modelo_trm_datos_mensuales.csv")
    monthly["fecha"] = pd.to_datetime(monthly["fecha"])
    sample_start = pd.Timestamp(metadata["muestra_inicio"]) - pd.DateOffset(months=1)
    sample_end = pd.Timestamp(metadata["muestra_fin"])
    monthly = monthly.loc[monthly["fecha"].between(sample_start, sample_end)].copy()

    base_headers = [
        "Mes", "TRM (COP/USD)", "Términos de intercambio", "Remesas (USD mill.)", "Remesas 12m (USD mill.)",
        "Tasa política Colombia (%)", "Fed funds (%)", "Diferencial tasas (pp)", "Balance fiscal mensual (miles mill. COP)",
        "Déficit fiscal 12m (% PIB)", "Índice dólar amplio", "VIX", "Dummy pandemia", "Reservas netas sin FLAR (USD mill.)",
        "EMBIG Colombia (pb)", "EMBIG Colombia (pp)", "Balanza comercial cambiaria (USD mill.)", "Flujo neto total de capital (USD mill.)",
        "TES pesos cero cupón 5 años (%)", "TES UVR cero cupón 5 años (%)", "BEI Colombia 5 años (%)", "BEI EE. UU. 5 años (%)",
        "Diferencial BEI 5 años (pp)", "BRL por USD", "CLP por USD", "MXN por USD", "PEN por USD",
        "Factor regional 3 monedas", "Factor regional 4 monedas", "ISE total DANE (índice)", "IPC Colombia (índice)",
    ]
    base_columns = [
        "fecha", "trm_cop_usd", "terminos_intercambio", "remesas_usd_millones", "remesas_12m_usd_millones",
        "tasa_politica_colombia_pct", "fed_funds_eeuu_pct", "diferencial_tasas_pp", "balance_fiscal_miles_millones_cop",
        "deficit_fiscal_12m_pct_pib", "indice_dolar_amplio", "vix", "dummy_pandemia_2020", "reservas_netas_sin_flar_usd_millones",
        "embig_colombia_pb", "embig_colombia_pp", "balanza_comercial_cambiaria_usd_millones", "flujos_capital_usd_millones",
        "tes_5y_pesos_colombia_pct", "tes_5y_uvr_colombia_pct", "bei_colombia_5y_pct", "bei_eeuu_5y_pct",
        "diferencial_bei_5y_pp", "brl_por_usd", "clp_por_usd", "mxn_por_usd", "pen_por_usd",
        "factor_monedas_regionales_3", "factor_monedas_regionales_4", "ise_total_dane", "ipc_colombia_indice",
    ]
    global_specs = [
        ("yield_real_10y_tips_pct", "TIPS real EE. UU. 10 años (%)"),
        ("yield_real_5y_us_pct", "Rendimiento real EE. UU. 5 años (%)"),
        ("yield_2y_us_pct", "Treasury EE. UU. 2 años (%)"),
        ("yield_10y_us_pct", "Treasury EE. UU. 10 años (%)"),
        ("spread_10y_2y_us_pct", "Pendiente Treasury 10Y−2Y (pp)"),
        ("breakeven_5y_us_pct", "BEI EE. UU. 5 años (%)"),
        ("breakeven_10y_us_pct", "BEI EE. UU. 10 años (%)"),
        ("epu_global", "Incertidumbre de política económica global"),
        ("estres_financiero_stl", "Estrés financiero STLFSI"),
        ("nfci_chicago", "NFCI Chicago"),
        ("anfci_chicago", "ANFCI Chicago"),
        ("desempleo_us_pct", "Desempleo EE. UU. activo (%)"),
        ("high_yield_oas_pct", "High-yield OAS (candidato)"),
        ("ted_spread_pct", "TED spread (candidato)"),
        ("desempleo_us_bls_pct", "Desempleo BLS UNRATE (candidato)"),
        ("precios_importacion_china", "Precios de importación China (candidato)"),
        ("produccion_industrial_china", "Producción industrial China (candidato)"),
        ("indicador_lider_china", "Indicador líder China (candidato)"),
        ("ipc_china", "IPC China (candidato)"),
        ("ln_brent_global", "ln Brent global"),
        ("ln_commodities_global", "ln commodities globales"),
        ("ln_empleo_manufactura_us", "ln empleo manufacturero EE. UU."),
        ("ln_produccion_industrial_us", "ln producción industrial EE. UU."),
        ("ln_fletes_transporte_us", "ln fletes/logística EE. UU."),
    ]
    headers = [*base_headers, *[label for _, label in global_specs]]
    data_rows = []
    for _, row in monthly.iterrows():
        data_rows.append([
            row["fecha"].strftime("%Y-%m"),
            *[row.get(column) for column in base_columns[1:]],
            *[row.get(column) for column, _ in global_specs],
        ])
    write_table(ws, "DatosFuenteTable", 5, 1, headers, data_rows)
    ws["A1"] = "Datos fuente mensuales y cobertura global"
    ws["A2"] = (
        "Niveles y transformaciones previas de las series utilizadas. Las columnas globales activas y candidatas se muestran "
        "para auditoría; ISE total DANE e IPC Colombia conservan niveles oficiales sin imputación; high-yield, TED, UNRATE y China "
        "conservan sus faltantes y no se imputan."
    )
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col > 29 else ws.column_dimensions[get_column_letter(col)].width
    ws.column_dimensions["A"].width = 12


def update_transformations(wb, metadata: dict[str, object]) -> None:
    """Rebuild the auditable transformation sheet, including ISE and IPC."""
    ws = wb["Transformaciones"]
    monthly = pd.read_csv(DATA / "modelo_trm_datos_mensuales.csv")
    monthly["fecha"] = pd.to_datetime(monthly["fecha"])
    sample_start = pd.Timestamp(metadata["muestra_inicio"]) - pd.DateOffset(months=1)
    sample_end = pd.Timestamp(metadata["muestra_fin"])
    monthly = monthly.loc[monthly["fecha"].between(sample_start, sample_end)].copy()

    headers = [
        "Mes", "ln TRM", "Δln TRM", "ln términos de intercambio", "Δln términos de intercambio",
        "ln remesas 12m", "Δln remesas 12m", "Diferencial tasas", "Δ diferencial", "Déficit 12m/PIB", "Δ déficit",
        "ln dólar amplio", "Δln dólar amplio", "ln VIX", "Δln VIX", "Pandemia", "ln reservas netas sin FLAR", "Δln reservas",
        "EMBIG Colombia (pp)", "Δ EMBIG Colombia", "asinh balanza (USD miles de millones)", "Δ asinh balanza",
        "asinh flujos (USD miles de millones)", "Δ asinh flujos", "Diferencial BEI 5 años", "Δ diferencial BEI",
        "Factor regional 3 monedas", "Factor regional 4 monedas", "ln ISE total DANE", "Δln ISE total DANE",
        "ln IPC Colombia", "Δln IPC Colombia",
    ]
    rows: list[list[object]] = []
    for offset in range(len(monthly)):
        r = 6 + offset
        first = offset == 0
        rows.append([
            f"='Datos_fuente'!A{r}",
            f"=IF('Datos_fuente'!B{r}=\"\",\"\",LN('Datos_fuente'!B{r}))",
            None if first else f'=IF(OR(B{r}=\"\",B{r - 1}=\"\"),\"\",B{r}-B{r - 1})',
            f"=IF('Datos_fuente'!C{r}=\"\",\"\",LN('Datos_fuente'!C{r}))",
            None if first else f'=IF(OR(D{r}=\"\",D{r - 1}=\"\"),\"\",D{r}-D{r - 1})',
            f"=IF('Datos_fuente'!E{r}=\"\",\"\",LN('Datos_fuente'!E{r}))",
            None if first else f'=IF(OR(F{r}=\"\",F{r - 1}=\"\"),\"\",F{r}-F{r - 1})',
            f"='Datos_fuente'!H{r}",
            None if first else f'=IF(OR(H{r}=\"\",H{r - 1}=\"\"),\"\",H{r}-H{r - 1})',
            f"='Datos_fuente'!J{r}",
            None if first else f'=IF(OR(J{r}=\"\",J{r - 1}=\"\"),\"\",J{r}-J{r - 1})',
            f"=IF('Datos_fuente'!K{r}=\"\",\"\",LN('Datos_fuente'!K{r}))",
            None if first else f'=IF(OR(L{r}=\"\",L{r - 1}=\"\"),\"\",L{r}-L{r - 1})',
            f"=IF('Datos_fuente'!L{r}=\"\",\"\",LN('Datos_fuente'!L{r}))",
            None if first else f'=IF(OR(N{r}=\"\",N{r - 1}=\"\"),\"\",N{r}-N{r - 1})',
            f"='Datos_fuente'!M{r}",
            f"=IF('Datos_fuente'!N{r}=\"\",\"\",LN('Datos_fuente'!N{r}))",
            None if first else f'=IF(OR(Q{r}=\"\",Q{r - 1}=\"\"),\"\",Q{r}-Q{r - 1})',
            f"='Datos_fuente'!P{r}",
            None if first else f'=IF(OR(S{r}=\"\",S{r - 1}=\"\"),\"\",S{r}-S{r - 1})',
            f"=IF('Datos_fuente'!Q{r}=\"\",\"\",ASINH('Datos_fuente'!Q{r}/1000))",
            None if first else f'=IF(OR(U{r}=\"\",U{r - 1}=\"\"),\"\",U{r}-U{r - 1})',
            f"=IF('Datos_fuente'!R{r}=\"\",\"\",ASINH('Datos_fuente'!R{r}/1000))",
            None if first else f'=IF(OR(W{r}=\"\",W{r - 1}=\"\"),\"\",W{r}-W{r - 1})',
            f"='Datos_fuente'!W{r}",
            None if first else f'=IF(OR(Y{r}=\"\",Y{r - 1}=\"\"),\"\",Y{r}-Y{r - 1})',
            f"='Datos_fuente'!AB{r}",
            f"='Datos_fuente'!AC{r}",
            f"=IF('Datos_fuente'!AD{r}=\"\",\"\",LN('Datos_fuente'!AD{r}))",
            None if first else f'=IF(OR(AC{r}=\"\",AC{r - 1}=\"\"),\"\",AC{r}-AC{r - 1})',
            f"=IF('Datos_fuente'!AE{r}=\"\",\"\",LN('Datos_fuente'!AE{r}))",
            None if first else f'=IF(OR(AE{r}=\"\",AE{r - 1}=\"\"),\"\",AE{r}-AE{r - 1})',
        ])
    write_table(ws, "TransformacionesTable", 5, 1, headers, rows)
    ws["A1"] = "Transformaciones auditables, incluidas las variables internas"
    ws["A2"] = (
        "Todas las columnas se enlazan a Datos_fuente. ISE total DANE e IPC Colombia se transforman con ln(x) y primera diferencia; "
        "los faltantes permanecen vacíos y no se interpolan ni se extrapolan."
    )


def update_sources(wb) -> None:
    ws = wb["Fuentes"]
    existing = [[ws.cell(row, col).value for col in range(1, 8)] for row in range(6, 28)]
    existing = [row for row in existing if any(item is not None for item in row)]
    for row in existing:
        if row[4] == "Marco macroeconómico integral" and row[1] == "EMBIG Colombia PD04715XD":
            row[5] = "Promedio mensual; pb/100; fuentes originales Reuters/J.P. Morgan"
    global_rows = [
        ["DANE", "ISE total; Cuadro 2; 12 agrupaciones", "Mensual", "2005–2026", "Marco macroeconómico integral", "Índice ajustado por efecto estacional y calendario; ln y primera diferencia; rezago 2 en pronóstico; sin imputación", "https://www.dane.gov.co/index.php/en/statistics-by-topic/national-accounts/economic-monitor-index-ise"],
        ["Banco de la República", "IPC Colombia, serie 15000", "Mensual", "1954–2026", "Marco macroeconómico integral", "Índice oficial; ln y primera diferencia; rezago 2 en pronóstico; sin imputación", "https://suameca.banrep.gov.co/graficador-series/rest/graficadorService/consultaSerieParaGraficar?idSerie=15000"],
        ["DANE", "ISE sectorial 9 y 12 agrupaciones", "Mensual", "2005–2026", "Candidata auditable", "Se conserva para cobertura y diagnóstico; no se activa junto al total por colinealidad", "https://www.dane.gov.co/index.php/en/statistics-by-topic/national-accounts/economic-monitor-index-ise"],
        ["DANE", "GEIH nacional y desestacionalizada", "Mensual", "Incompleta", "Candidata documentada", "Cobertura incompleta; se mantienen faltantes y no se empalma ni se imputa", "https://www.dane.gov.co/index.php/estadisticas-por-tema/mercado-laboral/empleo-y-desempleo/mercado-laboral-historicos"],
        ["DANE", "IPI total e IPP producción nacional", "Mensual", "Desde 2014", "Candidata documentada", "No cubren 2006–2026; no se extrapolan ni se empalman con una base anterior", "https://www.dane.gov.co/index.php/estadisticas-por-tema/industria/indice-de-produccion-industrial-ipi"],
        ["Federal Reserve Board / FRED", "DFII10, DFII5, DGS2, DGS10", "Diaria → mensual", "2003–2026", "Marco macroeconómico integral", "Rendimientos reales y nominales de EE. UU.; cambios contemporáneos o rezagados", "https://fred.stlouisfed.org/"],
        ["Federal Reserve Board / FRED", "T5YIE, T10YIE", "Diaria → mensual", "2003–2026", "Marco macroeconómico integral", "Expectativas/compensación de inflación a 5 y 10 años; cambios", "https://fred.stlouisfed.org/"],
        ["FRED", "DCOILBRENTEU, PALLFNFINDEXM", "Diaria/mensual", "2000–2026", "Marco macroeconómico integral", "Brent y commodities; logs y diferencias", "https://fred.stlouisfed.org/"],
        ["FRED", "USEPUINDXD, STLFSI4, NFCI, ANFCI", "Semanal/mensual", "2000–2026", "Marco macroeconómico integral", "Incertidumbre y condiciones financieras; diferencias", "https://fred.stlouisfed.org/"],
        ["FRED", "LRUN64TTUSM156S", "Mensual", "2000–2026", "Marco macroeconómico integral", "Desempleo estadounidense activo; diferencia; serie completa en la muestra", "https://fred.stlouisfed.org/series/LRUN64TTUSM156S"],
        ["FRED", "MANEMP, INDPRO, TSIFRGHT", "Mensual", "2000–2026", "Marco macroeconómico integral", "Empleo industrial, producción y fletes/logística; empleo/fletes con L2 en pronóstico", "https://fred.stlouisfed.org/"],
        ["FRED", "BAMLH0A0HYM2", "Diaria", "Incompleta", "Candidato documentado", "High-yield OAS; no cubre la muestra activa y no se imputa", "https://fred.stlouisfed.org/series/BAMLH0A0HYM2"],
        ["FRED", "TEDRATE", "Diaria", "Hasta 2022-01", "Candidato documentado", "TED spread; cobertura incompleta, fuera del modelo balanceado", "https://fred.stlouisfed.org/series/TEDRATE"],
        ["BLS / FRED", "UNRATE", "Mensual", "Faltante 2025-10", "Candidato documentado", "Desempleo BLS; se conserva el faltante, sin interpolación", "https://fred.stlouisfed.org/series/UNRATE"],
        ["OECD / FRED", "CHNTOT, CHNPRINTO01IXPYM, CHNLORSGPRTSTSAM, CHNCPIALLMINMEI", "Mensual", "Incompleta", "Candidato documentado", "Indicadores de China; cobertura y faltantes impiden inclusión activa", "https://fred.stlouisfed.org/"],
    ]
    write_table(ws, "FuentesTable", 5, 1,
                ["Organismo", "Serie/código", "Frecuencia", "Cobertura", "Uso", "Tratamiento", "URL"],
                [*existing, *global_rows])
    ws["A2"] = (
        "Enlaces oficiales o distribuidores públicos de las series activas y candidatas. Las candidatas incompletas "
        "se documentan para trazabilidad, pero no entran al modelo balanceado ni se rellenan artificialmente."
    )


def _rename_model_sheets(wb) -> None:
    """Alinea las dos hojas de especificación con sus nombres descriptivos."""
    target_names = ("Controles_externos", "Marco_macro_integral")
    for worksheet, target_name in zip(list(wb.worksheets)[3:5], target_names):
        if worksheet.title != target_name:
            worksheet.title = target_name


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"No existe el workbook base: {WORKBOOK_PATH}")
    metadata = pd.read_json(RESULTS / "metadata.json", typ="series").to_dict()
    wb = load_workbook(WORKBOOK_PATH)
    _rename_model_sheets(wb)
    update_summary(wb, metadata)
    update_weights_and_robustness(wb)
    update_bei(wb, metadata)
    update_forecast(wb, metadata)
    update_models(wb)
    update_source_sheet(wb, metadata)
    update_transformations(wb, metadata)
    update_sources(wb)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass
    wb.save(WORKBOOK_PATH)
    print(f"OK: workbook sincronizado con openpyxl -> {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
