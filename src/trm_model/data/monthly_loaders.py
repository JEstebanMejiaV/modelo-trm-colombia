from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from trm_model.monthly.specifications import (
    ROOT,
    RAW,
    DATA,
    MONTH_NUMBERS_ES,
    GLOBAL_BASE_FILE,
    GLOBAL_RAW_COMPONENTS,
    SAMPLE_START,
    SAMPLE_END,
)


DANE_MONTH_NUMBERS_ES = {
    **MONTH_NUMBERS_ES,
    "Enero": 1,
    "Febrero": 2,
    "Marzo": 3,
    "Abril": 4,
    "Mayo": 5,
    "Junio": 6,
    "Julio": 7,
    "Agosto": 8,
    "Septiembre": 9,
    "Octubre": 10,
    "Noviembre": 11,
    "Diciembre": 12,
}

ISE_COMPONENTS = {
    "Actividades primarias": "ise_actividades_primarias_dane",
    "Agricultura, ganadería, caza, silvicultura y pesca": "ise_agricultura_dane",
    "Explotación de minas y canteras": "ise_mineria_dane",
    "Actividades secundarias": "ise_actividades_secundarias_dane",
    "Industrias manufactureras": "ise_manufactura_dane",
    "Construcción": "ise_construccion_dane",
    "Actividades terciarias": "ise_actividades_terciarias_dane",
    "Suministro de electricidad, gas, vapor y aire acondicionado; Distribución de agua; Evacuación y tratamiento de aguas residuales, gestión de desechos y actividades de saneamiento ambiental": "ise_electricidad_agua_dane",
    "Comercio al por mayor y al por menor; Reparación de vehículos automotores y motocicletas; Transporte y almacenamiento; Alojamiento y servicios de comida": "ise_comercio_transporte_alojamiento_dane",
    "Información y comunicaciones": "ise_informacion_comunicaciones_dane",
    "Actividades financieras y de seguros": "ise_finanzas_dane",
    "Actividades inmobiliarias": "ise_inmobiliarias_dane",
    "Actividades profesionales, científicas y técnicas; Actividades de servicios administrativos y de apoyo": "ise_profesionales_administrativos_dane",
    "Administración pública y defensa; planes de seguridad social de afiliación obligatoria; Educación; Actividades de atención de la salud humana y de servicios sociales": "ise_administracion_educacion_salud_dane",
    "Actividades artísticas, de entretenimiento y recreación y otras actividades de servicios; Actividades de los hogares individuales en calidad de empleadores; Actividades no diferenciadas de los hogares individuales como productores de bienes y servicios para uso propio": "ise_arte_hogares_dane",
    "Indicador de Seguimiento a la Economía": "ise_total_dane",
}

GEIH_COMPONENTS = {
    "Tasa Global de Participación (TGP)": "tasa_participacion_dane_pct",
    "Tasa de Ocupación (TO)": "tasa_ocupacion_dane_pct",
    "Tasa de Desocupación (TD)": "tasa_desocupacion_dane_pct",
    "Población ocupada": "ocupados_dane_miles",
    "Población desocupada": "desocupados_dane_miles",
    "Población fuera de la fuerza de trabajo": "fuera_fuerza_trabajo_dane_miles",
}


def month_start(values: pd.Series | pd.DatetimeIndex) -> pd.DatetimeIndex:
    return pd.DatetimeIndex(values).to_period("M").to_timestamp()


def _dane_wide_dates(year_values, month_values) -> list[pd.Timestamp]:
    """Construye fechas mensuales desde encabezados DANE con años combinados."""
    dates: list[pd.Timestamp] = []
    year: int | None = None
    for year_value, month_value in zip(year_values, month_values):
        if year_value is not None:
            year_text = str(year_value).strip()
            if year_text[:4].isdigit():
                year = int(year_text[:4])
        month = DANE_MONTH_NUMBERS_ES.get(str(month_value).strip())
        if year is not None and month is not None:
            dates.append(pd.Timestamp(year=year, month=month, day=1))
        else:
            dates.append(pd.NaT)
    return dates


def _dane_wide_series(values, dates: list[pd.Timestamp]) -> pd.Series:
    series = pd.Series(
        pd.to_numeric(values, errors="coerce"),
        index=pd.DatetimeIndex(dates),
        dtype="float64",
    )
    return series[series.index.notna()].groupby(level=0).mean().sort_index()


def load_ise_dane(path: Path) -> pd.DataFrame:
    """Lee el ISE DANE ajustado por efecto estacional y calendario."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Cuadro 2"]
        rows = list(worksheet.iter_rows(values_only=True))
        dates = _dane_wide_dates(rows[10][3:], rows[11][3:])
        rows_by_concept: dict[str, tuple[object, ...]] = {}
        for row in rows[13:29]:
            concept = str(row[2] or "").replace("\n", " ").strip()
            if concept in ISE_COMPONENTS and concept not in rows_by_concept:
                rows_by_concept[concept] = row
        missing = sorted(set(ISE_COMPONENTS) - set(rows_by_concept))
        if missing:
            raise ValueError(f"No se encontraron conceptos ISE en {path}: {missing}")
        return pd.DataFrame(
            {
                output_name: _dane_wide_series(rows_by_concept[concept][3:], dates)
                for concept, output_name in ISE_COMPONENTS.items()
            }
        ).sort_index()
    finally:
        workbook.close()


def load_geih_dane(path: Path, seasonally_adjusted: bool = False) -> pd.DataFrame:
    """Lee las series mensuales nacionales de la GEIH publicadas por DANE."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Total nacional"]
        rows = list(worksheet.iter_rows(values_only=True))
        dates = _dane_wide_dates(rows[11][1:], rows[12][1:])
        rows_by_concept: dict[str, tuple[object, ...]] = {}
        for row in rows[13:]:
            concept = str(row[0] or "").strip()
            if concept in GEIH_COMPONENTS and concept not in rows_by_concept:
                rows_by_concept[concept] = row
        selected = {
            (
                output_name.replace("_dane_", "_dane_sa_")
                if seasonally_adjusted
                else output_name
            ): _dane_wide_series(rows_by_concept[concept][1:], dates)
            for concept, output_name in GEIH_COMPONENTS.items()
            if concept in rows_by_concept
        }
        if not selected:
            raise ValueError(f"No se encontraron conceptos GEIH en {path}.")
        return pd.DataFrame(selected).sort_index()
    finally:
        workbook.close()


def _ipp_month(value: object) -> pd.Timestamp | None:
    text = str(value or "").strip()
    if "-" not in text:
        return None
    month_text, year_text = text.split("-", 1)
    month = DANE_MONTH_NUMBERS_ES.get(month_text.strip())
    if month is None or not year_text[:2].isdigit():
        return None
    year = 2000 + int(year_text[:2])
    return pd.Timestamp(year=year, month=month, day=1)


def load_ipi_dane(path: Path) -> pd.Series:
    """Lee el índice total mensual del IPI DANE (serie disponible desde 2014)."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["3. Indices total por clase "]
        rows: list[tuple[pd.Timestamp, float]] = []
        for row in worksheet.iter_rows(min_row=10, values_only=True):
            if row[0] != "T_IPI" or row[1] is None or row[2] is None or row[4] is None:
                continue
            rows.append(
                (
                    pd.Timestamp(year=int(row[1]), month=int(row[2]), day=1),
                    float(row[4]),
                )
            )
        if not rows:
            raise ValueError(f"No se encontró el total IPI en {path}.")
        return pd.Series(dict(rows), name="ipi_total_dane").sort_index()
    finally:
        workbook.close()


def load_ipp_dane(path: Path) -> pd.Series:
    """Lee el índice de producción nacional total del IPP DANE."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["1.1"]
        rows: list[tuple[pd.Timestamp, float]] = []
        for column in range(4, worksheet.max_column + 1):
            date = _ipp_month(worksheet.cell(6, column).value)
            value = pd.to_numeric(worksheet.cell(7, column).value, errors="coerce")
            if date is not None and pd.notna(value):
                rows.append((date, float(value)))
        if not rows:
            raise ValueError(f"No se encontró el total IPP en {path}.")
        return pd.Series(dict(rows), name="ipp_produccion_nacional_dane").sort_index()
    finally:
        workbook.close()


def read_fred(path: Path, output_name: str, daily: bool = False) -> pd.Series:
    raw = pd.read_csv(path)
    frame = pd.DataFrame(
        {
            "fecha": pd.to_datetime(raw.iloc[:, 0].astype("string"), errors="coerce"),
            "valor": pd.to_numeric(raw.iloc[:, 1], errors="coerce"),
        }
    ).dropna()
    series = frame.set_index("fecha")["valor"].sort_index()
    if daily:
        series = series.resample("MS").mean()
    else:
        series.index = month_start(series.index)
        series = series.groupby(level=0).mean()
    series.name = output_name
    return series


def load_global_base() -> pd.DataFrame:
    """Carga las series globales activas sin imputar faltantes estructurales."""
    if not GLOBAL_BASE_FILE.exists():
        raise FileNotFoundError(
            f"No existe la base global mensual: {GLOBAL_BASE_FILE}. "
            "Ejecute python src/download_global_data.py."
        )

    raw = pd.read_csv(GLOBAL_BASE_FILE, parse_dates=["fecha"]).set_index("fecha")
    raw.index = month_start(raw.index)
    raw = raw.groupby(level=0).mean().sort_index()
    source_columns = [
        "yield_real_10y_tips_pct",
        "yield_real_5y_us_pct",
        "yield_2y_us_pct",
        "yield_10y_us_pct",
        "spread_10y_2y_us_pct",
        "breakeven_5y_us_pct",
        "breakeven_10y_us_pct",
        "brent_usd_barril",
        "commodities_index_imf",
        "epu_global",
        "estres_financiero_stl",
        "nfci_chicago",
        "anfci_chicago",
        "empleo_manufactura_us_miles",
        "produccion_industrial_us",
        "desempleo_us_pct",
        "fletes_transporte_us",
    ]
    optional_columns = [
        "high_yield_oas_pct",
        "ted_spread_pct",
        "desempleo_us_bls_pct",
        "precios_importacion_china",
        "produccion_industrial_china",
        "indicador_lider_china",
        "ipc_china",
    ]
    missing = [column for column in source_columns if column not in raw.columns]
    if missing:
        raise ValueError(f"Faltan series globales activas requeridas: {missing}")

    available_optional = [column for column in optional_columns if column in raw.columns]
    global_data = raw[source_columns + available_optional].copy()
    global_data["ln_brent_global"] = np.log(
        global_data.pop("brent_usd_barril").where(lambda value: value > 0)
    )
    global_data["ln_commodities_global"] = np.log(
        global_data.pop("commodities_index_imf").where(lambda value: value > 0)
    )
    global_data["ln_empleo_manufactura_us"] = np.log(
        global_data.pop("empleo_manufactura_us_miles").where(lambda value: value > 0)
    )
    global_data["ln_produccion_industrial_us"] = np.log(
        global_data.pop("produccion_industrial_us").where(lambda value: value > 0)
    )
    global_data["ln_fletes_transporte_us"] = np.log(
        global_data.pop("fletes_transporte_us").where(lambda value: value > 0)
    )
    if "precios_importacion_china" in global_data:
        global_data["ln_precios_importacion_china"] = np.log(
            global_data["precios_importacion_china"].where(lambda value: value > 0)
        )

    sample = global_data.loc[SAMPLE_START:SAMPLE_END, GLOBAL_RAW_COMPONENTS]
    if sample.isna().any().any():
        missing_sample = {
            column: sample.index[sample[column].isna()].strftime("%Y-%m").tolist()
            for column in sample.columns
            if sample[column].isna().any()
        }
        raise ValueError(
            "Las series globales activas no cubren la muestra completa; "
            f"no se imputan faltantes: {missing_sample}"
        )
    return global_data


def load_fed_gsw_breakeven_daily(path: Path) -> pd.Series:
    """Lee BKEVEN05 diario del archivo Gürkaynak-Sack-Wright."""
    raw = pd.read_csv(
        path,
        skiprows=18,
        usecols=["Date", "BKEVEN05"],
        na_values=["NA"],
    )
    frame = pd.DataFrame(
        {
            "fecha": pd.to_datetime(raw["Date"], errors="coerce"),
            "bei_eeuu_5y_pct": pd.to_numeric(raw["BKEVEN05"], errors="coerce"),
        }
    ).dropna()
    series = frame.set_index("fecha")["bei_eeuu_5y_pct"].sort_index()
    series.index = series.index.normalize()
    series = series.groupby(level=0).mean()
    series.name = "bei_eeuu_5y_pct"
    return series


def load_fed_gsw_breakeven(path: Path) -> pd.Series:
    daily = load_fed_gsw_breakeven_daily(path)
    monthly = daily.resample("MS").mean()
    monthly.name = "bei_eeuu_5y_pct"
    return monthly


def load_banrep_series(path: Path, output_name: str, daily: bool = False) -> pd.Series:
    """Lee el JSON publico del graficador de BanRep y lo lleva a frecuencia mensual."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    item = payload[0] if isinstance(payload, list) else payload
    data = pd.DataFrame(item["data"], columns=["timestamp_ms", output_name])
    dates = pd.to_datetime(data.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    data.index = pd.DatetimeIndex(dates)
    series = pd.to_numeric(data[output_name], errors="coerce").dropna().sort_index()
    if daily:
        # Se promedia cada mercado por separado; no se cruzan calendarios diarios.
        series = series.resample("MS").mean()
    else:
        series.index = month_start(series.index)
        series = series.groupby(level=0).mean()
    series.name = output_name
    return series


def load_banrep_observations(path: Path, output_name: str) -> pd.Series:
    """Lee observaciones BanRep sin agregarlas para comparar calendarios diarios."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    item = payload[0] if isinstance(payload, list) else payload
    frame = pd.DataFrame(item["data"], columns=["timestamp_ms", output_name])
    dates = pd.to_datetime(frame.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    frame.index = pd.DatetimeIndex(dates).normalize()
    series = pd.to_numeric(frame[output_name], errors="coerce").dropna().sort_index()
    series = series.groupby(level=0).mean()
    series.name = output_name
    return series


def build_bei_aggregations() -> pd.DataFrame:
    """Construye BEI mensual con medias separadas y con fechas diarias comunes."""
    nominal = load_banrep_observations(
        RAW / "tes_5y_pesos_banrep.json", "tes_5y_pesos_colombia_pct"
    )
    real = load_banrep_observations(
        RAW / "tes_5y_uvr_banrep.json", "tes_5y_uvr_colombia_pct"
    )
    us = load_fed_gsw_breakeven_daily(RAW / "bei_5y_eeuu_diario_fed.csv")

    separate = pd.concat(
        [
            nominal.resample("MS").mean(),
            real.resample("MS").mean(),
            us.resample("MS").mean(),
        ],
        axis=1,
        sort=True,
    )
    separate["bei_colombia_5y_pct"] = (
        separate["tes_5y_pesos_colombia_pct"]
        - separate["tes_5y_uvr_colombia_pct"]
    )
    separate["diferencial_bei_5y_pp"] = (
        separate["bei_colombia_5y_pct"] - separate["bei_eeuu_5y_pct"]
    )

    common_daily = pd.concat([nominal, real, us], axis=1, join="inner").dropna()
    common_daily["diferencial_bei_5y_comun_pp"] = (
        common_daily["tes_5y_pesos_colombia_pct"]
        - common_daily["tes_5y_uvr_colombia_pct"]
        - common_daily["bei_eeuu_5y_pct"]
    )
    common = common_daily.resample("MS").agg(
        tes_5y_pesos_comun_pct=("tes_5y_pesos_colombia_pct", "mean"),
        tes_5y_uvr_comun_pct=("tes_5y_uvr_colombia_pct", "mean"),
        bei_eeuu_5y_comun_pct=("bei_eeuu_5y_pct", "mean"),
        diferencial_bei_5y_comun_pp=("diferencial_bei_5y_comun_pp", "mean"),
        dias_comunes=("diferencial_bei_5y_comun_pp", "count"),
    )
    counts = pd.concat(
        [
            nominal.resample("MS").count().rename("dias_tes_pesos"),
            real.resample("MS").count().rename("dias_tes_uvr"),
            us.resample("MS").count().rename("dias_bei_eeuu"),
        ],
        axis=1,
        sort=True,
    )
    out = separate.join(common, how="outer").join(counts, how="outer")
    out["diferencia_comun_menos_separada_pp"] = (
        out["diferencial_bei_5y_comun_pp"] - out["diferencial_bei_5y_pp"]
    )
    out.index.name = "fecha"
    return out


def load_remittances(path: Path) -> pd.Series:
    payload = json.loads(path.read_text(encoding="utf-8"))[0]
    data = pd.DataFrame(payload["data"], columns=["timestamp_ms", "remesas_usd_millones"])
    dates = pd.to_datetime(data.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    data.index = month_start(dates)
    return data["remesas_usd_millones"].astype(float).sort_index()


def load_banrep_daily(path: Path, output_name: str) -> pd.Series:
    return load_banrep_series(path, output_name, daily=True)


def load_terms_of_trade(path: Path) -> pd.Series:
    payload = json.loads(path.read_text(encoding="utf-8"))
    series_payload = next(item for item in payload if item.get("id") == 15360)
    data = pd.DataFrame(series_payload["data"], columns=["timestamp_ms", "terminos_intercambio"])
    dates = pd.to_datetime(data.pop("timestamp_ms"), unit="ms", utc=True).dt.tz_convert(None)
    data.index = month_start(dates)
    return data["terminos_intercambio"].astype(float).sort_index()


def load_embig_bcrp(path: Path) -> pd.Series:
    """Lee EMBIG Colombia del JSON público del BCRP y promedia por mes."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[tuple[pd.Timestamp, float]] = []
    for observation in payload.get("periods", []):
        parts = str(observation.get("name", "")).strip().split(".")
        values = observation.get("values") or []
        if len(parts) != 3 or not values or parts[1] not in MONTH_NUMBERS_ES:
            continue
        year = int(parts[2])
        year += 2000 if year < 70 else 1900
        date = pd.Timestamp(year=year, month=MONTH_NUMBERS_ES[parts[1]], day=int(parts[0]))
        value = pd.to_numeric(str(values[0]).replace(",", "."), errors="coerce")
        if pd.notna(value):
            rows.append((date, float(value)))
    if not rows:
        raise ValueError(f"No se encontraron observaciones EMBIG válidas en {path}.")
    daily = pd.Series(
        (value for _, value in rows),
        index=pd.DatetimeIndex(date for date, _ in rows),
        name="embig_colombia_pb",
    ).sort_index()
    daily = daily.groupby(level=0).mean()
    monthly = daily.resample("MS").mean()
    monthly.name = "embig_colombia_pb"
    return monthly


def load_bcrp_monthly(path: Path, output_name: str) -> pd.Series:
    """Lee una serie mensual de BCRPData con periodos como Ene.2006."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[tuple[pd.Timestamp, float]] = []
    for observation in payload.get("periods", []):
        parts = str(observation.get("name", "")).strip().split(".")
        values = observation.get("values") or []
        if len(parts) != 2 or not values or parts[0] not in MONTH_NUMBERS_ES:
            continue
        value = pd.to_numeric(str(values[0]).replace(",", "."), errors="coerce")
        if pd.notna(value):
            rows.append(
                (
                    pd.Timestamp(year=int(parts[1]), month=MONTH_NUMBERS_ES[parts[0]], day=1),
                    float(value),
                )
            )
    if not rows:
        raise ValueError(f"No se encontraron observaciones mensuales válidas en {path}.")
    series = pd.Series(
        (value for _, value in rows),
        index=pd.DatetimeIndex(date for date, _ in rows),
        name=output_name,
    ).sort_index()
    return series.groupby(level=0).mean()


def _row_values(ws, row_number: int) -> tuple[list[object], list[object]]:
    dates = [cell.value for cell in ws[6]][1:]
    values = [cell.value for cell in ws[row_number]][1:]
    return dates, values


def load_fiscal(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    monthly_amounts = workbook.worksheets[0]
    monthly_pct = workbook.worksheets[3]

    dates, balance_values = _row_values(monthly_amounts, 31)
    _, income_values = _row_values(monthly_amounts, 8)
    pct_dates, income_pct_values = _row_values(monthly_pct, 8)

    fiscal = pd.DataFrame(
        {
            "fecha": pd.to_datetime(dates, errors="coerce"),
            "balance_fiscal_miles_millones_cop": pd.to_numeric(balance_values, errors="coerce"),
            "ingresos_totales_miles_millones_cop": pd.to_numeric(income_values, errors="coerce"),
        }
    ).dropna(subset=["fecha"])
    fiscal["fecha"] = month_start(fiscal["fecha"])
    fiscal = fiscal.set_index("fecha").sort_index()

    fiscal_pct = pd.DataFrame(
        {
            "fecha": pd.to_datetime(pct_dates, errors="coerce"),
            "ingresos_totales_pct_pib": pd.to_numeric(income_pct_values, errors="coerce"),
        }
    ).dropna(subset=["fecha"])
    fiscal_pct["fecha"] = month_start(fiscal_pct["fecha"])
    fiscal_pct = fiscal_pct.set_index("fecha").sort_index()

    fiscal = fiscal.join(fiscal_pct, how="left")
    valid = fiscal["ingresos_totales_pct_pib"].abs() > 1e-9
    fiscal.loc[valid, "pib_anual_miles_millones_cop_observado"] = (
        100.0
        * fiscal.loc[valid, "ingresos_totales_miles_millones_cop"]
        / fiscal.loc[valid, "ingresos_totales_pct_pib"]
    )
    year_gdp = fiscal.groupby(fiscal.index.year)["pib_anual_miles_millones_cop_observado"].median()
    fiscal["pib_anual_miles_millones_cop"] = fiscal.index.year.map(year_gdp)
    fiscal["balance_fiscal_12m_miles_millones_cop"] = fiscal[
        "balance_fiscal_miles_millones_cop"
    ].rolling(12, min_periods=12).sum()
    fiscal["deficit_fiscal_12m_pct_pib"] = (
        -100.0
        * fiscal["balance_fiscal_12m_miles_millones_cop"]
        / fiscal["pib_anual_miles_millones_cop"]
    )
    return fiscal


def build_dataset() -> pd.DataFrame:
    series = [
        load_banrep_daily(RAW / "trm_diaria_banrep.json", "trm_cop_usd"),
        load_banrep_daily(
            RAW / "tasa_politica_diaria_banrep.json", "tasa_politica_colombia_pct"
        ),
        read_fred(RAW / "fed_funds_mensual_fred.csv", "fed_funds_eeuu_pct"),
        read_fred(RAW / "dolar_amplio_diario_fred.csv", "indice_dolar_amplio", daily=True),
        read_fred(RAW / "vix_diario_fred.csv", "vix", daily=True),
        load_remittances(RAW / "remesas_mensuales_banrep.json"),
        load_terms_of_trade(RAW / "series_15360_15368.json"),
        load_banrep_series(
            RAW / "reservas_netas_sin_flar_banrep.json",
            "reservas_netas_sin_flar_usd_millones",
        ),
        build_bei_aggregations(),
        load_embig_bcrp(RAW / "embig_colombia_diario_bcrp.json"),
        load_banrep_series(
            RAW / "balanza_comercial_cambiaria_banrep.json",
            "balanza_comercial_cambiaria_usd_millones",
        ),
        load_banrep_series(
            RAW / "flujos_capital_totales_banrep.json",
            "flujos_capital_usd_millones",
        ),
        read_fred(RAW / "brl_usd_mensual_fred.csv", "brl_por_usd"),
        read_fred(RAW / "clp_usd_mensual_fred.csv", "clp_por_usd"),
        read_fred(RAW / "mxn_usd_mensual_fred.csv", "mxn_por_usd"),
        load_bcrp_monthly(RAW / "pen_usd_mensual_bcrp.json", "pen_por_usd"),
        load_banrep_series(
            RAW / "ipc_colombia_banrep.json", "ipc_colombia_indice", daily=False
        ),
        load_ise_dane(RAW / "ise_dane_12actividades_jun2026.xlsx"),
        load_geih_dane(RAW / "geih_dane_jun2026.xlsx"),
        load_geih_dane(
            RAW / "geih_dane_desestacionalizado_jun2026.xlsx",
            seasonally_adjusted=True,
        ),
        load_ipi_dane(RAW / "ipi_dane_jun2026.xlsx"),
        load_ipp_dane(RAW / "ipp_dane_jul2026.xlsx"),
    ]
    data = pd.concat(series, axis=1, sort=True).sort_index()
    data = data.join(load_fiscal(RAW / "balance_fiscal_gnc_mensual_trimestral.xlsx"), how="outer")
    data = data.join(load_global_base(), how="outer")

    data["remesas_12m_usd_millones"] = data["remesas_usd_millones"].rolling(
        12, min_periods=12
    ).sum()
    data["diferencial_tasas_pp"] = (
        data["tasa_politica_colombia_pct"] - data["fed_funds_eeuu_pct"]
    )
    data["embig_colombia_pp"] = data["embig_colombia_pb"] / 100.0
    data["asinh_balanza_comercial"] = np.arcsinh(
        data["balanza_comercial_cambiaria_usd_millones"] / 1000.0
    )
    data["asinh_flujos_capital"] = np.arcsinh(
        data["flujos_capital_usd_millones"] / 1000.0
    )

    regional_returns = pd.DataFrame(
        {
            currency: np.log(data[currency].where(data[currency] > 0)).diff()
            for currency in ["brl_por_usd", "clp_por_usd", "mxn_por_usd", "pen_por_usd"]
        }
    )
    calibration = regional_returns.loc["2006-01-01":"2019-12-01"]
    regional_z = (regional_returns - calibration.mean()) / calibration.std(ddof=0)
    data["factor_monedas_regionales_3"] = regional_z[
        ["brl_por_usd", "clp_por_usd", "mxn_por_usd"]
    ].mean(axis=1, skipna=False)
    data["factor_monedas_regionales_4"] = regional_z[
        ["brl_por_usd", "clp_por_usd", "mxn_por_usd", "pen_por_usd"]
    ].mean(axis=1, skipna=False)
    # Alias explícito del marco macroeconómico integral activo: cuatro monedas.
    data["factor_monedas_regionales"] = data["factor_monedas_regionales_4"]

    positive_logs = {
        "ln_trm": "trm_cop_usd",
        "ln_remesas_12m": "remesas_12m_usd_millones",
        "ln_dolar_amplio": "indice_dolar_amplio",
        "ln_vix": "vix",
        "ln_terminos_intercambio": "terminos_intercambio",
        "ln_reservas_netas_sin_flar": "reservas_netas_sin_flar_usd_millones",
        "ln_ise_total_dane": "ise_total_dane",
        "ln_ipc_colombia": "ipc_colombia_indice",
    }
    for target, source in positive_logs.items():
        data[target] = np.log(data[source].where(data[source] > 0))
    data["dln_vix"] = data["ln_vix"].diff()
    data["dummy_pandemia_2020"] = (
        (data.index >= pd.Timestamp("2020-03-01"))
        & (data.index <= pd.Timestamp("2020-05-01"))
    ).astype(int)
    data.index.name = "fecha"
    return data
